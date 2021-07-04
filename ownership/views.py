import os
from django.shortcuts import render,HttpResponseRedirect,HttpResponse, redirect
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
import datetime
from datetime import date, timedelta
from django.urls import reverse_lazy
from .models import (
   	Ownership,
   	Billing
)
from masterlist.models import EmployeeMasterlist
from vehicle_masterlist.models import VehicleMasterList
from . forms import (
    ownershipForm,
    billing_form,
    bidderForm
)
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                				)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

class ownershipListView(ListView):
	model = Ownership
	template_name = 'transfer_list.html'
	

def ownershipcreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    e_list = EmployeeMasterlist.objects.all()
    v_list = VehicleMasterList.objects.all()
    return render(request, 'transfer_new.html',{'Title':'Ownership - Transfer of Ownership', 'e_list':e_list,'v_list':v_list})


def ownership_submit(request):
	if request.method == 'POST':
		date_application = request.POST.get('date_application')
		req_employee_id = request.POST.get('req_employee_id')
		req_Lname = request.POST.get('req_Lname')
		req_Fname = request.POST.get('req_Fname')
		req_band = request.POST.get('req_band')
		req_cost = request.POST.get('req_cost')
		req_title = request.POST.get('req_title')
		plate_no = request.POST.get('plate_no')
		cond_sticker = request.POST.get('cond_sticker')
		vehicle_model = request.POST.get('vehicle_model')
		vehicle_brand = request.POST.get('vehicle_brand')
		vehicle_make = request.POST.get('vehicle_make')
		Vendor = request.POST.get('Vendor')
		Other_Vendor_Name = request.POST.get('Other_Vendor_Name')
		v_employee_id = request.POST.get('v_employee_id')
		v_fname = request.POST.get('v_fname')
		v_lname = request.POST.get('v_lname')
		v_band = request.POST.get('v_band')
		TOO_Purpose = request.POST.get('TOO_Purpose')
		transfer_fee = request.POST.get('transfer_fee')
		doc_date_completed = request.POST.get('doc_date_completed')
		deedofsale_date = request.POST.get('deedofsale_date')
		confirmation_status = request.POST.get('confirmation_status')
		emailed_to_casher = request.POST.get('emailed_to_casher')
		received_from_casher = request.POST.get('received_from_casher')
		deed_signed = request.POST.get('deed_signed')
		routed_to_jd = request.POST.get('routed_to_jd')
		approved_by_jd = request.POST.get('approved_by_jd')
		return_fleet_admin = request.POST.get('return_fleet_admin')
		forwarded_to_liason = request.POST.get('forwarded_to_liason')
		date_notarized = request.POST.get('date_notarized')
		endorosed_to_insurance = request.POST.get('endorosed_to_insurance')
		requested_for_pullout = request.POST.get('requested_for_pullout')
		# date_pulled = request.POST.get('date_pulled')
		# return_endorsementfleet = request.POST.get('return_endorsementfleet')
		forwarded_fleet_liason = request.POST.get('forwarded_fleet_liason')
		tmg_date_in = request.POST.get('tmg_date_in')
		tmg_location = request.POST.get('tmg_location')
		tmg_date_return = request.POST.get('tmg_date_return')
		lto_location = request.POST.get('lto_location')
		lto_date_in = request.POST.get('lto_date_in')
		lto_date_out = request.POST.get('lto_date_out')
		# lto_date_return = request.POST.get('lto_date_return')
		# date_docs_return = request.POST.get('date_docs_return')
		date_transfered_completed = request.POST.get('date_transfered_completed')
		date_comletion_vismin = request.POST.get('date_comletion_vismin')
		date_received_by = request.POST.get('received_by')
		TOO_SLA = request.POST.get('TOO_SLA')
		status = request.POST.get('status')
		
		saveto_own = Ownership(date_application = date_application,req_employee_id = req_employee_id,req_Fname = req_Fname,req_Lname = req_Lname,
			req_band = req_band,req_cost = req_cost,req_title = req_title,plate_no= plate_no,cond_sticker = cond_sticker,vehicle_model = vehicle_model,
			vehicle_brand = vehicle_brand,vehicle_make = vehicle_make,vendor = Vendor,vendor_name = Other_Vendor_Name,v_employee_id =  v_employee_id,
			v_fname = v_fname,v_lname = v_lname,v_band = v_band,purpose = TOO_Purpose,transfer_fee = transfer_fee,doc_date_completed = doc_date_completed,
			deedofsale_date = deedofsale_date,confirmation_status = confirmation_status,emailed_to_casher = emailed_to_casher,received_from_casher = received_from_casher,
			deed_signed = deed_signed,routed_to_jd =routed_to_jd ,approved_by_jd =approved_by_jd ,return_fleet_admin = return_fleet_admin,forwarded_to_liason = forwarded_to_liason,
			date_notarized = date_notarized,endorosed_to_insurance =endorosed_to_insurance ,requested_for_pullout = requested_for_pullout,
			forwarded_fleet_liason = forwarded_fleet_liason,tmg_date_in =tmg_date_in ,tmg_location = tmg_location,
			tmg_date_return = tmg_date_return,lto_date_in = lto_date_in,lto_date_out = lto_date_out, lto_location = lto_location,
			date_transfered_completed = date_transfered_completed,date_comletion_vismin = date_comletion_vismin, date_received_by = date_received_by, TOO_SLA=TOO_SLA, status=status)
		saveto_own.save()

		return HttpResponseRedirect('/Ownership/Ownership/')
def too_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = Ownership.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = Ownership.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = Ownership.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = Ownership.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = Ownership.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = Ownership.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'TOOdeadline.html',{'title':'TOO - TOO Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})


class ownershipUpdate(SuccessMessageMixin, UpdateView):
	model = Ownership
	form_class = ownershipForm
	template_name = 'transfer_form.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Transfer of Ownership Updated Successfully!"
		
class ownershipDetails(DetailView):
	model = Ownership
	template_name = 'transfer_details.html'
	
class ownershipDeleteView(BSModalDeleteView):
    model = Ownership
    template_name = 'transfer_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('ownership_list')

def ownershipHistoryView(request):
    if request.method == "GET":
       obj = Ownership.history.all()

       return render(request, 'transfer_history.html', context={'object': obj})


class billing(SuccessMessageMixin, CreateView):
	model = Billing
	form_class = billing_form
	template_name = 'billing/billing_form.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Billing Created Successfully!"


class billing_list(ListView):
	model = Billing
	template_name = 'billing/billing_list.html'


class billingDetails(DetailView):
	model = Billing
	template_name = 'billing/billing_details.html'


class billingDeleteView(BSModalDeleteView):
    model = Billing
    template_name = 'billing/billing_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('billing_list')


class billingUpdate(SuccessMessageMixin, UpdateView):
	model = Billing
	form_class = billing_form
	template_name = 'billing/billing_form.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Transfer of Ownership Updated Successfully!"

def billingHistoryView(request):
    if request.method == "GET":
       obj = Billing.history.all()

       return render(request, 'billing/billing_history.html', context={'object': obj})

class bidder(SuccessMessageMixin, CreateView):
	model = Ownership
	form_class = bidderForm
	template_name = 'bidder_form.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Transfer of Ownership Bidder Created Successfully!"

def ownership_excel(request):
    own_queryset = Ownership.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Transfer of Ownership.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Transfer of Ownership'

    columns = [
			'Date Application Received', 
		    'Requisitioner Employee ID', 
		    'Requisitioner First Name',
		    'Requisitioner Last Name',
		    'Requisitioner Band',
		    'Requisitioner Cost Center',
		    'Requisitioner Title',
		    'Plate Number',
		    'Conduction Sticker ',
		    'Vehicle Model' ,
		    'Vehicle Brand' ,
		    'Vehicle Make' ,
		    'Vendor' ,
		    'Other Vendor Name' ,
		    'Vendee Employee ID',
		    'Vendee First Name',
		    'Vendee Last Name',
		    'Vendee Band',
		    'TOO Purpose' ,
		    'Transfer Fee' ,
		    'Document Completed Date' ,
		    'Date Created Deed of Sale' ,
		    'Confirmation Status' ,
		    'Date OR Emailed to Casher' ,
		    'Date OR Received from Casher' ,
		    'Signed Deed of Sale Completed Date' ,
		    'Date Routed to JD/ECS' ,
		    'Date Approved By JD/ECS' ,
		    'Date Return to Fleet Admin Assistant' ,
		    'Date Forwarded to Liason Officer' ,
		    'Date Notarized' ,
		    'Date Endorsed to Insurance Company' ,
		    'Date Received Pull-out of OR/CR' ,
		    'Forwarded Fleet liason',
		    'TMG Schedule' ,
		    'TMG Location' ,
		    'TMG Date Out',
		    'LTO Location' ,
		    'LTO Date In' ,
		    'LTO Date Out' ,
		    'Date Transfer Completed' ,
		    'Date Transfer Completed for Visayas/Mindanao' ,
		    'SLA' ,
		    'Date Initiated' ,
		    'Date Received By' ,
		    'Status',
		    'Deadline' ,
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for own in own_queryset:
        row_num += 1
        row = [
				own.date_application,
				own.req_employee_id,
				own.req_Fname ,
				own.req_Lname ,
				own.req_band ,
				own.req_cost ,
				own.req_title ,
				own.plate_no ,
				own.cond_sticker ,
				own.vehicle_model ,
				own.vehicle_brand ,
				own.vehicle_make ,
				own.vendor ,
				own.vendor_name ,
				own.v_employee_id ,
				own.v_fname ,
				own.v_lname ,
				own.v_band ,
				own.purpose ,
				own.transfer_fee ,
				own.doc_date_completed ,
				own.deedofsale_date ,
				own.confirmation_status ,
				own.emailed_to_casher ,
				own.received_from_casher ,
				own.deed_signed ,
				own.routed_to_jd ,
				own.approved_by_jd ,
				own.return_fleet_admin ,
				own.forwarded_to_liason ,
				own.date_notarized ,
				own.endorosed_to_insurance ,
				own.requested_for_pullout ,
				own.forwarded_fleet_liason,
				own.tmg_date_in ,
				own.tmg_location ,
				own.tmg_date_return ,
				own.lto_location ,
				own.lto_date_in ,
				own.lto_date_out ,
				own.date_transfered_completed ,
				own.date_comletion_vismin ,
				own.TOO_SLA ,
				own.date_initiated ,
				own.date_received_by ,
				own.status,
				own.Deadline ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def billing_excel(request):
    own_queryset = Billing.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Billing.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Billing'

    columns = [
			'Reference No',
            'SOA Number',
            'In Payment Of',
            'Cost Center',
            'Date Of Bill',
            'Total Amount'
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for bill in own_queryset:
        row_num += 1
        row = [
			bill.ref_no ,
    		bill.in_payment_of ,
    		bill.soa_no ,
    		bill.cost_center ,
    		bill.date_bill ,
    		bill.total_amount 
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

##Daily report Details
def ownership_report_details(request):
	date = datetime.datetime.today()
	notorized = Ownership.objects.filter(status = 'NOTARIZED').count()
	routing_count = Ownership.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
	tmg_schedule = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
	tmg_appearance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
	lto_transfer = Ownership.objects.filter(status = 'LTO TRANSFER').count()
	fleet_vismin = Ownership.objects.filter(status = 'FLEET VISMIN').count()
	with_tmg_etching = Ownership.objects.filter(status = 'WITH_MACRO ETCHING').count()
	total = notorized + routing_count + tmg_schedule + tmg_appearance + lto_transfer + fleet_vismin + with_tmg_etching
	return render(request, 'report_details.html',{'Title':'TOO - TOO Report', 'notorized':notorized,'routing_count':routing_count,
	'tmg_schedule':tmg_schedule, 'tmg_appearance':tmg_appearance, 'lto_transfer':lto_transfer, 'fleet_vismin':fleet_vismin,
	'with_tmg_etching':with_tmg_etching, 'total':total,'date':date})

# Registration Daily Report
def ownership_report(request):
	username = os.getlogin()
	date = datetime.datetime.today()

	notorized = Ownership.objects.filter(status = 'NOTARIZED').count()
	routing_count = Ownership.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
	tmg_schedule = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
	tmg_appearance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
	lto_transfer = Ownership.objects.filter(status = 'LTO TRANSFER').count()
	fleet_vismin = Ownership.objects.filter(status = 'FLEET VISMIN').count()
	with_tmg_etching = Ownership.objects.filter(status = 'WITH_MACRO ETCHING').count()
	total = notorized + routing_count + tmg_schedule + tmg_appearance + lto_transfer + fleet_vismin + with_tmg_etching
	from openpyxl import Workbook, load_workbook
	output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	file_name = "TOO_Report.xlsx"
	output['Content-Disposition'] = 'attachment; filename='+ file_name
	wb = Workbook()
	ws = wb.active

	#header
	ws.title = "TOO Report"
	ws['A1'].value = "SUBJECT:"
	ws['B1'].value = "TOO"
	ws['A3'].value = "PERSONNEL:"
	ws['B3'].value = "Jessie"
	ws['A4'].value = "Date"
	ws['A5'].value = ""
	ws.append(['Summary', 'Total','Remarks'])
	ws['A8'].value = "DOAS Creation (New)"
	ws['A9'].value = "DOAS Receive"
	ws['A10'].value = "For Routing"
	ws['A11'].value = "Notarized"
	ws['A12'].value = ""
	ws['A13'].value = "For TMG"
	ws['A14'].value = "For MACRO Etching"
	ws['A15'].value = "With TMG Schedule"
	ws['A16'].value = ""
	ws['A17'].value = "With MACRO Etching"
	ws['A18'].value = ""
	ws['A19'].value = "Fleet VisMin"
	ws['A20'].value = ""
	ws['A21'].value = "LTO Transfer"
	ws['A22'].value = ""
	ws['A23'].value = "Total"
	ws['A24'].value = ""
	ws['A25'].value = "Transfered"
	#data
	ws['B4'].value = date
	ws['B10'].value = routing_count
	ws['B11'].value = notorized
	ws['B14'].value = tmg_appearance
	ws['B15'].value = tmg_schedule
	ws['B17'].value = with_tmg_etching
	ws['B19'].value = fleet_vismin
	ws['B21'].value = lto_transfer
	ws['B23'].value = total

	# style
	ws['A6'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['B6'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['C6'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['D6'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['E6'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['F6'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['G6'].fill = PatternFill("solid", fgColor="00FFCC99")
	#TMG color
	ws['A13'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['B13'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['C13'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['D13'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['E13'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['F13'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['G13'].fill = PatternFill("solid", fgColor="00FFCC99")
	#vismin color
	ws['A19'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['B19'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['C19'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['D19'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['E19'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['F19'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['G19'].fill = PatternFill("solid", fgColor="00FFCC99")
	#lto transfer color
	ws['A21'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['B21'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['C21'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['D21'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['E21'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['F21'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws['G21'].fill = PatternFill("solid", fgColor="00FFCC99")
	#transfer color
	ws['A25'].fill = PatternFill("solid", fgColor="00FF99CC")
	ws['B25'].fill = PatternFill("solid", fgColor="00FF99CC")
	ws['C25'].fill = PatternFill("solid", fgColor="00FF99CC")
	ws['D25'].fill = PatternFill("solid", fgColor="00FF99CC")
	ws['E25'].fill = PatternFill("solid", fgColor="00FF99CC")
	ws['F25'].fill = PatternFill("solid", fgColor="00FF99CC")
	ws['G25'].fill = PatternFill("solid", fgColor="00FF99CC")

	wb.save(output)
	return output
	# wb.save(f"/Users/{username}/Desktop/TOO_Report.xlsx")
	# return redirect('/Ownership/Ownership')


