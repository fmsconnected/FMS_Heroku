from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
# Create your views here.
from django.views import generic
from django.shortcuts import render,HttpResponseRedirect, get_list_or_404,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
import datetime
from datetime import date, timedelta
from .models import (
		Corrective,
)
from request.models import Vehicle_Repair
from masterlist.models import (
    EmployeeMasterlist
    )
from vehicle_masterlist.models import VehicleMasterList
from leasingmasterlist.models import Leasing
from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
# from django.views.generic.edit import (
    # DeleteView,
# )
from . forms import (
    correctiveform,
    )

from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

class correctiveListView(ListView):
    model = Corrective
    template_name='corrective_list.html'

def correctiveOngoing(request):
    if request.method == "GET":
       obj = Corrective.objects.filter(status="Ongoing")

       return render(request, 'corrective_ongoing.html', context={'object': obj})
def correctiveCompleted(request):
    if request.method == "GET":
       obj = Corrective.objects.filter(status="Completed")

       return render(request, 'corrective_completed.html', context={'object': obj})

def correctiveCreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'corrective_form.html',{'Title':'Corrective - Corrective Maintenance','emplist':emplist,'vlist':vlist})

def correctivesubmit(request):
    if request.method == 'POST':
        request_date = request.POST.get('request_date')
        emp_id = request.POST.get('emp_id')
        cost_center = request.POST.get('cost_center')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        c_no = request.POST.get('c_no')
        company = request.POST.get('company')
        department = request.POST.get('department')
        group = request.POST.get('group')
        plate_no = request.POST.get('plate_no')
        v_brand = request.POST.get('v_brand')
        engine = request.POST.get('engine')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        chassis = request.POST.get('chassis')
        v_band = request.POST.get('v_band')
        cs_no = request.POST.get('cs_no')
        eq_no = request.POST.get('eq_no')
        fleet_area = request.POST.get('fleet_area')
        particulars = request.POST.get('particulars')
        category = request.POST.get('category')
        maintenance_type1 = request.POST.get('maintenance_type1')
        maintenance_type2 = request.POST.get('maintenance_type2')
        scope_work1 = request.POST.get('scope_work1')
        scope_work2 = request.POST.get('scope_work2')
        recomendations = request.POST.get('recomendations')
        service_reminder = request.POST.get('service_reminder')
        repair_verified_by = request.POST.get('repair_verified_by')
        work_order1 = request.POST.get('work_order1')
        work_order2 = request.POST.get('work_order2')
        work_order3 = request.POST.get('work_order3')
        date_work_created = request.POST.get('date_work_created')
        repair_shop = request.POST.get('repair_shop')
        memo_app = request.POST.get('memo_app')
        date_forward = request.POST.get('date_forward')
        estimate_no = request.POST.get('estimate_no')
        maintenance_amount = request.POST.get('maintenance_amount')
        less_discount = request.POST.get('less_discount')
        estimate_remark = request.POST.get('estimate_remark')
        estimate_attach = request.POST.get('estimate_attach')
        approved_by = request.POST.get('approved_by')
        kilo_reading = request.POST.get('kilo_reading')
        status = request.POST.get('status')

        saveto_corretive = Corrective(request_date=request_date, employee=emp_id, cost_center=cost_center, first_name=fname,
            last_name=lname, contact_no=c_no, company=company, department=department, group_section=group,
            plate_no=plate_no, v_brand=v_brand, engine=engine, v_make=v_make, v_model=v_model, chassis=chassis,
            band=v_band, cond_sticker=cs_no, equipment_no=eq_no, fleet_area=fleet_area, particulars=particulars,
            category=category, maintenance_type1=maintenance_type1, scope_work1=scope_work1, maintenance_type2=maintenance_type2,
            scope_work2=scope_work2, recommendations=recomendations, service_reminder=service_reminder, verified_by=repair_verified_by,
            work_order1=work_order1, work_order2=work_order2, work_order3=work_order3, datework_created=date_work_created,
            Shop_vendor=repair_shop, date_forwarded=date_forward, estimate_no=estimate_no, maintenance_amount=maintenance_amount,
            less_discount=less_discount, estimate_remarks=estimate_remark, estimate_attached=estimate_attach, approvedby=approved_by,
            meter_reading=kilo_reading, memo_app=memo_app,status=status
    )
        saveto_corretive.save()

        return HttpResponseRedirect('/Corrective/Corrective/')

class correctiveDetailView(DetailView):
    model = Corrective
    template_name = 'corrective_details.html'

class correctiveUpdateView(SuccessMessageMixin, UpdateView):
    model = Corrective
    form_class = correctiveform
    template_name = 'corrective_update.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Corrective Maintenance Updated Successfully!"

class correctiveDeleteView(BSModalDeleteView):
    model = Corrective
    template_name = 'corrective_delete.html'
    success_message = 'Success: Corrective Data was deleted.'
    success_url = reverse_lazy('corrective_list')

def correctiveHistoryView(request):
    if request.method == "GET":
       obj = Corrective.history.all()

       return render(request, 'corrective_history.html', context={'object': obj})

def corrective_excel(request):
    corrective_queryset = Corrective.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Corrective.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Corrective'

    columns = [
                'Request Date' ,
                'Employee' ,
                'Cost Center' ,
                'First Name' ,
                'Last Name' ,
                'Contact No' ,
                'Company' ,
                'Department' ,
                'Group Section' ,
                'Plate No' ,
                'Brand' ,
                'Engine' ,
                'Make' ,
                'Model' ,
                'Chassis' ,
                'Band' ,
                'Conduction Sticker' ,
                'Equipment No' ,
                'Fleet Area' ,
                'Particulars' ,
                'Category' ,
                'Maintenance Type 1' ,
                'Scope Work 1' ,
                'Maintenance Type 2' ,
                'Scope Work 2' ,
                'Recommendations' ,
                'Service Reminder' ,
                'Verified By' ,
                'Work Order 1' ,
                'Work Order 2' ,
                'Work Order 3' ,
                'Date Work Created' ,
                'Shop Vendor' ,
                'Memo App Number',
                'Date Forwarded' ,
                'Estimate No' ,
                'Maintenance Amount' ,
                'Less Discount' ,
                'Estimate Remarks' ,
                'Estimate Attached' ,
                'Approved By' ,
                'Meter Reading' ,
                'Date Initiated' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for corrective in corrective_queryset:
        row_num += 1
        row = [
                corrective.request_date,
                corrective.employee ,
                corrective.cost_center ,
                corrective.first_name ,
                corrective.last_name ,
                corrective.contact_no ,
                corrective.company ,
                corrective.department ,
                corrective.group_section ,
                corrective.plate_no ,
                corrective.v_brand ,
                corrective.engine ,
                corrective.v_make ,
                corrective.v_model ,
                corrective.chassis ,
                corrective.band ,
                corrective.cond_sticker ,
                corrective.equipment_no ,
                corrective.fleet_area ,
                corrective.particulars ,
                corrective.category ,
                corrective.maintenance_type1 ,
                corrective.scope_work1 ,
                corrective.maintenance_type2 ,
                corrective.scope_work2 ,
                corrective.recommendations ,
                corrective.service_reminder ,
                corrective.verified_by ,
                corrective.work_order1 ,
                corrective.work_order2 ,
                corrective.work_order3 ,
                corrective.datework_created ,
                corrective.Shop_vendor ,
                corrective.memo_app,
                corrective.date_forwarded ,
                corrective.estimate_no ,
                corrective.maintenance_amount ,
                corrective.less_discount ,
                corrective.estimate_remarks ,
                corrective.estimate_attached ,
                corrective.approvedby ,
                corrective.meter_reading ,
                corrective.date_initiated ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


##Daily report Details
def vehicle_maintenance_report_details(request):
    date = datetime.datetime.today()
    # notorized = Ownership.objects.filter(status = 'NOTARIZED').count()
    # routing_count = Ownership.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
    # tmg_schedule = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
    # tmg_appearance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
    
    return render(request, 'report/corrective_report.html',{'Title: Vehicle Maintenance':'Vehicle Maintenance','date':date})

# Registration Daily Report
def vehicle_maintenance_report(request):
    date = datetime.datetime.today()

    # preventive_count = Vehicle_Repair.objects.filter(status = 'NOTARIZED').count()
    # corrective_count = Corrective.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
    # tire_battery = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
    # insurance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
    
    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Vehicle Maitenance.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Vehicle Maintenance"
    ws['A1'].value = "SUBJECT:"
    ws['B1'].value = "Vehicle Maintenance"
    ws['A3'].value = "PERSONNEL:"
    ws['B3'].value = "Shane Santos"
    ws['A4'].value = "Date"
    ws['A5'].value = ""
    ws.append(['Vehicle Maitenance', 'Total','Remarks'])
    ws['A8'].value = "Preventive Maitenance"
    ws['A9'].value = "Corrective Maitenance"
    ws['A10'].value = "Tires and Battery"
    ws['A11'].value = ""
    ws['A12'].value = ""
    ws['A13'].value = "Insurance Claims"
    #data
    # ws['B4'].value = date
    # ws['B10'].value = routing_count
    # ws['B11'].value = notorized
    # ws['B14'].value = tmg_appearance
    # ws['B15'].value = tmg_schedule
    # ws['B17'].value = with_tmg_etching
    # ws['B19'].value = fleet_vismin
    # ws['B21'].value = lto_transfer
    # ws['B23'].value = total

    # style
    ws['A6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['B6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['C6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['D6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['E6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['F6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['G6'].fill = PatternFill("solid", fgColor="00FFCC99")
    #TMG color
    ws['A13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['B13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['C13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['D13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['E13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['F13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['G13'].fill = PatternFill("solid", fgColor="00FFCC99")
   
    wb.save(output)
    return output




