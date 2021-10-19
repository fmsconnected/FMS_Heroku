from django.shortcuts import render,HttpResponseRedirect,reverse,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
from datetime import date, timedelta
from .models import (
    Vehicle_Repair_payment,
)
from masterlist.models import EmployeeMasterlist
from vehicle_masterlist.models import VehicleMasterList

from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from . forms import (
    vrepair_form
)
from car_rental_payment.models import CarRental
from fuel_supplier_payment.models import Fuel_supplier
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )


class vrepair_payment(ListView):
    model = Vehicle_Repair_payment
    template_name = 'payment/vehicle_repair/vehicle_repair_paymentList.html'

def vrepair_payment_create(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'payment/vehicle_repair/vehicle_repair_form.html',{'emplist':emplist,'vlist':vlist})

class vrepairDetailView(DetailView):
    model = Vehicle_Repair_payment
    template_name = 'payment/vehicle_repair/vehicle_repair_details.html'

class vrepairDeleteView(BSModalDeleteView):
    model = Vehicle_Repair_payment
    template_name = 'payment/vehicle_repair/vehicle_repair_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('vehiclerepair_payment')

class vrepairUpdate(SuccessMessageMixin, UpdateView):
    model = Vehicle_Repair_payment
    form_class = vrepair_form
    template_name = 'payment/vehicle_repair/vehicle_repair_update.html'

    def get_success_message(self, cleaned_data):
        print(cleaned_data)
        return "Vehicle Repair Payment Update Successfully!"

def vrepairlHistoryView(request):
    if request.method == "GET":
       obj = Vehicle_Repair_payment.history.all()

       return render(request, 'payment/vehicle_repair/vehicle_repair_history.html', context={'object': obj})

def vrepairsubmit(request):
    if request.method == 'POST':
        request_date = request.POST.get('request_date')
        emp_id = request.POST.get('emp_id')
        cost_center = request.POST.get('cost_center')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        c_no = request.POST.get('c_no')
        company = request.POST.get('company')
        department = request.POST.get('department')
        group = request.POST.get('group')
        plate_no = request.POST.get('plate_no')
        v_brand = request.POST.get('v_brand')
        engine = request.POST.get('engine')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        chassis = request.POST.get('chassis')
        v_band = request.POST.get('v_band')
        cs_no = request.POST.get('cs_no')
        eq_no = request.POST.get('eq_no')
        dealership = request.POST.get('dealer')
        amount = request.POST.get('amount')
        service_type = request.POST.get('service_type')
        rfp_no = request.POST.get('rfp_no')
        invoice_number2 = request.POST.get('invoice_num')
        invoice_date = request.POST.get('invoice_date')
        status = request.POST.get('status')


        saveto_vrp = Vehicle_Repair_payment(request_date=request_date, employee=emp_id, cost_center=cost_center, first_name=fname,
            last_name=lname, contact_no=c_no, company=company, department=department, group_section=group,
            plate_no=plate_no, v_brand=v_brand, engine=engine, v_make=v_make, v_model=v_model, chassis=chassis,
            band=v_band, cond_sticker=cs_no, equipment_no=eq_no, dealership=dealership, amount=amount, service_type=service_type, rfp_no =rfp_no, invoice_number2=invoice_number2,
            invoice_date=invoice_date,Status=status
    )
        saveto_vrp.save()

        return HttpResponseRedirect('/Payment/VehicleRepair/')

def vrepair_excel(request):
    vrepair_queryset = Vehicle_Repair_payment.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Repair Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Repair Payment'

    columns = [
            'Activity Id', 
            'Request Date', 
            'Employee', 
            'Cost Center', 
            'First Name', 
            'Last Name', 
            'Contact No', 
            'Company', 
            'Department', 
            'Group Section', 
            'Plate No', 
            'Brand', 
            'Engine', 
            'Make', 
            'Model', 
            'Chassis', 
            'Band', 
            'Conduction Sticker', 
            'Equipment No', 
            'Dealership', 
            'Amount', 
            'Service Type', 
            'Date Initiated',
            'rfpno',
            'invoice_number2',
            'invoice_date', 
            'Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vrp in vrepair_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
            vrp.Activity_id, 
            vrp.request_date, 
            vrp.employee, 
            vrp.cost_center, 
            vrp.first_name, 
            vrp.last_name, 
            vrp.contact_no, 
            vrp.company, 
            vrp.department, 
            vrp.group_section, 
            vrp.plate_no, 
            vrp.v_brand, 
            vrp.engine, 
            vrp.v_make, 
            vrp.v_model, 
            vrp.chassis, 
            vrp.band, 
            vrp.cond_sticker, 
            vrp.equipment_no, 
            vrp.dealership, 
            vrp.amount, 
            vrp.service_type, 
            vrp.date_initiated,
            vrp.rfp_no,
            vrp.invoice_number2,
            vrp.invoice_date,
            vrp.Status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


## Car rental Daily report Details
def car_report_details(request):
    date = datetime.datetime.today()
    carreport_SAFARI = CarRental.objects.filter(car_provider="SAFARI", sqa_number="").count()
    carreport_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND", sqa_number="").count()
    carreport_TG = CarRental.objects.filter(car_provider="Tiger City", sqa_number="").count()
    carreport_ORIX = CarRental.objects.filter(car_provider="ORIX", sqa_number="").count()
    
    vreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX", invoice_date="").count()
    vreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI", invoice_date="").count()
    vreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND", invoice_date="").count()
    vreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City", invoice_date="").count()
    
    processed_ORIX = CarRental.objects.filter(car_provider="ORIX").exclude(sqa_number='').count()
    processed_SAFARI = CarRental.objects.filter(car_provider="SAFARI").exclude(sqa_number='').count()
    processed_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND").exclude(sqa_number='').count()
    processed_TG = CarRental.objects.filter(car_provider="Tiger City").exclude(sqa_number='').count()

    proreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX").exclude(invoice_date="").count()
    proreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI").exclude(invoice_date="").count()
    proreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND").exclude(invoice_date="").count()
    proreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City").exclude(invoice_date="").count()
    
    shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL", SOA_billdate="").count()
    pro_shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL").exclude(SOA_billdate="").count()
    petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation", SOA_billdate="").count()
    pro_petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation").exclude(SOA_billdate="").count()

    orix = carreport_ORIX + vreport_ORIX
    safari = carreport_SAFARI + vreport_SAFARI
    diamond = carreport_DIAMOND + vreport_DIAMOND
    tg = carreport_TG + vreport_TG

    tp_orix = processed_ORIX + proreport_ORIX
    tp_safari = processed_SAFARI + proreport_SAFARI
    tp_diamond = processed_DIAMOND + proreport_DIAMOND
    tp_tg = processed_TG + proreport_TG

    return render(request, 'payment_report.html',{'title':'Payment','date':date,'orix':orix, 'safari':safari, 'diamond':diamond, 'tg':tg,'tp_orix':tp_orix, 'tp_safari':tp_safari, 'tp_diamond':tp_diamond, 'tp_tg':tp_tg})

# Car rental Daily Report
def car_report(request):
    date = datetime.datetime.today()
    yr = datetime.datetime.now().year
    months = ['zero','January','February','March','April','May','June','July','August','September','October','November','December']
    month = months[date.month]
    datem = datetime.datetime(date.year, date.month, 1)
    carreport_SAFARI = CarRental.objects.filter(car_provider="SAFARI", sqa_number="").count()
    carreport_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND", sqa_number="").count()
    carreport_TG = CarRental.objects.filter(car_provider="Tiger City", sqa_number="").count()
    carreport_ORIX = CarRental.objects.filter(car_provider="ORIX", sqa_number="").count()
    
    vreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX", invoice_date="").count()
    vreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI", invoice_date="").count()
    vreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND", invoice_date="").count()
    vreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City", invoice_date="").count()
    
    processed_ORIX = CarRental.objects.filter(car_provider="ORIX").exclude(sqa_number='').count()
    processed_SAFARI = CarRental.objects.filter(car_provider="SAFARI").exclude(sqa_number='').count()
    processed_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND").exclude(sqa_number='').count()
    processed_TG = CarRental.objects.filter(car_provider="Tiger City").exclude(sqa_number='').count()

    proreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX").exclude(invoice_date="").count()
    proreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI").exclude(invoice_date="").count()
    proreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND").exclude(invoice_date="").count()
    proreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City").exclude(invoice_date="").count()
    
    shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL", SOA_billdate="").count()
    pro_shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL").exclude(SOA_billdate="").count()
    petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation", SOA_billdate="").count()
    pro_petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation").exclude(SOA_billdate="").count()

    orix = carreport_ORIX + vreport_ORIX
    safari = carreport_SAFARI + vreport_SAFARI
    diamond = carreport_DIAMOND + vreport_DIAMOND
    tg = carreport_TG + vreport_TG

    tp_orix = processed_ORIX + proreport_ORIX
    tp_safari = processed_SAFARI + proreport_SAFARI
    tp_diamond = processed_DIAMOND + proreport_DIAMOND
    tp_tg = processed_TG + proreport_TG


    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Insurance Daily Report.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Insurance Daily Report"
    ws['A1'].value = "Personel"
    ws['B1'].value = "Janine Manzo"
    ws['B2'].value = datem
    ws['A2'].value = "Date"
    ws.append(['Vendor','Total Unpaid SOA','Total Processed','For this Month Total SOA Processed'])
    ws['A4'].value = "ORIX"
    ws['A5'].value = "DIAMOND"
    ws['A6'].value = "PETRON"
    ws['A7'].value = "SHELL"
    ws['A8'].value = "GR8"
    ws['A9'].value = "JXM"
    ws['A10'].value = "SAFARI (Car Rental)"
    ws['A11'].value = "Tiger City (Car Rental)"
    ws['A12'].value = "Created Work Order"
    ws['A13'].value = "TopServe (FLEET Driver)"
    ws['A14'].value = "CARPLAN 95%"
    ws['A15'].value = "CARPLAN 5%"
    #data
    ws['B4'].value = orix
    ws['B5'].value = diamond
    ws['B6'].value = petron
    ws['B7'].value = shell
    ws['B8'].value = ""
    ws['B9'].value = ""
    ws['B10'].value = safari
    ws['B11'].value = tg
    ws['B12'].value = ""
    ws['B13'].value = ""
    ws['B14'].value = ""
    ws['B15'].value = ""

    ws['C4'].value = tp_orix
    ws['C5'].value = tp_diamond
    ws['C6'].value = pro_petron
    ws['C7'].value = pro_shell
    ws['C8'].value = ""
    ws['C9'].value = ""
    ws['C10'].value = tp_safari
    ws['C11'].value = tp_tg
    ws['C12'].value = ""
    ws['C13'].value = ""
    ws['C14'].value = ""
    ws['C15'].value = ""

   
    wb.save(output)
    return output
