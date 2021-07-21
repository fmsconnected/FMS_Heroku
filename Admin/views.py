from .models import UserReport
from django.shortcuts import render, HttpResponseRedirect, HttpResponse, reverse
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
from rest_framework import viewsets
from rest_framework.response import Response
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render

from django.views.generic import (
    ListView,
    CreateView,
    DetailView,
    UpdateView,
)

from .models import (
    UserReport
)


# class userListView(ListView):
#     model = UserReport
#     template_name = 'userreport_list.html'

def userListView(request):
    user = UserReport.objects.all().order_by('-id')
    return render(request, 'userreport_list.html', {'title': 'User - User', 'user': user})

def user_report_excel(request):
    report = UserReport.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=User Login Report.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'User Login Report'

    columns = [

        'User Name',
        'Action',
        'Ip',
        'Date and Time',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for report in report:
        row_num += 1
        date = report.date.strftime('%m/%d/%Y')
        row = [
            report.username,
            report.action,
            report.ip,
            report.date,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
