from django.shortcuts import render,HttpResponseRedirect,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from openpyxl import Workbook
from django.urls import reverse_lazy
import datetime
from datetime import date, timedelta
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from .models import (
   	Fata_monitoring,
)

from rest_framework import viewsets
from rest_framework.response import Response
from django.db.models import Q
from . forms import (
    FATAmonitoringForm,
    reg_updateForm
)
from masterlist.models import EmployeeMasterlist
from vehicle_masterlist.models import VehicleMasterList

from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                				)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

from .serializers import (
    monitoringSerializer
    )

class monitoringViewSet(viewsets.ModelViewSet):
    queryset = Fata_monitoring.objects.all().order_by('id')
    serializer_class = monitoringSerializer


class monitoringListView(ListView):
	model = Fata_monitoring
	template_name = 'fata_monitoringlist.html'

def monitoring_create(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    e_list = EmployeeMasterlist.objects.all()
    v_list = VehicleMasterList.objects.all()
    return render(request, 'fata_monitoring_new.html',{'Title':'Monitoring - Fata Monitoring', 'e_list':e_list,'v_list':v_list})

def monitoring_ongoing(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    ongoing = Fata_monitoring.objects.filter(Status="Ongoing")
    return render(request, 'fata_ongoing.html',{'Title':'Monitoring - Fata Monitoring', 'ongoing':ongoing})

def monitoring_completed(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    completed = Fata_monitoring.objects.filter(Status="Completed")
    return render(request, 'fata_completed.html',{'Title':'Monitoring - Fata Monitoring', 'completed':completed})

def monitoring_submit(request):
	if request.method == 'POST':
		Fata_no = request.POST.get('Fata_no')
		Date_transfer = request.POST.get('Date_transfer')
		Date_received = request.POST.get('Date_received')
		Plate_no = request.POST.get('Plate_no')
		v_make = request.POST.get('v_make')
		v_brand = request.POST.get('v_brand')
		Certificate_of_Reg = request.POST.get('Certificate_of_Reg')
		v_model = request.POST.get('v_model')
		Transferor_employee = request.POST.get('Transferor_employee')
		Transferor_Fname = request.POST.get('Transferor_Fname')
		Transferor_Lname = request.POST.get('Transferor_Lname')
		Recipient_Employee = request.POST.get('Recipient_Employee')
		Recipient_Fname = request.POST.get('Recipient_Fname')
		Recipient_Lname = request.POST.get('Recipient_Lname')
		Date_endorsed_Globe = request.POST.get('Date_endorsed_Globe')
		Date_endorsed_Innove = request.POST.get('Date_endorsed_Innove')
		Clearing_accountability = request.POST.get('Clearing_accountability')
		Globe_fixed_asset = request.POST.get('Globe_fixed_asset')
		Innove_fixed_asset = request.POST.get('Innove_fixed_asset')
		status = request.POST.get('status')
		

		saveto_fata = Fata_monitoring(Fata_no=Fata_no ,Date_transfer=Date_transfer ,Date_received=Date_received ,Plate_no=Plate_no ,
			Vehicle_make=v_make ,Vehicle_brand=v_brand ,Certificate_of_Reg=Certificate_of_Reg ,Vehicle_model=v_model ,
			Transferor_employee=Transferor_employee ,Transferor_Fname=Transferor_Fname ,Transferor_Lname=Transferor_Lname ,
			Recipient_Employee=Recipient_Employee ,Recipient_Fname=Recipient_Fname ,Recipient_Lname=Recipient_Lname ,Date_endorsed_Globe=Date_endorsed_Globe ,
			Date_endorsed_Innove=Date_endorsed_Innove ,Clearing_accountability=Clearing_accountability ,Globe_fixed_asset=Globe_fixed_asset ,Innove_fixed_asset=Innove_fixed_asset, Status=status)
		saveto_fata.save()

		return HttpResponseRedirect('/Monitoring/Monitoring/')
		
class monitoringCreateView(SuccessMessageMixin, CreateView):
	def dispatch(self, *args, **kwargs):
	    return super().dispatch(*args, **kwargs)
	template_name = 'fata_monitoring.html'
	form_class = FATAmonitoringForm

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "FATA Monitoring Has been Created!"

class monitoringUpdate(SuccessMessageMixin, UpdateView):
	model = Fata_monitoring
	form_class = FATAmonitoringForm
	template_name = 'fata_monitoring.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "FATA Monitoring Updated Successfully!"
		
class monitoringDetails(DetailView):
	model = Fata_monitoring
	template_name = 'fata_monitoring_details.html'
	
class monitoringDeleteView(BSModalDeleteView):
    model = Fata_monitoring
    template_name = 'fata_monitoring_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('Monitoring_list')

class regUpdate(SuccessMessageMixin, UpdateView):
	model = VehicleMasterList
	form_class = reg_updateForm
	template_name = 'regupdate.html'
	success_url = reverse_lazy('Monitoring_jan_reg')
	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Registrations Updated Successfully!"

def regUpdate(request, pk):
	if request.method == 'POST':
		last_reg = request.POST.get('last_reg')
		smoke_date = request.POST.get('smoke_date')
		coc_date = request.POST.get('coc_date')
		remarks = request.POST.get('remarks')
		Status = request.POST.get('Status')
		smoke_due = request.POST.get('smoke_due')
	
	VehicleMasterList.objects.filter(id=pk).update(Last_Registration_Date=last_reg, Smoke_Emission_Date=smoke_date, COC_Date=coc_date, 
		Remarks=remarks, Status=Status, Smoke_due=smoke_due)

	return HttpResponseRedirect('/Masterlist/Registration/Details/{}'.format(pk))

def monitoringHistoryView(request):
    if request.method == "GET":
       obj = Fata_monitoring.history.all()

       return render(request, 'fata_monitoring_history.html', context={'object': obj})

def janRegView(request):
	context = {
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", Last_Registration_Date__isnull=True),
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", Smoke_Emission_Date__isnull=True),
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", COC_Date__isnull=True)	
		}

	return render(request, 'month_reg/regJan_monitoring.html', context)

def febRegView(request):
	context = {
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Last_Registration_Date__isnull=True),
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Smoke_Emission_Date__isnull=True),
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", COC_Date__isnull=True)
		}

	return render(request, 'month_reg/regFeb_monitoring.html', context)

def marRegView(request):
	context = {
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR",Last_Registration_Date__isnull=True),
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR",Smoke_Emission_Date__isnull=True),
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR",COC_Date__isnull=True)
		}

	return render(request, 'month_reg/regMar_monitoring.html', context)

def aprRegView(request):
	context = {
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Last_Registration_Date__isnull=True),
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Smoke_Emission_Date__isnull=True),
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", COC_Date__isnull=True)

		}

	return render(request, 'month_reg/regApr_monitoring.html', context)

def mayRegView(request):
	context = {
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Last_Registration_Date__isnull=True),
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Smoke_Emission_Date__isnull=True),
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", COC_Date__isnull=True)
		}

	return render(request, 'month_reg/regMay_monitoring.html', context)

def junRegView(request):
	context = {
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Last_Registration_Date__isnull=True),
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Smoke_Emission_Date__isnull=True),
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", COC_Date__isnull=True)
		}

	return render(request, 'month_reg/regJun_monitoring.html', context)

def julRegView(request):
	context = {
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Last_Registration_Date__isnull=True),
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Smoke_Emission_Date__isnull=True),
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", COC_Date__isnull=True)
		}
	return render(request, 'month_reg/regJul_monitoring.html', context)

def augRegView(request):
	context = {
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Last_Registration_Date__isnull=True),
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Smoke_Emission_Date__isnull=True),
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", COC_Date__isnull=True)
		}

	return render(request, 'month_reg/regAug_monitoring.html', context)

def sepRegView(request):
	context = {
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Last_Registration_Date__isnull=True),
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Smoke_Emission_Date__isnull=True),
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", COC_Date__isnull=True)
		}

	return render(request, 'month_reg/regSep_monitoring.html', context)

def octRegView(request):
	context = {
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Last_Registration_Date__isnull=True),
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Smoke_Emission_Date__isnull=True),
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", COC_Date__isnull=True)
		}

	return render(request, 'month_reg/regOct_monitoring.html', context)


# SUMMARY BY MONTH
def janSumView(request):
	context = {
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", Last_Registration_Date__isnull=False ),
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", Smoke_Emission_Date__isnull=False),
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumJan_monitoring.html', context)

def febSumView(request):
	context = {
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Last_Registration_Date__isnull=False ),
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Smoke_Emission_Date__isnull=False),
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumFeb_monitoring.html', context)

def marSumView(request):
	context = {
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR", Last_Registration_Date__isnull=False ),
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR", Smoke_Emission_Date__isnull=False),
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumMar_monitoring.html', context)

def aprSumView(request):
	context = {
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Last_Registration_Date__isnull=False ),
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Smoke_Emission_Date__isnull=False),
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumApr_monitoring.html', context)

def maySumView(request):
	context = {
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Last_Registration_Date__isnull=False ),
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Smoke_Emission_Date__isnull=False),
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumMay_monitoring.html', context)

def junSumView(request):
	context = {
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Last_Registration_Date__isnull=False ),
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Smoke_Emission_Date__isnull=False),
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumJun_monitoring.html', context)

def julSumView(request):
	context = {
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Last_Registration_Date__isnull=False ),
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Smoke_Emission_Date__isnull=False),
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumJul_monitoring.html', context)

def augSumView(request):
	context = {
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Last_Registration_Date__isnull=False ),
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Smoke_Emission_Date__isnull=False),
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumAug_monitoring.html', context)

def sepSumView(request):
	context = {
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Last_Registration_Date__isnull=False ),
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Smoke_Emission_Date__isnull=False),
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumSep_monitoring.html', context)

def octSumView(request):
	context = {
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Last_Registration_Date__isnull=False ),
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Smoke_Emission_Date__isnull=False),
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", COC_Date__isnull=False)	
		}

	return render(request, 'sum_reg/sumOct_monitoring.html', context)

def plateMonitoringView(request):
	context = {
			'plate_monitoring': VehicleMasterList.objects.filter(PLATE_NO__isnull=True)
		}

	return render(request, 'plate_monitoring.html', context)

def fata_excel(request):
    fata_queryset = Fata_monitoring.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= FATA Monitoring.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'FATA Monitoring'

    columns = [
    			'Activity id',
				'FATA Number' ,
				'Date Transfer' ,
				'Date Received' ,
				'Plate No' ,
				'Vehicle Make' ,
				'Vehicle Brand' ,
				'Certificate Of Registrations Name' ,
				'Vehicle Model' ,
				'Cr Name',
				'Transferor Employee' ,
				'Transferor First Name' ,
				'Transferor Last Name' ,
				'Recipient Employee' ,
				'Recipient Fist Name' ,
				'Recipient Last Name' ,
				'Date Endorsed Globe' ,
				'Date Endorsed Innove' ,
				'Clearing of Accountability' ,
				'Globe Fixed Asset Recepient' ,
				'Innove Fixed Asset Recepient' ,
				'Remarks',
				'Date Initiated' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for fata in fata_queryset:
        row_num += 1
        row = [
				fata.Activity_id,
				fata.Fata_no,
				fata.Date_transfer,
				fata.Date_received,
				fata.Plate_no,
				fata.Vehicle_make,
				fata.Vehicle_brand,
				fata.Certificate_of_Reg,
				fata.Vehicle_model,
				fata.Cr_name,
				fata.Transferor_employee,
				fata.Transferor_Fname,
				fata.Transferor_Lname,
				fata.Recipient_Employee,
				fata.Recipient_Fname,
				fata.Recipient_Lname,
				fata.Date_endorsed_Globe,
				fata.Date_endorsed_Innove,
				fata.Clearing_accountability,
				fata.Globe_fixed_asset,
				fata.Innove_fixed_asset,
				fata.Remarks,
				fata.Date_initiated,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

#SUMMARY EXPORT
def sum_jan_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_January.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring January'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_feb_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_Febuary.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring Febuary'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_mar_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_March.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring March'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_apr_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_April.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring April'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_may_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_May.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring May'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_jun_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_June.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring June'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_jul_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_July.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring July'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_aug_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_August.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring August'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_sep_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_September.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring September'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def sum_oct_excel(request):
    sum_queryset = VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Last_Registration_Date__isnull=False, Smoke_Emission_Date__isnull=False, COC_Date__isnull=False )   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Summary_Monitoring_October.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Summary Monitoring October'

    columns = [
				'Plate Number' ,
				'Conductions Sticker No.:' ,
				'CR Name' ,
				'Model' ,
				'Brand' ,
				'Vehicle Make' ,
				'Assignee Last Name' ,
				'Assignee First Name' ,
				'MV-File' ,
				'Last Registration Date' ,
				'Smoke Emission Date' ,
				'COC Date' ,
				'Remarks' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sum in sum_queryset:
        row_num += 1
        row = [
				sum.PLATE_NO ,
				sum.CS_NO ,
				sum.CR_NAME ,
				sum.MODEL ,
				sum.BRAND ,
				sum.VEHICLE_MAKE ,
				sum.ASSIGNEE_LAST_NAME ,
				sum.ASSIGNEE_FIRST_NAME ,
				sum.MV_FILE_NO ,
				sum.Last_Registration_Date ,
				sum.Smoke_Emission_Date ,
				sum.COC_Date ,
				sum.Remarks ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
	
def janReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			 'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN",Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		 }

	return render(request, 'reg_due/regJan_monitoring_due.html', context)

def febReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
			}

	return render(request, 'reg_due/regFeb_monitoring_due.html', context)

def marReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR",Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		}

	return render(request, 'reg_due/regMar_monitoring_due.html', context)

def aprReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)

		}

	return render(request, 'reg_due/regApr_monitoring_due.html', context)

def mayReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		}

	return render(request, 'reg_due/regMay_monitoring_due.html', context)

def junReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		}

	return render(request, 'reg_due/regJun_monitoring_due.html', context)

def julReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		}
	return render(request, 'reg_due/regJul_monitoring_due.html', context)

def augReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		}

	return render(request, 'reg_due/regAug_monitoring_due.html', context)

def sepReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		}

	return render(request, 'reg_due/regSep_monitoring_due.html', context)

def octReg_dueView(request):
	today = datetime.datetime.now(tz=timezone.utc)
	context = {
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Smoke_Emission_Date__lt=today, Smoke_due__gt=today)
		}

	return render(request, 'reg_due/regOct_monitoring_due.html', context)


##Daily report Details
def report_details(request):
	# date = datetime.datetime.today()
	# notorized = Ownership.objects.filter(status = 'NOTARIZED').count()
	# routing_count = Ownership.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
	# tmg_schedule = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
	# tmg_appearance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
	# lto_transfer = Ownership.objects.filter(status = 'LTO TRANSFER').count()
	# fleet_vismin = Ownership.objects.filter(status = 'FLEET VISMIN').count()
	# with_tmg_etching = Ownership.objects.filter(status = 'WITH_MACRO ETCHING').count()
	# total = notorized + routing_count + tmg_schedule + tmg_appearance + lto_transfer + fleet_vismin + with_tmg_etching
	# return render(request, 'report_details.html',{'Title':'TOO - TOO Report', 'notorized':notorized,'routing_count':routing_count,
	# 'tmg_schedule':tmg_schedule, 'tmg_appearance':tmg_appearance, 'lto_transfer':lto_transfer, 'fleet_vismin':fleet_vismin,
	# 'with_tmg_etching':with_tmg_etching, 'total':total,'date':date})
	return render(request, 'report.html')
# Registration Daily Report
def report(request):
# 	date = datetime.datetime.today()

# 	notorized = Ownership.objects.filter(status = 'NOTARIZED').count()
# 	routing_count = Ownership.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
# 	tmg_schedule = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
# 	tmg_appearance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
# 	lto_transfer = Ownership.objects.filter(status = 'LTO TRANSFER').count()
# 	fleet_vismin = Ownership.objects.filter(status = 'FLEET VISMIN').count()
# 	with_tmg_etching = Ownership.objects.filter(status = 'WITH_MACRO ETCHING').count()
# 	total = notorized + routing_count + tmg_schedule + tmg_appearance + lto_transfer + fleet_vismin + with_tmg_etching
	from openpyxl import Workbook, load_workbook
	output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	file_name = "FATA_Report.xlsx"
	output['Content-Disposition'] = 'attachment; filename='+ file_name
	wb = Workbook()
	ws = wb.active

# 	#header
	ws.title = "FATA Report"
	ws['A1'].value = "Acquisition / Stephanie Warde"
	ws.append(['', 'Total','Remarks'])
	ws['A3'].value = "* New PR / PO (Ariba)"
	ws['A4'].value = "2nd Hand"
	ws['A5'].value = "FATA /Returned /Inhouse"
	ws['A6'].value = "Delivered Units"
	ws['A7'].value = "For Good Receipt (Approval in Ariba)"
	ws['A8'].value = ""
	ws['A9'].value = "* Documents"
	ws['A10'].value = "FATA"
	ws['A11'].value = "ORCR Monitoring"
	ws['A12'].value = "Stencil Monitoring"
	ws['A13'].value = "Spare Key Monitoring"
	ws['A14'].value = ""
	ws['A15'].value = "* Plate Monitoring"
	ws['A16'].value = "Received Plate"
	ws['A17'].value = "Aging Plate"
	ws['A18'].value = ""
	ws['A19'].value = "* Other Task"
	ws['A20'].value = "Scanning ORCR and Invoices"
# 	#data
# 	ws['B4'].value = date
# 	ws['B10'].value = routing_count
# 	ws['B11'].value = notorized
# 	ws['B14'].value = tmg_appearance
# 	ws['B15'].value = tmg_schedule
# 	ws['B17'].value = with_tmg_etching
# 	ws['B19'].value = fleet_vismin
# 	ws['B21'].value = lto_transfer
# 	ws['B23'].value = total

# 	# style
# 	ws['A6'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['B6'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['C6'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['D6'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['E6'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['F6'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['G6'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	#TMG color
# 	ws['A13'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['B13'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['C13'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['D13'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['E13'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['F13'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['G13'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	#vismin color
# 	ws['A19'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['B19'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['C19'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['D19'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['E19'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['F19'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['G19'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	#lto transfer color
# 	ws['A21'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['B21'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['C21'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['D21'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['E21'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['F21'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	ws['G21'].fill = PatternFill("solid", fgColor="00FFCC99")
# 	#transfer color
# 	ws['A25'].fill = PatternFill("solid", fgColor="00FF99CC")
# 	ws['B25'].fill = PatternFill("solid", fgColor="00FF99CC")
# 	ws['C25'].fill = PatternFill("solid", fgColor="00FF99CC")
# 	ws['D25'].fill = PatternFill("solid", fgColor="00FF99CC")
# 	ws['E25'].fill = PatternFill("solid", fgColor="00FF99CC")
# 	ws['F25'].fill = PatternFill("solid", fgColor="00FF99CC")
# 	ws['G25'].fill = PatternFill("solid", fgColor="00FF99CC")

	wb.save(output)
	return output
		
		