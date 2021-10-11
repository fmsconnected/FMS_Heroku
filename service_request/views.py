from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from django.views import generic
import schedule
import time
from django.core.mail import get_connection,send_mail
from django.core import mail
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.template import Context
from django.shortcuts import render,HttpResponseRedirect, get_list_or_404,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
import datetime
from datetime import date, timedelta
from django.db.models import Q, Count
from .models import (
        service_vehicle
)
from masterlist.models import (
    EmployeeMasterlist,
    )
from vehicle_masterlist.models import VehicleMasterList
from leasingmasterlist.models import Leasing

from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from CustomerLog.models import (
    CS_log
)
from . forms import (
    serviceform
    )

from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

                      #####################################  
                    #########################################
                   #####.     Service Vehicle Request    #####
                    #########################################
                     ######################################


class serviceListView(ListView):
    model = service_vehicle
    template_name = 'service_list.html'

def serviceCreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = Leasing.objects.all()
    return render(request, 'service_new.html',{'Title':'Car - Car Request', 'emplist':emplist,'vlist':vlist})

def servicesubmit(request):
    if request.method == 'POST':
        req_employee_id = request.POST.get('req_employee_id')
        request_date = request.POST.get('request_date')
        deadline = request.POST.get('deadline')
        req_lname = request.POST.get('req_lname')
        req_fname = request.POST.get('req_fname')
        assignee_employee_id = request.POST.get('assignee_employee_id')
        Assignee_Group = request.POST.get('Assignee_Group')
        assignee_fname = request.POST.get('assignee_fname')
        assignee_lname = request.POST.get('assignee_lname')
        assignee_costcenter = request.POST.get('assignee_costcenter')
        assignee_section = request.POST.get('assignee_section')
        assignee_location = request.POST.get('assignee_location')
        assignee_atd = request.POST.get('assignee_atd')
        new_employee_id = request.POST.get('new_employee_id')
        new_employee_fname = request.POST.get('new_employee_fname')
        new_employee_lname = request.POST.get('new_employee_lname')
        new_employee_cost = request.POST.get('new_employee_cost')
        new_temporary_atd = request.POST.get('new_temporary_atd')
        prefered_vehicle = request.POST.get('prefered_vehicle')
        justification = request.POST.get('justification')
        E_plate_no = request.POST.get('E_plate_no')
        E_con_sticker = request.POST.get('E_con_sticker')
        E_model_year = request.POST.get('E_model_year')
        E_brand = request.POST.get('E_brand')
        E_make = request.POST.get('E_make')
        E_type = request.POST.get('E_type')
        approved_by = request.POST.get('approved_by')
        approved_date = request.POST.get('approved_date')
        vehicle_provider = request.POST.get('vehicle_provider')
        vehicle_plate_no = request.POST.get('vehicle_plate_no')
        vehicle_CS_no = request.POST.get('vehicle_CS_no')
        vehicle_model = request.POST.get('vehicle_model')
        vehicle_brand = request.POST.get('vehicle_brand')
        vehicle_make = request.POST.get('vehicle_make')
        vehicle_fuel_type = request.POST.get('vehicle_fuel_type')
        svv_sla = request.POST.get('svv_sla')
        sv_status = request.POST.get('sv_status')

        saveto_service = service_vehicle(request_date=request_date, req_employee_id=req_employee_id, req_lname=req_lname, req_fname =req_fname,
            assignee_employee_id=assignee_employee_id, assignee_group=Assignee_Group, assignee_fname=assignee_fname, assignee_lname=assignee_lname,
            assignee_costcenter=assignee_costcenter, assignee_section=assignee_section, assignee_location=assignee_location, assignee_atd=assignee_atd,
            new_employee_id=new_employee_id, new_employee_fname=new_employee_fname, new_employee_lname=new_employee_lname, new_employee_cost=new_employee_cost,
            new_temporary_atd=new_temporary_atd, prefered_vehicle=prefered_vehicle, justification=justification, E_plate_no=E_plate_no, E_con_sticker=E_con_sticker,
            E_model_year =E_model_year, E_brand=E_brand, E_make=E_make, E_type=E_type, approved_by=approved_by, approved_date=approved_date,
            vehicle_provider =vehicle_provider, vehicle_plate_no=vehicle_plate_no, vehicle_CS_no=vehicle_CS_no, vehicle_model =vehicle_model,
            vehicle_brand =vehicle_brand, vehicle_make=vehicle_make, vehicle_fuel_type=vehicle_fuel_type, SVV_SLA =svv_sla, Deadline=deadline,Status=sv_status)
        saveto_service.save()

        return HttpResponseRedirect('/ServiceRequest/Service/')

class serviceCreateView(SuccessMessageMixin, CreateView):
    model = service_vehicle
    form_class = serviceform
    template_name = 'service_form.html'

    def get_success_message(self, cleaned_data):
        print(cleaned_data)
        return "New Service Vehicle Request Has been Created!"

class serviceUpdateView(SuccessMessageMixin, UpdateView):
    model = service_vehicle
    form_class = serviceform
    template_name = 'service_form.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Service Vehicle Request Update Successfully!"

class serviceDetailView(DetailView):
    model = service_vehicle
    template_name = 'service_details.html'

class serviceDeleteView(BSModalDeleteView):
    model = service_vehicle
    template_name = 'service_delete.html'
    success_message = 'Success: Service Vehicle Request was deleted.'
    success_url = reverse_lazy('service_list')

def serviceHistoryView(request):
    if request.method == "GET":
       obj = service_vehicle.history.all()

       return render(request, 'service_history.html', context={'object': obj})

def service_report(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = Leasing.objects.all()
    return render(request, 'service_vehicle_report.html')


def svr_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = service_vehicle.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = service_vehicle.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = service_vehicle.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = service_vehicle.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = service_vehicle.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = service_vehicle.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'svrdeadline.html',{'title':'Service - Service vehicle Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})


def service_request_excel(request):
    service_queryset = service_vehicle.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Service Vehicle Request.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Service Vehicle Request'

    columns = [
        'Request Date' ,
        'Employee Id' ,
        'Last Name' ,
        'First Name' ,
        'Assignee Employee Id' ,
        'Assignee Group' ,
        'Assignee First Name' ,
        'Assignee Last Name' ,
        'Assignee Cost Center' ,
        'Assignee Section' ,
        'Assignee Location' ,
        'Assignee ATD' ,
        'New Employee Id' ,
        'New Employee First Name' ,
        'New Employee Last Name' ,
        'New Employee Cost Center' ,
        'New Temporary ATD' ,
        'Prefered Vehicle' ,
        'Justification' ,
        'Plate No' ,
        'Conduction Sticker' ,
        'Model' ,
        'Brand' ,
        "Make" ,
        'Type' ,
        'Approved By' ,
        'Approved Date' ,
        'Vehicle Provider' ,
        'Vehicle Plate No' ,
        'Vehicle CS No' ,
        'Vehicle Model' ,
        'Vehicle Brand' ,
        'Vehicle Make' ,
        'Vehicle Fuel Type' ,
        'SLA' ,
        'Date Initiated' ,
        'Deadline',
        'Status',

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for service in service_queryset:
        row_num += 1
        row = [
            service.request_date ,
            service.req_employee_id ,
            service.req_lname ,
            service.req_fname ,
            service.assignee_employee_id ,
            service.assignee_group ,
            service.assignee_fname ,
            service.assignee_lname ,
            service.assignee_costcenter ,
            service.assignee_section ,
            service.assignee_location ,
            service.assignee_atd ,
            service.new_employee_id ,
            service.new_employee_fname ,
            service.new_employee_lname ,
            service.new_employee_cost ,
            service.new_temporary_atd ,
            service.prefered_vehicle ,
            service.justification,
            service.E_plate_no ,
            service.E_con_sticker ,
            service.E_model_year ,
            service.E_brand ,
            service.E_make ,
            service.E_type ,
            service.approved_by ,
            service.approved_date ,
            service.vehicle_provider ,
            service.vehicle_plate_no ,
            service.vehicle_CS_no ,
            service.vehicle_model ,
            service.vehicle_brand ,
            service.vehicle_make ,
            service.vehicle_fuel_type ,
            service.SVV_SLA ,
            service.date_initiated ,
            service.Deadline,
            service.Status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def service_vehicle_report(request):
    date = datetime.datetime.today().year
    m_date = datetime.datetime.today()
    months = ['zero','January','February','March','April','May','June','July','August','September','October','November','December']
    month = months[m_date.month]
    # preventive_count = Vehicle_Repair.objects.filter(status = 'NOTARIZED').count()
    # corrective_count = Corrective.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
    # tire_battery = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
    # insurance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
    
    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Service Vehicle Request.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Service Vehicle Request"
    ws.append([month, date])
    ws['B2'].value = "Monday"
    ws['C2'].value = "Tuesday"
    ws['D2'].value = "Wednesday"
    ws['E2'].value = "Thursday"
    ws['F2'].value = "Friday"
    ws['G2'].value = "Total"

    ws['A3'].value = "A. Leasing/Process"
    ws['A4'].value = "SVRF Line-Up"
    ws['A5'].value = "* New Hire"
    ws['A6'].value = "* Replacement Request"
    ws['A7'].value = ""
    ws['A8'].value = "Deliveries"
    ws['A9'].value = "* Deliveries Endorsement (New)"
    ws['A10'].value = "* Safekeeping (Allocated)"
    ws['A11'].value = ""
    ws['A12'].value = "PULLOUT Endorsement"
    ws['A13'].value = "* Endorsed for Pullout"
    ws['A14'].value = "* Returned (Matured/Terminated)"
    ws['A15'].value = "* Sold"
    ws['A16'].value = ""
    ws['A17'].value = "TOA (Transefer of Accountability)"
    ws['A18'].value = "* Change In Masterlist"
    ws['A19'].value = "* For Processing Waiting Approval"
    ws['A20'].value = ""
    ws['A21'].value = "ENCODE (Masterlist and SAP)"
    ws['A22'].value = "* Input New Vahicle Details"
    ws['A23'].value = "* Create Contract-SAP"
    ws['A24'].value = ""
    ws['A25'].value = "B. Car Rental/Process"
    ws['A26'].value = "* Request"
    ws['A27'].value = "* Booked"
    ws['A28'].value = ""
    ws['A29'].value = "C. Other Task"
    ws['A30'].value = "* Cascading Replacement Vehicle"
    ws['A31'].value = "* Allocation of Temporary Vehicle"

    #data
    # ws['B4'].value = date
    # ws['B10'].value = routing_count
    # ws['B11'].value = notorized
    # ws['B14'].value = tmg_appearance
    # ws['B15'].value = tmg_schedule
    # ws['B17'].value = with_tmg_etching
    # ws['B19'].value = fleet_vismin
    # ws['B21'].value = lto_transfer
    # ws['B23'].value = total

    # style
    ws['A4'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['A8'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['A12'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['A17'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['A21'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['A25'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['A29'].fill = PatternFill("solid", fgColor="00FFCC99")
   
    wb.save(output)
    return output
