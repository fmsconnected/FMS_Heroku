from django.shortcuts import render,HttpResponseRedirect,HttpResponse, redirect
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
import datetime
from datetime import date, timedelta
from django.urls import reverse_lazy

from .serializers import (
    shell_report_Serializer
    )
from . models import shell_report
from rest_framework import viewsets
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                                DeleteView,
                				)
from django.db.models import Sum
from . forms import shell_form

from bootstrap_modal_forms.generic import BSModalDeleteView
                                           
def monthly_report_shell(request):
	return render(request, 'shell/shell_list.html')

class shell_report_ViewSet(viewsets.ModelViewSet):
    queryset = shell_report.objects.all().order_by('id')
    serializer_class = shell_report_Serializer

class shell_new(CreateView):
    model = shell_report
    form_class = shell_form
    template_name = 'shell/shell_create.html'


class shell_update(UpdateView):
    model = shell_report
    form_class = shell_form
    template_name = 'shell/shell_create.html'

# class shell_delete(BSModalDeleteView):
#     model = shell_report
#     template_name = 'shell/shell_delete.html'
#     success_message = 'Success: Item was deleted.'
    # success_url = reverse_lazy('report_shell_list')
class shell_delete(DeleteView):
    model = shell_report
    success_url = reverse_lazy('report_shell_list')
    template_name = 'shell/shell_delete.html'

class shell_details(DetailView):
    model = shell_report
    template_name = 'shell/shell_detail.html'

def reportsummary(request):
    date = datetime.datetime.today()
    BB14_B10=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B10").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_B2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_B4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B5").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_B6=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B6").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_B7=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B7").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_B8=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B8").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    BB14_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMB4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMB4-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG12_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG12-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_B1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_C1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_C2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_C3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_C4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_C5_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_C5_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_C5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_D1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_D2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_D3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_E2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_E3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_F1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_F2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG2_F3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG3_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG3-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG4_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG4_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG4_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG4_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-F").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG4_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG5_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG5_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CMG6_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG6-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    CRA6_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CRA6-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    EIG09_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="EIG09-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN13_K1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN13_K2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN22_D1b=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D1b").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN22_D2a=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D2a").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN23_C1 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN23-C1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN9_C0301=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-C0301").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN9_E02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-E02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN9_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-F0201").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    FIN_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN-F0201").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-F").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_H=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-H").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG02_K =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-K").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GEG04_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG04-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    GENT_F2011=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GENT-F2011").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_D06_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D06-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_D5_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_D5_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-05").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_D6_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_D6_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-05").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_F4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-F4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_G4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-G4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT2_H3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-H3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT3_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-B5").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT3_D5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-D5").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_B05A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B05A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_B06=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B06").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_C01 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-C01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_D01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_D02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_E03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_E04=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E04").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_G01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_G02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_G03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_H01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_H03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_I01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_I02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_I03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_J01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_J02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT5_J03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT6_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT6-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    NTT7_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT7-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    OP12_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="OP12-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG02_O=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG02_O2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG02_P=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-P").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG02_Q=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG02_Q2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG02_R=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-R").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG02_U=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-U").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    SG11_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG11-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    ST1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="ST1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0) or 0
    

    rebate_BB14_B10 = shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B10").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_B2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_B4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B5").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_B6=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B6").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_B7=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B7").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_B8=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B8").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_BB14_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMB4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMB4-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG12_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG12-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_B1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_C1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_C2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_C3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_C4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_C5_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_C5_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_C5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_D1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_D2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_D3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_E2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_E3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_F1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_F2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG2_F3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG3_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG3-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG4_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG4_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG4_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG4_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-F").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG4_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG5_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG5_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CMG6_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG6-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_CRA6_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CRA6-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_EIG09_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="EIG09-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN13_K1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN13_K2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN22_D1b=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D1b").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN22_D2a=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D2a").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN23_C1 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN23-C1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN9_C0301=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-C0301").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN9_E02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-E02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN9_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-F0201").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_FIN_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN-F0201").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-F").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_H=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-H").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG02_K =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-K").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GEG04_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG04-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_GENT_F2011=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GENT-F2011").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_D06_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D06-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_D5_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_D5_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-05").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_D6_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_D6_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-05").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_F4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-F4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_G4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-G4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT2_H3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-H3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT3_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-B5").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT3_D5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-D5").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_B05A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B05A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_B06=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B06").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_C01 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-C01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_D01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_D02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_E03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_E04=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E04").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_G01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_G02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_G03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_H01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_H03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_I01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_I02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_I03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_J01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_J02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT5_J03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT6_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT6-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_NTT7_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT7-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_OP12_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="OP12-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG02_O=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG02_O2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG02_P=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-P").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG02_Q=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG02_Q2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG02_R=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-R").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG02_U=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-U").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_SG11_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG11-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    rebate_ST1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="ST1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0) or 0
    
    grand_total=BB14_B10+BB14_B2+BB14_B3+BB14_B4+BB14_B5+BB14_B6+BB14_B7
    +BB14_B8+BB14_E+CMB4_B+CMG12_D+CMG2_B1+CMG2_B3+CMG2_C1+CMG2_C2
    +CMG2_C3+CMG2_C4+CMG2_C5_B+CMG2_C5_C+CMG2_C5_G+CMG2_D1+CMG2_D2
    +CMG2_D3+CMG2_E2+CMG2_E3+CMG2_F1+CMG2_F2+CMG2_F3+CMG3_A+CMG4_B
    +CMG4_C+CMG4_D+CMG4_E+CMG4_F+CMG4_G+CMG5_D+CMG5_E+CMG5_G
    +CMG6_E+CRA6_A+EIG09_A+FIN13_K1+FIN13_K2+FIN22_D1b+FIN22_D2a
    +FIN23_C1 +FIN9_C0301+FIN9_E02+FIN9_F0201+FIN_F0201+GEG02_A+GEG02_B
    +GEG02_C+GEG02_D+GEG02_E+GEG02_F+GEG02_G+GEG02_H+GEG02_K
    +GEG04_B+GENT_F2011+NTT2_D06_E+NTT2_D5_03+NTT2_D5_05+NTT2_D6_03
    +NTT2_D6_05+NTT2_F4+NTT2_G4+NTT2_H3+NTT3_B5+NTT3_D5+NTT5_B05A
    +NTT5_B06+NTT5_C01+NTT5_D01+NTT5_D02+NTT5_E03+NTT5_E04+NTT5_G01
    +NTT5_G02+NTT5_G03+NTT5_H01+NTT5_H03+NTT5_I01+NTT5_I02+NTT5_I03
    +NTT5_J01+NTT5_J02+NTT5_J03+NTT6_C+NTT7_E+OP12_A+SG02_O
    +SG02_O2+SG02_P+SG02_Q+SG02_Q2+SG02_R+SG02_U+SG11_A+ST1

    reb_grand_total=rebate_BB14_B10+rebate_BB14_B2+rebate_BB14_B3+rebate_BB14_B4+rebate_BB14_B5+rebate_BB14_B6+rebate_BB14_B7
    +rebate_BB14_B8+rebate_BB14_E+rebate_CMB4_B+rebate_CMG12_D+rebate_CMG2_B1+rebate_CMG2_B3+rebate_CMG2_C1+rebate_CMG2_C2
    +rebate_CMG2_C3+rebate_CMG2_C4+rebate_CMG2_C5_B+rebate_CMG2_C5_C+rebate_CMG2_C5_G+rebate_CMG2_D1+rebate_CMG2_D2
    +rebate_CMG2_D3+rebate_CMG2_E2+rebate_CMG2_E3+rebate_CMG2_F1+rebate_CMG2_F2+rebate_CMG2_F3+rebate_CMG3_A+rebate_CMG4_B
    +rebate_CMG4_C+rebate_CMG4_D+rebate_CMG4_E+rebate_CMG4_F+rebate_CMG4_G+rebate_CMG5_D+rebate_CMG5_E+rebate_CMG5_G
    +rebate_CMG6_E+rebate_CRA6_A+rebate_EIG09_A+rebate_FIN13_K1+rebate_FIN13_K2+rebate_FIN22_D1b+rebate_FIN22_D2a
    +rebate_FIN23_C1 +rebate_FIN9_C0301+rebate_FIN9_E02+rebate_FIN9_F0201+rebate_FIN_F0201+rebate_GEG02_A+rebate_GEG02_B
    +GEG02_C+rebate_GEG02_D+rebate_GEG02_E+rebate_GEG02_F+rebate_GEG02_G+rebate_GEG02_H+rebate_GEG02_K
    +rebate_GEG04_B+rebate_GENT_F2011+rebate_NTT2_D06_E+rebate_NTT2_D5_03+rebate_NTT2_D5_05+rebate_NTT2_D6_03
    +rebate_NTT2_D6_05+rebate_NTT2_F4+rebate_NTT2_G4+rebate_NTT2_H3+rebate_NTT3_B5+rebate_NTT3_D5+rebate_NTT5_B05A
    +rebate_NTT5_B06+rebate_NTT5_C01+rebate_NTT5_D01+rebate_NTT5_D02+rebate_NTT5_E03+rebate_NTT5_E04+rebate_NTT5_G01
    +rebate_NTT5_G02+rebate_NTT5_G03+rebate_NTT5_H01+rebate_NTT5_H03+rebate_NTT5_I01+rebate_NTT5_I02+rebate_NTT5_I03
    +rebate_NTT5_J01+rebate_NTT5_J02+rebate_NTT5_J03+rebate_NTT6_C+rebate_NTT7_E+rebate_OP12_A+rebate_SG02_O
    +rebate_SG02_O2+rebate_SG02_P+rebate_SG02_Q+rebate_SG02_Q2+rebate_SG02_R+rebate_SG02_U+rebate_SG11_A+rebate_ST1


    return render(request, 'shell/shell_summary.html',{'title' : 'Shell Report', 'BB14_B10':BB14_B10,'BB14_B2':BB14_B2,'BB14_B3':BB14_B3,'BB14_B4':BB14_B4,'BB14_B5':BB14_B5,'BB14_B6':BB14_B6,'BB14_B7':BB14_B7,'BB14_B8':BB14_B8,'BB14_E':BB14_E,'CMB4_B':CMB4_B,'CMG12_D':CMG12_D,'CMG2_B1':CMG2_B1,
    'CMG2_B3':CMG2_B3,'CMG2_C1':CMG2_C1,'CMG2_C2':CMG2_C2,'CMG2_C3':CMG2_C3,'CMG2_C4':CMG2_C4,'CMG2_C5_B':CMG2_C5_B,'CMG2_C5_C':CMG2_C5_C,'CMG2_C5_G':CMG2_C5_G,'CMG2_D1':CMG2_D1,'CMG2_D2':CMG2_D2,'CMG2_D3':CMG2_D3,
    'CMG2_E2':CMG2_E2,'CMG2_E3':CMG2_E3,'CMG2_F1':CMG2_F1,'CMG2_F2':CMG2_F2,'CMG2_F3':CMG2_F3,'CMG3_A':CMG3_A,'CMG4_B':CMG4_B,'CMG4_C':CMG4_C,'CMG4_D':CMG4_D,'CMG4_E':CMG4_E,'CMG4_F':CMG4_F,'CMG4_G':CMG4_G,'CMG5_D':CMG5_D,
    'CMG5_E':CMG5_E,'CMG5_G':CMG5_G,'CMG6_E':CMG6_E,'CRA6_A':CRA6_A,'EIG09_A':EIG09_A,'FIN13_K1':FIN13_K1,'FIN13_K2':FIN13_K2,'FIN22_D1b':FIN22_D1b,'FIN22_D2a':FIN22_D2a,'FIN23_C1':FIN23_C1 ,
    'FIN9_C0301':FIN9_C0301,'FIN9_E02':FIN9_E02,'FIN9_F0201':FIN9_F0201,'FIN_F0201':FIN_F0201,'GEG02_A':GEG02_A,'GEG02_B':GEG02_B,'GEG02_C':GEG02_C,'GEG02_D':GEG02_D,'GEG02_E':GEG02_E,
    'GEG02_F':GEG02_F,'GEG02_G':GEG02_G,'GEG02_H':GEG02_H,'GEG02_K':GEG02_K ,'GEG04_B':GEG04_B,'GENT_F2011':GENT_F2011,'NTT2_D06_E':NTT2_D06_E,'NTT2_D5_03':NTT2_D5_03,'NTT2_D5_05':NTT2_D5_05,'NTT2_D6_03':NTT2_D6_03,
    'NTT2_D6_05':NTT2_D6_05,'NTT2_F4':NTT2_F4,'NTT2_G4':NTT2_G4,'NTT2_H3':NTT2_H3,'NTT3_B5':NTT3_B5,'NTT3_D5':NTT3_D5,'NTT5_B05A':NTT5_B05A,'NTT5_B06':NTT5_B06,'NTT5_C01':NTT5_C01 ,'NTT5_D01':NTT5_D01,'NTT5_D02':NTT5_D02,
    'NTT5_E03':NTT5_E03,'NTT5_E04':NTT5_E04,'NTT5_G01':NTT5_G01,'NTT5_G02':NTT5_G02,'NTT5_G03':NTT5_G03,'NTT5_H01':NTT5_H01,'NTT5_H03':NTT5_H03,'NTT5_I01':NTT5_I01,'NTT5_I02':NTT5_I02,'NTT5_I03':NTT5_I03,'NTT5_J01':NTT5_J01,
    'NTT5_J02':NTT5_J02,'NTT5_J03':NTT5_J03,'NTT6_C':NTT6_C,'NTT7_E':NTT7_E,'OP12_A':OP12_A,'SG02_O':SG02_O,'SG02_O2':SG02_O2,'SG02_P':SG02_P,'SG02_Q':SG02_Q,'SG02_Q2':SG02_Q2,'SG02_R':SG02_R,'SG02_U':SG02_U,'SG11_A':SG11_A,'ST1':ST1,
    'rebate_BB14_B10':rebate_BB14_B10,'rebate_BB14_B2':rebate_BB14_B2,'rebate_BB14_B3':rebate_BB14_B3,'rebate_BB14_B4':rebate_BB14_B4,'rebate_BB14_B5':rebate_BB14_B5,'rebate_BB14_B6':rebate_BB14_B6,'rebate_BB14_B7':rebate_BB14_B7,'rebate_BB14_B8':rebate_BB14_B8,'rebate_BB14_E':rebate_BB14_E,'rebate_CMB4_B':rebate_CMB4_B,'rebate_CMG12_D':rebate_CMG12_D,'rebate_CMG2_B1':rebate_CMG2_B1,
    'rebate_CMG2_B3':rebate_CMG2_B3,'rebate_CMG2_C1':rebate_CMG2_C1,'rebate_CMG2_C2':rebate_CMG2_C2,'rebate_CMG2_C3':rebate_CMG2_C3,'rebate_CMG2_C4':rebate_CMG2_C4,'rebate_CMG2_C5_B':rebate_CMG2_C5_B,'rebate_CMG2_C5_C':rebate_CMG2_C5_C,'rebate_CMG2_C5_G':rebate_CMG2_C5_G,'rebate_CMG2_D1':rebate_CMG2_D1,'rebate_CMG2_D2':rebate_CMG2_D2,'rebate_CMG2_D3':rebate_CMG2_D3,
    'rebate_CMG2_E2':rebate_CMG2_E2,'rebate_CMG2_E3':rebate_CMG2_E3,'rebate_CMG2_F1':rebate_CMG2_F1,'rebate_CMG2_F2':rebate_CMG2_F2,'rebate_CMG2_F3':rebate_CMG2_F3,'rebate_CMG3_A':rebate_CMG3_A,'rebate_CMG4_B':rebate_CMG4_B,'rebate_CMG4_C':rebate_CMG4_C,'rebate_CMG4_D':rebate_CMG4_D,'rebate_CMG4_E':rebate_CMG4_E,'rebate_CMG4_F':rebate_CMG4_F,'rebate_CMG4_G':rebate_CMG4_G,'rebate_CMG5_D':rebate_CMG5_D,
    'rebate_CMG5_E':rebate_CMG5_E,'rebate_CMG5_G':rebate_CMG5_G,'rebate_CMG6_E':rebate_CMG6_E,'rebate_CRA6_A':rebate_CRA6_A,'rebate_EIG09_A':rebate_EIG09_A,'rebate_FIN13_K1':rebate_FIN13_K1,'rebate_FIN13_K2':rebate_FIN13_K2,'rebate_FIN22_D1b':rebate_FIN22_D1b,'rebate_FIN22_D2a':rebate_FIN22_D2a,'rebate_FIN23_C1':rebate_FIN23_C1 ,
    'rebate_FIN9_C0301':rebate_FIN9_C0301,'rebate_FIN9_E02':rebate_FIN9_E02,'rebate_FIN9_F0201':rebate_FIN9_F0201,'rebate_FIN_F0201':rebate_FIN_F0201,'rebate_GEG02_A':rebate_GEG02_A,'rebate_GEG02_B':rebate_GEG02_B,'rebate_GEG02_C':rebate_GEG02_C,'rebate_GEG02_D':rebate_GEG02_D,'rebate_GEG02_E':rebate_GEG02_E,
    'rebate_GEG02_F':rebate_GEG02_F,'rebate_GEG02_G':rebate_GEG02_G,'rebate_GEG02_H':rebate_GEG02_H,'rebate_GEG02_K':rebate_GEG02_K ,'rebate_GEG04_B':rebate_GEG04_B,'rebate_GENT_F2011':rebate_GENT_F2011,'rebate_NTT2_D06_E':rebate_NTT2_D06_E,'rebate_NTT2_D5_03':rebate_NTT2_D5_03,'rebate_NTT2_D5_05':rebate_NTT2_D5_05,'rebate_NTT2_D6_03':rebate_NTT2_D6_03,
    'rebate_NTT2_D6_05':rebate_NTT2_D6_05,'rebate_NTT2_F4':rebate_NTT2_F4,'rebate_NTT2_G4':rebate_NTT2_G4,'rebate_NTT2_H3':rebate_NTT2_H3,'rebate_NTT3_B5':rebate_NTT3_B5,'rebate_NTT3_D5':rebate_NTT3_D5,'rebate_NTT5_B05A':rebate_NTT5_B05A,'rebate_NTT5_B06':rebate_NTT5_B06,'rebate_NTT5_C01':rebate_NTT5_C01 ,'rebate_NTT5_D01':rebate_NTT5_D01,'rebate_NTT5_D02':rebate_NTT5_D02,
    'rebate_NTT5_E03':rebate_NTT5_E03,'rebate_NTT5_E04':rebate_NTT5_E04,'rebate_NTT5_G01':rebate_NTT5_G01,'rebate_NTT5_G02':rebate_NTT5_G02,'rebate_NTT5_G03':rebate_NTT5_G03,'rebate_NTT5_H01':rebate_NTT5_H01,'rebate_NTT5_H03':rebate_NTT5_H03,'rebate_NTT5_I01':rebate_NTT5_I01,'rebate_NTT5_I02':rebate_NTT5_I02,'rebate_NTT5_I03':rebate_NTT5_I03,'rebate_NTT5_J01':rebate_NTT5_J01,
    'rebate_NTT5_J02':rebate_NTT5_J02,'rebate_NTT5_J03':rebate_NTT5_J03,'rebate_NTT6_C':rebate_NTT6_C,'rebate_NTT7_E':rebate_NTT7_E,'rebate_OP12_A':rebate_OP12_A,'rebate_SG02_O':rebate_SG02_O,'rebate_SG02_O2':rebate_SG02_O2,'rebate_SG02_P':rebate_SG02_P,'rebate_SG02_Q':rebate_SG02_Q,'rebate_SG02_Q2':rebate_SG02_Q2,'rebate_SG02_R':rebate_SG02_R,'rebate_SG02_U':rebate_SG02_U,'rebate_SG11_A':rebate_SG11_A,'rebate_ST1':rebate_ST1,
    'grand_total':grand_total,'reb_grand_total':reb_grand_total
    })


def shellreport_export(request):
    date = datetime.datetime.today()
    shell_data = shell_report.objects.filter(InvoiceDate__year=date.year,InvoiceDate__month=date.month)
    BB14_B10 = shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B10").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_B2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_B4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B5").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_B6=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B6").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_B7=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B7").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_B8=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B8").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    BB14_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMB4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMB4-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG12_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG12-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_B1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_C1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_C2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_C3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_C4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_C5_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_C5_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_C5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_D1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_D2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_D3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_E2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_E3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_F1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_F2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG2_F3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG3_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG3-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG4_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG4_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG4_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG4_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-F").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG4_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG5_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG5_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CMG6_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG6-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    CRA6_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CRA6-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    EIG09_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="EIG09-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN13_K1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN13_K2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN22_D1b=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D1b").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN22_D2a=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D2a").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN23_C1 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN23-C1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN9_C0301=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-C0301").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN9_E02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-E02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN9_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-F0201").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    FIN_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN-F0201").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-D").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-F").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-G").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_H=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-H").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG02_K =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-K").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GEG04_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG04-B").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    GENT_F2011=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GENT-F2011").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_D06_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D06-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_D5_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_D5_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-05").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_D6_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_D6_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-05").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_F4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-F4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_G4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-G4").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT2_H3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-H3").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT3_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-B5").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT3_D5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-D5").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_B05A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B05A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_B06=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B06").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_C01 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-C01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_D01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_D02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_E03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_E04=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E04").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_G01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_G02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_G03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_H01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_H03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_I01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_I02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_I03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_J01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J01").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_J02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J02").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT5_J03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J03").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT6_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT6-C").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    NTT7_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT7-E").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    OP12_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="OP12-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG02_O=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG02_O2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG02_P=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-P").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG02_Q=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG02_Q2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q2").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG02_R=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-R").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG02_U=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-U").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    SG11_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG11-A").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    ST1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="ST1").aggregate(Sum('DelcoGrossValue')).get('DelcoGrossValue__sum', 0.00) or 0
    
    grand_total=BB14_B10+BB14_B2+BB14_B3+BB14_B4+BB14_B5+BB14_B6+BB14_B7
    +BB14_B8+BB14_E+CMB4_B+CMG12_D+CMG2_B1+CMG2_B3+CMG2_C1+CMG2_C2
    +CMG2_C3+CMG2_C4+CMG2_C5_B+CMG2_C5_C+CMG2_C5_G+CMG2_D1+CMG2_D2
    +CMG2_D3+CMG2_E2+CMG2_E3+CMG2_F1+CMG2_F2+CMG2_F3+CMG3_A+CMG4_B
    +CMG4_C+CMG4_D+CMG4_E+CMG4_F+CMG4_G+CMG5_D+CMG5_E+CMG5_G
    +CMG6_E+CRA6_A+EIG09_A+FIN13_K1+FIN13_K2+FIN22_D1b+FIN22_D2a
    +FIN23_C1 +FIN9_C0301+FIN9_E02+FIN9_F0201+FIN_F0201+GEG02_A+GEG02_B
    +GEG02_C+GEG02_D+GEG02_E+GEG02_F+GEG02_G+GEG02_H+GEG02_K
    +GEG04_B+GENT_F2011+NTT2_D06_E+NTT2_D5_03+NTT2_D5_05+NTT2_D6_03
    +NTT2_D6_05+NTT2_F4+NTT2_G4+NTT2_H3+NTT3_B5+NTT3_D5+NTT5_B05A
    +NTT5_B06+NTT5_C01+NTT5_D01+NTT5_D02+NTT5_E03+NTT5_E04+NTT5_G01
    +NTT5_G02+NTT5_G03+NTT5_H01+NTT5_H03+NTT5_I01+NTT5_I02+NTT5_I03
    +NTT5_J01+NTT5_J02+NTT5_J03+NTT6_C+NTT7_E+OP12_A+SG02_O
    +SG02_O2+SG02_P+SG02_Q+SG02_Q2+SG02_R+SG02_U+SG11_A+ST1

    rebate_BB14_B10 = shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B10").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_B2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_B4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B5").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_B6=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B6").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_B7=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B7").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_B8=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-B8").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_BB14_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="BB14-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMB4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMB4-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG12_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG12-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_B1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_B3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-B3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_C1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_C2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_C3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_C4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_C5_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_C5_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_C5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-C5-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_D1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_D2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_D3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-D3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_E2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_E3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-E3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_F1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_F2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG2_F3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG2-F3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG3_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG3-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG4_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG4_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG4_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG4_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG4_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-F").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG4_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG4-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG5_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG5_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG5_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG5-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CMG6_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CMG6-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_CRA6_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="CRA6-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_EIG09_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="EIG09-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN13_K1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN13_K2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN13-K2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN22_D1b=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D1b").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN22_D2a=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN22-D2a").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN23_C1 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN23-C1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN9_C0301=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-C0301").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN9_E02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-E02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN9_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN9-F0201").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_FIN_F0201=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="FIN-F0201").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_D=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-D").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_F=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-F").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_G=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-G").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_H=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-H").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG02_K =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG02-K").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GEG04_B=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GEG04-B").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_GENT_F2011=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="GENT-F2011").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_D06_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D06-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_D5_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_D5_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D5-05").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_D6_03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_D6_05=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-D6-05").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_F4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-F4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_G4=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-G4").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT2_H3=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT2-H3").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT3_B5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-B5").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT3_D5=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT3-D5").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_B05A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B05A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_B06=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-B06").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_C01 =shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-C01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_D01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_D02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-D02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_E03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_E04=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-E04").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_G01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_G02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_G03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-G03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_H01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_H03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-H03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_I01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_I02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_I03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-I03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_J01=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J01").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_J02=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J02").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT5_J03=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT5-J03").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT6_C=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT6-C").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_NTT7_E=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="NTT7-E").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_OP12_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="OP12-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG02_O=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG02_O2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-O2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG02_P=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-P").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG02_Q=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG02_Q2=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-Q2").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG02_R=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-R").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG02_U=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG02-U").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_SG11_A=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="SG11-A").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    rebate_ST1=shell_report.objects.filter(InvoiceDate__year=date.year, 
        InvoiceDate__month=date.month, Supplier='Shell', CostCenter="ST1").aggregate(Sum('RebateCustAmount')).get('RebateCustAmount__sum', 0.00) or 0
    
    reb_grand_total=rebate_BB14_B10+rebate_BB14_B2+rebate_BB14_B3+rebate_BB14_B4+rebate_BB14_B5+rebate_BB14_B6+rebate_BB14_B7
    +rebate_BB14_B8+rebate_BB14_E+rebate_CMB4_B+rebate_CMG12_D+rebate_CMG2_B1+rebate_CMG2_B3+rebate_CMG2_C1+rebate_CMG2_C2
    +rebate_CMG2_C3+rebate_CMG2_C4+rebate_CMG2_C5_B+rebate_CMG2_C5_C+rebate_CMG2_C5_G+rebate_CMG2_D1+rebate_CMG2_D2
    +rebate_CMG2_D3+rebate_CMG2_E2+rebate_CMG2_E3+rebate_CMG2_F1+rebate_CMG2_F2+rebate_CMG2_F3+rebate_CMG3_A+rebate_CMG4_B
    +rebate_CMG4_C+rebate_CMG4_D+rebate_CMG4_E+rebate_CMG4_F+rebate_CMG4_G+rebate_CMG5_D+rebate_CMG5_E+rebate_CMG5_G
    +rebate_CMG6_E+rebate_CRA6_A+rebate_EIG09_A+rebate_FIN13_K1+rebate_FIN13_K2+rebate_FIN22_D1b+rebate_FIN22_D2a
    +rebate_FIN23_C1 +rebate_FIN9_C0301+rebate_FIN9_E02+rebate_FIN9_F0201+rebate_FIN_F0201+rebate_GEG02_A+rebate_GEG02_B
    +GEG02_C+rebate_GEG02_D+rebate_GEG02_E+rebate_GEG02_F+rebate_GEG02_G+rebate_GEG02_H+rebate_GEG02_K
    +rebate_GEG04_B+rebate_GENT_F2011+rebate_NTT2_D06_E+rebate_NTT2_D5_03+rebate_NTT2_D5_05+rebate_NTT2_D6_03
    +rebate_NTT2_D6_05+rebate_NTT2_F4+rebate_NTT2_G4+rebate_NTT2_H3+rebate_NTT3_B5+rebate_NTT3_D5+rebate_NTT5_B05A
    +rebate_NTT5_B06+rebate_NTT5_C01+rebate_NTT5_D01+rebate_NTT5_D02+rebate_NTT5_E03+rebate_NTT5_E04+rebate_NTT5_G01
    +rebate_NTT5_G02+rebate_NTT5_G03+rebate_NTT5_H01+rebate_NTT5_H03+rebate_NTT5_I01+rebate_NTT5_I02+rebate_NTT5_I03
    +rebate_NTT5_J01+rebate_NTT5_J02+rebate_NTT5_J03+rebate_NTT6_C+rebate_NTT7_E+rebate_OP12_A+rebate_SG02_O
    +rebate_SG02_O2+rebate_SG02_P+rebate_SG02_Q+rebate_SG02_Q2+rebate_SG02_R+rebate_SG02_U+rebate_SG11_A+rebate_ST1

    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Shell Summary.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Summary"
    ws['A1'].value = ""
    ws['A2'].value = ""
    ws.append(['Cost Center', 'Sum of DelcoGrossValue','Sum of RebateCustAmount'])
    ws['A4'].value = "BB14-B10"
    ws['A5'].value = "BB14-B2"
    ws['A6'].value = "BB14-B3"
    ws['A7'].value = "BB14-B4"
    ws['A8'].value = "BB14-B5"
    ws['A9'].value = "BB14-B6"
    ws['A10'].value = "BB14-B7"
    ws['A11'].value = "BB14-B8"
    ws['A12'].value = "BB14-E"
    ws['A13'].value = "CMB4-B"
    ws['A14'].value = "CMG12-D"
    ws['A15'].value = "CMG2-B1"
    ws['A16'].value = "CMG2-B3"
    ws['A17'].value = "CMG2-C1"
    ws['A18'].value = "CMG2-C2"
    ws['A19'].value = "CMG2-C3"
    ws['A20'].value = "CMG2-C4"
    ws['A21'].value = "CMG2-C5-B"
    ws['A22'].value = "CMG2-C5-C"
    ws['A23'].value = "CMG2-C5-G"
    ws['A24'].value = "CMG2-D1"
    ws['A25'].value = "CMG2-D2"
    ws['A26'].value = "CMG2-D3"
    ws['A27'].value = "CMG2-E2"
    ws['A28'].value = "CMG2-E3"
    ws['A29'].value = "CMG2-F1"
    ws['A30'].value = "CMG2-F2"
    ws['A31'].value = "CMG2-F3"   
    ws['A32'].value = "CMG3-A"
    ws['A33'].value = "CMG4-B"
    ws['A34'].value = "CMG4-C"
    ws['A35'].value = "CMG4-D"
    ws['A36'].value = "CMG4-E"
    ws['A37'].value = "CMG4-F"
    ws['A38'].value = "CMG4-G"
    ws['A39'].value = "CMG5-D"
    ws['A40'].value = "CMG5-E"
    ws['A41'].value = "CMG5-G"
    ws['A42'].value = "CMG6-E"
    ws['A43'].value = "CRA6-A"
    ws['A44'].value = "EIG09-A"
    ws['A45'].value = "FIN13-K1"  
    ws['A46'].value = "FIN13-K2"
    ws['A47'].value = "FIN22-D1b"
    ws['A48'].value = "FIN22-D2a"
    ws['A49'].value = "FIN23-C1"    
    ws['A50'].value = "FIN9-C0301"
    ws['A51'].value = "FIN9-E02"
    ws['A52'].value = "FIN9-F0201"
    ws['A53'].value = "FIN-F0201"
    ws['A54'].value = "GEG02-A"
    ws['A55'].value = "GEG02-B"
    ws['A56'].value = "GEG02-C" 
    ws['A57'].value = "GEG02-D"
    ws['A58'].value = "GEG02-E"
    ws['A59'].value = "GEG02-F"
    ws['A60'].value = "GEG02-G"
    ws['A61'].value = "GEG02-H"
    ws['A62'].value = "GEG02-K"  
    ws['A63'].value = "GEG04-B"
    ws['A64'].value = "GENT-F2011"
    ws['A65'].value = "NTT2-D06-E"
    ws['A66'].value = "NTT2-D5-03"
    ws['A67'].value = "NTT2-D5-05"
    ws['A68'].value = "NTT2-D6-03"
    ws['A69'].value = "NTT2-D6-05"
    ws['A70'].value = "NTT2-F4"
    ws['A71'].value = "NTT2-G4"
    ws['A72'].value = "NTT2-H3"
    ws['A73'].value = "NTT3-B5"
    ws['A74'].value = "NTT3-D5"
    ws['A75'].value = "NTT5-B05A"
    ws['A76'].value = "NTT5-B06"
    ws['A77'].value = "NTT5-C01"
    ws['A78'].value = "NTT5-D01"
    ws['A79'].value = "NTT5-D02"
    ws['A80'].value = "NTT5-E03"
    ws['A81'].value = "NTT5-E04"
    ws['A82'].value = "NTT5-G01"  
    ws['A83'].value = "NTT5-G02"
    ws['A84'].value = "NTT5-G03"
    ws['A85'].value = "NTT5-H01"   
    ws['A86'].value = "NTT5-H03"
    ws['A87'].value = "NTT5-I01"
    ws['A88'].value = "NTT5-I02"
    ws['A89'].value = "NTT5-I03"
    ws['A90'].value = "NTT5-J01"
    ws['A91'].value = "NTT5-J02"
    ws['A92'].value = "NTT5-J03"
    ws['A93'].value = "NTT6-C"
    ws['A94'].value = "NTT7-E"
    ws['A95'].value = "OP12-A"
    ws['A96'].value = "SG02-O"
    ws['A97'].value = "SG02-O2"
    ws['A98'].value = "SG02-P"
    ws['A99'].value = "SG02-Q"
    ws['A100'].value = "SG02-Q2"
    ws['A101'].value = "SG02-R"
    ws['A102'].value = "SG02-U"
    ws['A103'].value = "SG11-A"
    ws['A104'].value = "ST1"
    ws['A105'].value = "Grand Total"


    # Data DelcoGrossValue
    ws['B4'].value = BB14_B10
    ws['B5'].value = BB14_B2
    ws['B6'].value = BB14_B3
    ws['B7'].value = BB14_B4
    ws['B8'].value = BB14_B5
    ws['B9'].value = BB14_B6
    ws['B10'].value = BB14_B7
    ws['B11'].value = BB14_B8
    ws['B12'].value = BB14_E
    ws['B13'].value = CMB4_B
    ws['B14'].value = CMG12_D
    ws['B15'].value = CMG2_B1
    ws['B16'].value = CMG2_B3
    ws['B17'].value = CMG2_C1
    ws['B18'].value = CMG2_C2
    ws['B19'].value = CMG2_C3
    ws['B20'].value = CMG2_C4
    ws['B21'].value = CMG2_C5_B
    ws['B22'].value = CMG2_C5_C
    ws['B23'].value = CMG2_C5_G
    ws['B24'].value = CMG2_C5_G   
    ws['B25'].value = CMG2_D1
    ws['B26'].value = CMG2_D2
    ws['B27'].value = CMG2_D3
    ws['B28'].value = CMG2_E2
    ws['B29'].value = CMG2_E3
    ws['B30'].value = CMG2_F1
    ws['B31'].value = CMG2_F2
    ws['B32'].value = CMG2_F3
    ws['B33'].value = CMG3_A
    ws['B34'].value = CMG4_B
    ws['B35'].value = CMG4_C
    ws['B36'].value = CMG4_D
    ws['B37'].value = CMG4_E
    ws['B38'].value = CMG4_F
    ws['B39'].value = CMG4_G
    ws['B40'].value = CMG5_D
    ws['B41'].value = CMG5_E
    ws['B42'].value = CMG5_G
    ws['B43'].value = CMG6_E
    ws['B44'].value = CRA6_A
    ws['B45'].value = EIG09_A
    ws['B46'].value = FIN13_K1  
    ws['B47'].value = FIN13_K2
    ws['B48'].value = FIN22_D1b
    ws['B49'].value = FIN22_D2a
    ws['B50'].value = FIN23_C1    
    ws['B51'].value = FIN9_C0301
    ws['B52'].value = FIN9_E02
    ws['B53'].value = FIN9_F0201
    ws['B54'].value = FIN_F0201
    ws['B55'].value = GEG02_A
    ws['B56'].value = GEG02_B
    ws['B57'].value = GEG02_C
    ws['B58'].value = GEG02_D
    ws['B59'].value = GEG02_E
    ws['B60'].value = GEG02_F
    ws['B61'].value = GEG02_G
    ws['B62'].value = GEG02_H
    ws['B63'].value = GEG02_K   
    ws['B64'].value = GEG04_B
    ws['B65'].value = GENT_F2011
    ws['B66'].value = NTT2_D06_E
    ws['B67'].value = NTT2_D5_03
    ws['B68'].value = NTT2_D5_05
    ws['B69'].value = NTT2_D6_03
    ws['B70'].value = NTT2_D6_05
    ws['B71'].value = NTT2_F4
    ws['B72'].value = NTT2_G4
    ws['B73'].value = NTT2_H3
    ws['B74'].value = NTT3_B5
    ws['B75'].value = NTT5_B05A
    ws['B76'].value = NTT5_B06
    ws['B77'].value = NTT5_C01
    ws['B78'].value = NTT5_D01
    ws['B79'].value = NTT5_D02
    ws['B80'].value = NTT5_E03
    ws['B81'].value = NTT5_E04
    ws['B82'].value = NTT5_G01  
    ws['B83'].value = NTT5_G02
    ws['B84'].value = NTT5_G03
    ws['B85'].value = NTT5_H01   
    ws['B86'].value = NTT5_H03
    ws['B87'].value = NTT5_I01
    ws['B88'].value = NTT5_I02
    ws['B89'].value = NTT5_I03
    ws['B90'].value = NTT5_J01
    ws['B91'].value = NTT5_J02
    ws['B92'].value = NTT5_J03
    ws['B93'].value = NTT6_C
    ws['B94'].value = NTT7_E
    ws['B95'].value = OP12_A
    ws['B96'].value = SG02_O
    ws['B97'].value = SG02_O2
    ws['B98'].value = SG02_P
    ws['B99'].value = SG02_Q
    ws['B100'].value = SG02_Q2
    ws['B101'].value = SG02_R
    ws['B102'].value = SG02_U
    ws['B103'].value = SG11_A
    ws['B104'].value = ST1
    ws['B105'].value = grand_total

    #Data RebateCustAmount
    ws['C4'].value = rebate_BB14_B10
    ws['C5'].value = rebate_BB14_B2
    ws['C6'].value = rebate_BB14_B3
    ws['C7'].value = rebate_BB14_B4
    ws['C8'].value = rebate_BB14_B5
    ws['C9'].value = rebate_BB14_B6
    ws['C10'].value = rebate_BB14_B7
    ws['C11'].value = rebate_BB14_B8
    ws['C12'].value = rebate_BB14_E
    ws['C13'].value = rebate_CMB4_B
    ws['C14'].value = rebate_CMG12_D
    ws['C15'].value = rebate_CMG2_B1
    ws['C16'].value = rebate_CMG2_B3
    ws['C17'].value = rebate_CMG2_C1
    ws['C18'].value = rebate_CMG2_C2
    ws['C19'].value = rebate_CMG2_C3
    ws['C20'].value = rebate_CMG2_C4
    ws['C21'].value = rebate_CMG2_C5_B
    ws['C22'].value = rebate_CMG2_C5_C
    ws['C23'].value = rebate_CMG2_C5_G 
    ws['C24'].value = rebate_CMG2_D1
    ws['C25'].value = rebate_CMG2_D2
    ws['C26'].value = rebate_CMG2_D3
    ws['C27'].value = rebate_CMG2_E2
    ws['C28'].value = rebate_CMG2_E3
    ws['C29'].value = rebate_CMG2_F1
    ws['C30'].value = rebate_CMG2_F2
    ws['C31'].value = rebate_CMG2_F3
    ws['C32'].value = rebate_CMG3_A
    ws['C33'].value = rebate_CMG4_B
    ws['C34'].value = rebate_CMG4_C
    ws['C35'].value = rebate_CMG4_D
    ws['C36'].value = rebate_CMG4_E
    ws['C37'].value = rebate_CMG4_F
    ws['C38'].value = rebate_CMG4_G
    ws['C39'].value = rebate_CMG5_D
    ws['C40'].value = rebate_CMG5_E
    ws['C41'].value = rebate_CMG5_G
    ws['C42'].value = rebate_CMG6_E
    ws['C43'].value = rebate_CRA6_A
    ws['C44'].value = rebate_EIG09_A
    ws['C45'].value = rebate_FIN13_K1  
    ws['C46'].value = rebate_FIN13_K2
    ws['C47'].value = rebate_FIN22_D1b
    ws['C48'].value = rebate_FIN22_D2a
    ws['C49'].value = rebate_FIN23_C1    
    ws['C50'].value = rebate_FIN9_C0301
    ws['C51'].value = rebate_FIN9_E02
    ws['C52'].value = rebate_FIN9_F0201
    ws['C53'].value = rebate_FIN_F0201
    ws['C54'].value = rebate_GEG02_A
    ws['C55'].value = rebate_GEG02_B
    ws['C56'].value = rebate_GEG02_C
    ws['C57'].value = rebate_GEG02_D
    ws['C58'].value = rebate_GEG02_E
    ws['C59'].value = rebate_GEG02_F
    ws['C60'].value = rebate_GEG02_G
    ws['C61'].value = rebate_GEG02_H
    ws['C62'].value = rebate_GEG02_K   
    ws['C63'].value = rebate_GEG04_B
    ws['C64'].value = rebate_GENT_F2011
    ws['C65'].value = rebate_NTT2_D06_E
    ws['C66'].value = rebate_NTT2_D5_03
    ws['C67'].value = rebate_NTT2_D5_05
    ws['C68'].value = rebate_NTT2_D6_03
    ws['C69'].value = rebate_NTT2_D6_05
    ws['C70'].value = rebate_NTT2_F4
    ws['C71'].value = rebate_NTT2_G4
    ws['C72'].value = rebate_NTT2_H3
    ws['C73'].value = rebate_NTT3_B5
    ws['C74'].value = rebate_NTT3_D5
    ws['C75'].value = rebate_NTT5_B05A
    ws['C76'].value = rebate_NTT5_B06
    ws['C77'].value = rebate_NTT5_C01 
    ws['C78'].value = rebate_NTT5_D01
    ws['C79'].value = rebate_NTT5_D02
    ws['C80'].value = rebate_NTT5_E03
    ws['C81'].value = rebate_NTT5_E04
    ws['C82'].value = rebate_NTT5_G01 
    ws['C83'].value = rebate_NTT5_G02
    ws['C84'].value = rebate_NTT5_G03
    ws['C85'].value = rebate_NTT5_H01   
    ws['C86'].value = rebate_NTT5_H03
    ws['C87'].value = rebate_NTT5_I01
    ws['C88'].value = rebate_NTT5_I02
    ws['C89'].value = rebate_NTT5_I03
    ws['C90'].value = rebate_NTT5_J01
    ws['C91'].value = rebate_NTT5_J02
    ws['C92'].value = rebate_NTT5_J03
    ws['C93'].value = rebate_NTT6_C
    ws['C94'].value = rebate_NTT7_E
    ws['C95'].value = rebate_OP12_A
    ws['C96'].value = rebate_SG02_O
    ws['C97'].value = rebate_SG02_O2
    ws['C98'].value = rebate_SG02_P
    ws['C99'].value = rebate_SG02_Q
    ws['C100'].value = rebate_SG02_Q2
    ws['C101'].value = rebate_SG02_R
    ws['C102'].value = rebate_SG02_U
    ws['C103'].value = rebate_SG11_A
    ws['C104'].value = rebate_ST1
    ws['C105'].value = reb_grand_total

    ws1 = wb.create_sheet("Raw Data")
    ws1.title = "Raw Data"
    columns = [
            'AccountNumber',
            'CustomerName1',
            'FullCardNumber',
            'DelcoListPrice',
            'NetCustomerAmount',
            'CardSequenceNumber',
            'CheckDigit',
            'DelcoGrossValue',
            'DelcoVatAmount',
            'DelcoVatRate',
            'RebateCustAmount',
            'VATNetAmount',
            'DelcoRebateRate',
            'VoucherNumber',
            'FleetId',
            'Quantity',
            'DialogueIndicator',
            'IssuerCode1',
            'NetInvoiceIndicator',
            'NetworkCode',
            'CardGroupName',
            'CardHolderName',
            'CardVehicleRegistrationNumber',
            'DeliveryNetAmount',
            'PinIndicator',
            'RebateType',
            'RecordType',
            'ReleaseCode',
            'Time',
            'TransactionStatus',
            'TransactionType1',
            'VATIndicator',
            'VatUsage',
            'IssuerCountry',
            'DelcoCode',
            'DelcoCountryCode1',
            'TransactionCurrency',
            'InvoiceNumber',
            'InvoiceDate',
            'ProductCode',
            'SiteCode',
            'Site',
            'DeliveryDate',
            'OdometerReading1',
            'DelcoCustRate1',
            'CustomerCurrencyName1',
            'DelcoPumpPrice',
            'DelcoRebateAmount1',
            'CustRebateRate1',
            'PayerDisplayName',
            'PayerCode',
            'CardGroupID',
            'ProductName',
            'CreditDebitCode',
            'CorrectionFlag',
            'DelcoCountryCode',
            'VATApplicable',
            'CustomerEuroRate',
            'DelcoEuroRate',
            'EuroRebateAmount',
            'EuroVATAmount',
            'InvoiceSequenceNo',
            'ListPriceInTransactionCurrency',
            'NetEuroAmount',
            'PumpPriceInTransactionCurrency',
            'RebateonNetAmountInCustomerCurrency',
            'RebateonNetAmountInTransactionCurrency',
            'RebateRate',
            'UnInvoiceSequenceNo',
            'NetworkName',
            'Additional1',
            'Additional2',
            'Additional3',
            'Additional4',
            'TransactionIdentifier',
            'TransactionAuthorisationCode',
            'DelcoName',
            'SiteGroupid',
            'SiteGroupName',
            'Cardtype',
            'PayerGroup',
            'PayerNumber',
            'RefundFlag',
            'RefundOriginalTransactionId',
            'PostingDate',
            'PostingTime ',
            'CustomerReferenceValue',
            'CustomerReferenceDescription',
            'CardPayerAssociationCode',
            'CardPayerAssociation',
            'IncomingProductCode',
            'TransactionProviderId',
            'TransactionProviderName',
            'FileName',
            'FileDate',
            'AccountGroup3Code',
            'AccountGroup3Name',
            'ContractedRebateRate',
            'IsBestOfPricingApplicable',
            'CostCenter',

    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws1.cell(row=row_num, column=col_num)
        cell.value = column_title

    for shell_data in shell_data:
        row_num += 1
        row = [
            shell_data.AccountNumber,
            shell_data.CustomerName1,
            shell_data.FullCardNumber,
            shell_data.DelcoListPrice,
            shell_data.NetCustomerAmount,
            shell_data.CardSequenceNumber,
            shell_data.CheckDigit,
            shell_data.DelcoGrossValue,
            shell_data.DelcoVatAmount,
            shell_data.DelcoVatRate,
            shell_data.RebateCustAmount,
            shell_data.VATNetAmount,
            shell_data.DelcoRebateRate,
            shell_data.VoucherNumber,
            shell_data.FleetId,
            shell_data.Quantity,
            shell_data.DialogueIndicator,
            shell_data.IssuerCode1,
            shell_data.NetInvoiceIndicator,
            shell_data.NetworkCode,
            shell_data.CardGroupName,
            shell_data.CardHolderName,
            shell_data.CardVehicleRegistrationNumber,
            shell_data.DeliveryNetAmount,
            shell_data.PinIndicator,
            shell_data.RebateType,
            shell_data.RecordType,
            shell_data.ReleaseCode,
            shell_data.Time,
            shell_data.TransactionStatus,
            shell_data.TransactionType1,
            shell_data.VATIndicator,
            shell_data.VatUsage,
            shell_data.IssuerCountry,
            shell_data.DelcoCode,
            shell_data.DelcoCountryCode1,
            shell_data.TransactionCurrency,
            shell_data.InvoiceNumber,
            shell_data.InvoiceDate,
            shell_data.ProductCode,
            shell_data.SiteCode,
            shell_data.Site,
            shell_data.DeliveryDate,
            shell_data.OdometerReading1,
            shell_data.DelcoCustRate1,
            shell_data.CustomerCurrencyName1,
            shell_data.DelcoPumpPrice,
            shell_data.DelcoRebateAmount1,
            shell_data.CustRebateRate1,
            shell_data.PayerDisplayName,
            shell_data.PayerCode,
            shell_data.CardGroupID,
            shell_data.ProductName,
            shell_data.CreditDebitCode,
            shell_data.CorrectionFlag,
            shell_data.DelcoCountryCode,
            shell_data.VATApplicable,
            shell_data.CustomerEuroRate,
            shell_data.DelcoEuroRate,
            shell_data.EuroRebateAmount,
            shell_data.EuroVATAmount,
            shell_data.InvoiceSequenceNo,
            shell_data.ListPriceInTransactionCurrency,
            shell_data.NetEuroAmount,
            shell_data.PumpPriceInTransactionCurrency,
            shell_data.RebateonNetAmountInCustomerCurrency,
            shell_data.RebateonNetAmountInTransactionCurrency,
            shell_data.RebateRate,
            shell_data.UnInvoiceSequenceNo,
            shell_data.NetworkName,
            shell_data.Additional1,
            shell_data.Additional2,
            shell_data.Additional3,
            shell_data.Additional4,
            shell_data.TransactionIdentifier,
            shell_data.TransactionAuthorisationCode,
            shell_data.DelcoName,
            shell_data.SiteGroupid,
            shell_data.SiteGroupName,
            shell_data.Cardtype,
            shell_data.PayerGroup,
            shell_data.PayerNumber,
            shell_data.RefundFlag,
            shell_data.RefundOriginalTransactionId,
            shell_data.PostingDate,
            shell_data.PostingTime ,
            shell_data.CustomerReferenceValue,
            shell_data.CustomerReferenceDescription,
            shell_data.CardPayerAssociationCode,
            shell_data.CardPayerAssociation,
            shell_data.IncomingProductCode,
            shell_data.TransactionProviderId,
            shell_data.TransactionProviderName,
            shell_data.FileName,
            shell_data.FileDate,
            shell_data.AccountGroup3Code,
            shell_data.AccountGroup3Name,
            shell_data.ContractedRebateRate,
            shell_data.IsBestOfPricingApplicable,
            shell_data.CostCenter,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(output)
    return output
