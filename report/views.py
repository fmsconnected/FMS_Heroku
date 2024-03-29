from django.shortcuts import render,HttpResponseRedirect,reverse,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from django.urls import reverse_lazy
import datetime
from openpyxl import Workbook
import datetime
from datetime import date, timedelta
from ajax_select import register
from .models import (
   	vehicle_report,
)
from . forms import (
    reportform,
)
from masterlist.models import EmployeeMasterlist
from vehicle_masterlist.models import VehicleMasterList
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                				)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

def report_new(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    elist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'report_new.html',{'title': 'Report - Vehicle Report', 'elist': elist, 'vlist': vlist})

def report_ongoing(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    ongoing = vehicle_report.objects.filter(Status="Ongoing")
    return render(request, 'report_ongoing.html',{'title': 'Report - Vehicle Report', 'ongoing': ongoing})

def report_completed(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    completed = vehicle_report.objects.filter(Status="Completed")
    return render(request, 'report_completed.html',{'title': 'Report - Vehicle Report', 'completed': completed})

class reportListView(ListView):
	model = vehicle_report
	template_name = 'report_list.html'

def report_submit(request):
	if request.method == 'POST':
		mvar_date = request.POST.get('mvar_date')
		acc_type = request.POST.get('acc_type')
		supp_docs = request.POST.get('supp_docs')
		plate_number = request.POST.get('plate_number')
		v_model = request.POST.get('v_model')
		v_make = request.POST.get('v_make')
		c_sticker = request.POST.get('c_sticker')
		a_emp_id = request.POST.get('a_emp_id')
		a_emp_fname = request.POST.get('a_emp_fname')
		a_emp_lname = request.POST.get('a_emp_lname')
		a_emp_number = request.POST.get('a_emp_number')
		a_emp_company = request.POST.get('a_emp_company')
		a_emp_group = request.POST.get('a_emp_group')
		a_emp_div = request.POST.get('a_emp_div')
		a_emp_dept = request.POST.get('a_emp_dept')
		ai_emp_id = request.POST.get('ai_emp_id')
		ai_emp_fname = request.POST.get('ai_emp_fname')
		ai_emp_lname = request.POST.get('ai_emp_lname')
		inform_inspection = request.POST.get('inform_inspection')
		date_inspection = request.POST.get('date_inspection')
		inspection_remarks = request.POST.get('inspection_remarks')
		date_filed = request.POST.get('date_filed')
		date_received = request.POST.get('date_received')
		date_forward = request.POST.get('date_forward')
		SLA = request.POST.get('MVAR_SLA')
		vr_status = request.POST.get('vr_status')

		saveto_report = vehicle_report(received_date=mvar_date, v_accident_type=acc_type, support_docs=supp_docs, plate_number=plate_number, v_model=v_model, v_make=v_make, 
			cond_sticker=c_sticker, a_employee_id=a_emp_id, a_employee_fname=a_emp_fname, a_employee_lname=a_emp_lname, a_employee_no=a_emp_number, a_employee_company=a_emp_company, a_employee_group=a_emp_group, 
			a_employee_division=a_emp_div, a_employee_dept=a_emp_dept, sup_employee_id=ai_emp_id, sup_employee_fname=ai_emp_fname, sup_employee_lname=ai_emp_lname, inform_assignee=inform_inspection, 
			date_of_inspection=date_inspection, inspection_remarks=inspection_remarks, date_filed_alarm=date_filed, date_cert_received=date_received, date_forwarded=date_forward, MVAR_SLA=SLA, Status=vr_status)
		saveto_report.save()

		def get_success_message(self, cleaned_data):
			print(cleaned_data)
			return "New Vehicle Report Has been Created!"  

		return HttpResponseRedirect('/Report/Report/')                                                                                         

class reportUpdate(SuccessMessageMixin, UpdateView):
	model = vehicle_report
	form_class = reportform
	template_name = 'report_form.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Vehicle Report Updated Successfully!"

class reportDetails(DetailView):
	model = vehicle_report
	template_name = 'vehiclereport_details.html'
	
class reportDeleteView(BSModalDeleteView):
    model = vehicle_report
    template_name = 'report_delete.html'
    success_message = 'Success: Report Vehicle was deleted.'
    success_url = reverse_lazy('report_list')

def reportHistoryView(request):
    if request.method == "GET":
       obj = vehicle_report.history.all()

       return render(request, 'report_history.html', context={'object': obj})

def report_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = vehicle_report.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = vehicle_report.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = vehicle_report.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = vehicle_report.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = vehicle_report.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = vehicle_report.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'reportdeadline.html',{'title':'Report - Report Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})


def report_excel(request):
    report_queryset = vehicle_report.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Motor Vehicle Accident Report.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Motor Vehicle Accident Report'

    columns = [
				'Received Date' ,
			    'Accident Type' ,
			    'Support Documents' ,
			    'Plate Number' ,
			    'Model' ,
			    'Make' ,
			    'Conduction Sticker' ,
			    'Employee Id' ,
			    'Employee Fisrt Name' ,
			    'Employee Last Name' ,
			    'Employee No' ,
			    'Employee Company' ,
			    'Employee Group' ,
			    'Employee Division' ,
			    'Employee Department' ,
			    'Supervisor Employee Id' ,
			    'Supervisor Employee First Name' ,
			    'Supervisor Employee Last Name' ,
			    'Inform Assignee of Vehicle Inspection' ,
			    'Date of Vehicle Inspection' ,
			    'Inspection Remarks' ,
			    'Date Filed Alarm Sheet at PNP TMG' ,
			    'Date Received Certificateof Non-Recovery' ,
			    'Date Forward Documents to Insurance Company' ,
			    'Date Initiated' ,
			    'SLA' ,
			    'Status',

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for report in report_queryset:
        row_num += 1
        row = [
			    report.received_date ,
			    report.v_accident_type ,
			    report.support_docs ,
			    report.plate_number ,
			    report.v_model ,
			    report.v_make ,
			    report.cond_sticker ,
			    report.a_employee_id ,
			    report.a_employee_fname ,
			    report.a_employee_lname ,
			    report.a_employee_no ,
			    report.a_employee_company ,
			    report.a_employee_group ,
			    report.a_employee_division ,
			    report.a_employee_dept ,
			    report.sup_employee_id ,
			    report.sup_employee_fname ,
			    report.sup_employee_lname ,
			    report.inform_assignee ,
			    report.date_of_inspection ,
			    report.inspection_remarks ,
			    report.date_filed_alarm ,
			    report.date_cert_received ,
			    report.date_forwarded ,
			    report.date_initiated ,
			    report.MVAR_SLA ,
			    report.Status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response



        