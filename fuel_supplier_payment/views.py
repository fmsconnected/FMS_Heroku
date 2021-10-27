from django.shortcuts import render,HttpResponseRedirect,reverse,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
from datetime import date, timedelta
from .models import (
    Fuel_supplier
)
from masterlist.models import EmployeeMasterlist
from vehicle_masterlist.models import VehicleMasterList

from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from . forms import (
    FuelsupplierForm
)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )


class FuelDetailView(DetailView):
    model = Fuel_supplier
    template_name = 'fuel_supplierSummary.html'

class FuelListView(ListView):
    model = Fuel_supplier
    template_name = 'fuel_supplierList.html'

class FuelCreateView(SuccessMessageMixin, CreateView):
    model = Fuel_supplier
    form_class = FuelsupplierForm
    template_name = 'fuel_supplier.html'

    def get_success_message(self, cleaned_data):
        print(cleaned_data)
        return "Fuel Supplier Has been Created!"

class FuelUpdateView(SuccessMessageMixin, UpdateView):
    model = Fuel_supplier
    form_class = FuelsupplierForm
    template_name = 'fuel_supplier.html'

    def get_success_message(self, cleaned_data):
        print(cleaned_data)
        return "Fuel Supplier Update Successfully!"

class FuelDeleteView(BSModalDeleteView):
    model = Fuel_supplier
    template_name = 'fuel_supplier_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('Fuel_supplierList')

def FuelHistoryView(request):
    if request.method == "GET":
       obj = Fuel_supplier.history.all()

       return render(request, 'fuel_supplier_history.html', context={'object': obj})

def Fuel_ongoing(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    ongoing = Fuel_supplier.objects.filter(status="Ongoing")

    return render(request, 'fuel_ongoing.html', context={'ongoing': ongoing})

def Fuel_completed(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    completed = Fuel_supplier.objects.filter(status="Completed")

    return render(request, 'fuel_completed.html', context={'completed': completed})

def fuel_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = Fuel_supplier.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = Fuel_supplier.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = Fuel_supplier.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = Fuel_supplier.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = Fuel_supplier.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = Fuel_supplier.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'fueldeadline.html',{'title':'Fuel - Fuel Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})


def fuel_excel(request):
    fuel_queryset = Fuel_supplier.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Fuel Supplier Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Fuel Supplier Payment'

    columns = [
    'Date Received',
    'Fuel Provider',
    'Bill Date',
    'Current Amount',
    'Outstanding Amount',
    'Payee',
    'Attached',
    'Deadline',
    'Date Forwarded',
    'Date Initiated',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for fuel in fuel_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                fuel.SOA_Date_received,
                fuel.Fuel_provider,
                fuel.SOA_billdate,
                fuel.SOA_current_amount,
                fuel.SOA_outstanding_amount,
                fuel.Payee,
                fuel.SOA_attached,
                fuel.Deadline,
                fuel.Date_forwarded,
                fuel.Date_initiated,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
