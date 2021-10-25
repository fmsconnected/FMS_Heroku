from django.shortcuts import render,HttpResponseRedirect,reverse,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
from datetime import date, timedelta
from .models import (
    CarRental
)
from masterlist.models import EmployeeMasterlist
from vehicle_masterlist.models import VehicleMasterList

from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

class CarListView(ListView):
    model = CarRental
    template_name = 'carrental_list.html'

def car_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    dl = CarRental.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = CarRental.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = CarRental.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = CarRental.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = CarRental.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = CarRental.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'cardeadline.html',{'title':'Car - Car Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})

def car_ongoing(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    ongoing = CarRental.objects.filter(status = "Ongoing")
    return  render(request, 'carrental_ongoing.html',{'title':'Car - Car Rental', 'ongoing':ongoing})
def car_completed(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    completed = CarRental.objects.filter(status = "Completed")
    return  render(request, 'carrental_completed.html',{'title':'Car - Car Rental', 'completed':completed})

class CarRentalDetailView(DetailView):
    model = CarRental
    template_name = 'carrental_summary.html'

def Carrentalpayment(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    e_list = EmployeeMasterlist.objects.all()
    v_list = VehicleMasterList.objects.all()
    return render(request, 'car_rental1.html',{'title': 'Car - Car Rental', 'e_list': e_list, 'v_list': v_list})


def Carrental_submit(request):
    if request.method == 'POST':
    #<---Assign Details--->
        Bill_date = request.POST.get('Bdate')
        employee_id = request.POST.get('Eid')
        firstname = request.POST.get('Efm')
        lastname = request.POST.get('Elm')
        Ass_company = request.POST.get('Acom')
        Cost_center = request.POST.get('Acost')
        Date_initiated = request.POST.get('date')
        #<-- other assign-->
        Other_assigneeFM = request.POST.get('Ofname')
        Other_assigneeLM = request.POST.get('Olname')
        Other_cost = request.POST.get('Ocost')
        #<--Vehicle Details-->
        Plate_no = request.POST.get('Pnumber')
        V_brand = request.POST.get('Vbrand')
        V_model = request.POST.get('Vmodel')
        V_make = request.POST.get('Vmake')
        #<--Rental Details-->
        Delivered_V = request.POST.get('Ddelivered')
        S_rental = request.POST.get('Srental')
        E_rental = request.POST.get('Erental')
        #<--Expense Details-->
        R_cost = request.POST.get('Rcost')
        Gas_cost = request.POST.get('Gcost')
        Toll_fee = request.POST.get('Tfee')
        Park_fee = request.POST.get('Pfee')
        Del_fee = request.POST.get('Dfee')
        Driverfee = request.POST.get('Driverfee')
        Meal_cost = request.POST.get('M_cost')
        Other_exp = request.POST.get('Other_exp')
        Total = request.POST.get('Total')
        C_SLA = request.POST.get('SLA')
        car_provider = request.POST.get('car_provider')
        sqa_number = request.POST.get('sqa_number')
        rfp_no2 = request.POST.get('rfp_no2')
        status = request.POST.get('status')

        #############################
        ######Date calculator########
        #############################

        d1 = datetime.datetime.strptime(S_rental, '%Y-%m-%d')
        d2 = datetime.datetime.strptime(E_rental, '%Y-%m-%d')
        Rduration = (d2 - d1)
        
        

        saveto_CRP = CarRental(Bill_date=Bill_date, Employee_id=employee_id, L_name=lastname, F_name=firstname,
            Assignee_company=Ass_company, Cost_center=Cost_center, O_Fname=Other_assigneeFM, O_Lname=Other_assigneeLM,
            O_cost_center=Other_cost, Plate_no=Plate_no, V_brand=V_brand, V_make=V_make,
            D_vehicle=Delivered_V, S_rental=S_rental, E_rental=E_rental, R_duration=Rduration, R_Cost=R_cost,
            G_cost=Gas_cost, T_fee=Toll_fee, P_fee=Park_fee, Del_fee=Del_fee, Dri_fee=Driverfee, M_cost=Meal_cost,
            O_expenses=Other_exp, T_expenses=Total, Date_initiated=Date_initiated, C_SLA=C_SLA, car_provider = car_provider, sqa_number = sqa_number,
            rfp_no2 = rfp_no2, status=status)
        saveto_CRP.save()

        return HttpResponseRedirect('/Car/Car/')

def Carrental_update(request, pk):
    if request.method == 'POST':
    #<---Assign Details--->
        Bill_date = request.POST.get('Bdate')
        employee_id = request.POST.get('Eid')
        firstname = request.POST.get('Efm')
        lastname = request.POST.get('Elm')
        Ass_company = request.POST.get('Acom')
        Cost_center = request.POST.get('Acost')
        #<-- other assign-->
        Other_assigneeFM = request.POST.get('Ofname')
        Other_assigneeLM = request.POST.get('Olname')
        Other_cost = request.POST.get('Ocost')
        #<--Vehicle Details-->
        Plate_no = request.POST.get('Pnumber')
        V_brand = request.POST.get('Vbrand')
        V_make = request.POST.get('Vmake')
        #<--Rental Details-->
        Delivered_V = request.POST.get('Ddelivered')
        S_rental = request.POST.get('Startrental')
        E_rental = request.POST.get('Endrental')
        #<--Expense Details-->
        R_cost = request.POST.get('Rentcost')
        Gas_cost = request.POST.get('Gascost')
        Toll_fee = request.POST.get('Tollfee')
        Park_fee = request.POST.get('Parkingfee')
        Del_fee = request.POST.get('Delfee')
        Driverfee = request.POST.get('Drifee')
        Meal_cost = request.POST.get('Mealcost')
        Other_exp = request.POST.get('Otherexp')
        Total = request.POST.get('TotalExp')
        car_provider = request.POST.get('Vprovider')
        sqa_number = request.POST.get('sqa_number')
        rfp_no2 = request.POST.get('rfp_no2')
        status = request.POST.get('status')

        #############################
        ######Date calculator########
        #############################

        d1 = datetime.datetime.strptime(S_rental, '%Y-%m-%d').date()
        d2 = datetime.datetime.strptime(E_rental, '%Y-%m-%d').date()
        Rduration = (d2 - d1)

        CarRental.objects.filter(id=pk).update(Bill_date=Bill_date,Employee_id=employee_id, L_name=lastname, F_name=firstname,
            Assignee_company=Ass_company, Cost_center=Cost_center, O_Fname=Other_assigneeFM, O_Lname=Other_assigneeLM,
            O_cost_center=Other_cost, Plate_no=Plate_no, V_brand=V_brand, V_make=V_make,
            D_vehicle=Delivered_V, S_rental=S_rental, E_rental=E_rental, R_duration=Rduration, R_Cost=R_cost,
            G_cost=Gas_cost, T_fee=Toll_fee, P_fee=Park_fee, Del_fee=Del_fee, Dri_fee=Driverfee, M_cost=Meal_cost,
            O_expenses=Other_exp, T_expenses=Total,sqa_number=sqa_number,car_provider=car_provider, status=status)

        return HttpResponseRedirect('/Car/Car/')

class carrentalDeleteView(BSModalDeleteView):
    model = CarRental
    template_name = 'carrental_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('carrental_list')

def carrentalHistoryView(request):
    if request.method == "GET":
        obj = CarRental.history.all()
        return render(request, 'carrental_history.html', context={'object': obj})

def rent(request):
    emp = EmployeeMasterlist.objects.all()
    vechicle = VehicleMasterList.objects.all()
    return render(request, 'payment/car_rental.html', {'title': 'CarRental - Create New Car Rental Request', 'emp': emp, 
    'vechicle': vechicle})


def car_excel(request):
    car_queryset = CarRental.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Car Rental Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Car Rental Payment'

    columns = [
            'Bill date',
            'Employee Id',
            'Last Name',
            'First Name',
            'Company',
            'Cost Center',
            'Date Created',
            'Other Assignee First Name',
            'Other Assignee Last Name',
            'Other Cost Center',
            'Plate No',
            'Model',
            'Brand',
            'Make',
            'Date Rental',
            'Start Date Rental',
            'End Date Rental',
            'Rent Duration',
            'Rent Cost',
            'Gasoline Cost',
            'Toll Fee',
            'Parking Fee',
            'Delivery Fee',
            'Driver Fee',
            'Meal Cost',
            'Other Expense',
            'Vat',
            'Total Expense',
            'car_provider',
            'sqa_number',
            'rfp_no2',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for car in car_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                car.Bill_date,
                car.Employee_id,
                car.L_name,
                car.F_name,
                car.Assignee_company,
                car.Cost_center,
                car.Date_initiated,
                car.O_Fname,
                car.O_Lname,
                car.O_cost_center,
                car.Plate_no,
                car.V_model,
                car.V_brand,
                car.V_make,
                car.D_vehicle,
                car.S_rental,
                car.E_rental,
                car.R_duration,
                car.R_Cost,
                car.G_cost,
                car.T_fee,
                car.P_fee,
                car.Del_fee,
                car.Dri_fee,
                car.M_cost,
                car.O_expenses,
                car.VAT,
                car.T_expenses,
                car.car_provider,
                car.sqa_number,
                car.rfp_no2,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


## Car rental Daily report Details
def car_report_details(request):
    date = datetime.datetime.today()
    carreport_SAFARI = CarRental.objects.filter(car_provider="SAFARI", sqa_number="").count()
    carreport_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND", sqa_number="").count()
    carreport_TG = CarRental.objects.filter(car_provider="Tiger City", sqa_number="").count()
    carreport_ORIX = CarRental.objects.filter(car_provider="ORIX", sqa_number="").count()
    
    vreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX", invoice_date="").count()
    vreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI", invoice_date="").count()
    vreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND", invoice_date="").count()
    vreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City", invoice_date="").count()
    
    processed_ORIX = CarRental.objects.filter(car_provider="ORIX").exclude(sqa_number='').count()
    processed_SAFARI = CarRental.objects.filter(car_provider="SAFARI").exclude(sqa_number='').count()
    processed_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND").exclude(sqa_number='').count()
    processed_TG = CarRental.objects.filter(car_provider="Tiger City").exclude(sqa_number='').count()

    proreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX").exclude(invoice_date="").count()
    proreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI").exclude(invoice_date="").count()
    proreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND").exclude(invoice_date="").count()
    proreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City").exclude(invoice_date="").count()
    
    shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL", SOA_billdate="").count()
    pro_shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL").exclude(SOA_billdate="").count()
    petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation", SOA_billdate="").count()
    pro_petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation").exclude(SOA_billdate="").count()

    orix = carreport_ORIX + vreport_ORIX
    safari = carreport_SAFARI + vreport_SAFARI
    diamond = carreport_DIAMOND + vreport_DIAMOND
    tg = carreport_TG + vreport_TG

    tp_orix = processed_ORIX + proreport_ORIX
    tp_safari = processed_SAFARI + proreport_SAFARI
    tp_diamond = processed_DIAMOND + proreport_DIAMOND
    tp_tg = processed_TG + proreport_TG

    return render(request, 'payment_report.html',{'title':'Payment','date':date,'orix':orix, 'safari':safari, 'diamond':diamond, 'tg':tg,'tp_orix':tp_orix, 'tp_safari':tp_safari, 'tp_diamond':tp_diamond, 'tp_tg':tp_tg})

# Car rental Daily Report
def car_report(request):
    date = datetime.datetime.today()
    yr = datetime.datetime.now().year
    months = ['zero','January','February','March','April','May','June','July','August','September','October','November','December']
    month = months[date.month]
    datem = datetime.datetime(date.year, date.month, 1)
    carreport_SAFARI = CarRental.objects.filter(car_provider="SAFARI", sqa_number="").count()
    carreport_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND", sqa_number="").count()
    carreport_TG = CarRental.objects.filter(car_provider="Tiger City", sqa_number="").count()
    carreport_ORIX = CarRental.objects.filter(car_provider="ORIX", sqa_number="").count()
    
    vreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX", invoice_date="").count()
    vreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI", invoice_date="").count()
    vreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND", invoice_date="").count()
    vreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City", invoice_date="").count()
    
    processed_ORIX = CarRental.objects.filter(car_provider="ORIX").exclude(sqa_number='').count()
    processed_SAFARI = CarRental.objects.filter(car_provider="SAFARI").exclude(sqa_number='').count()
    processed_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND").exclude(sqa_number='').count()
    processed_TG = CarRental.objects.filter(car_provider="Tiger City").exclude(sqa_number='').count()

    proreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX").exclude(invoice_date="").count()
    proreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI").exclude(invoice_date="").count()
    proreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND").exclude(invoice_date="").count()
    proreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City").exclude(invoice_date="").count()
    
    shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL", SOA_billdate="").count()
    pro_shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL").exclude(SOA_billdate="").count()
    petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation", SOA_billdate="").count()
    pro_petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation").exclude(SOA_billdate="").count()

    orix = carreport_ORIX + vreport_ORIX
    safari = carreport_SAFARI + vreport_SAFARI
    diamond = carreport_DIAMOND + vreport_DIAMOND
    tg = carreport_TG + vreport_TG

    tp_orix = processed_ORIX + proreport_ORIX
    tp_safari = processed_SAFARI + proreport_SAFARI
    tp_diamond = processed_DIAMOND + proreport_DIAMOND
    tp_tg = processed_TG + proreport_TG


    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Insurance Daily Report.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Insurance Daily Report"
    ws['A1'].value = "Personel"
    ws['B1'].value = "Janine Manzo"
    ws['B2'].value = datem
    ws['A2'].value = "Date"
    ws.append(['Vendor','Total Unpaid SOA','Total Processed','For this Month Total SOA Processed'])
    ws['A4'].value = "ORIX"
    ws['A5'].value = "DIAMOND"
    ws['A6'].value = "PETRON"
    ws['A7'].value = "SHELL"
    ws['A8'].value = "GR8"
    ws['A9'].value = "JXM"
    ws['A10'].value = "SAFARI (Car Rental)"
    ws['A11'].value = "Tiger City (Car Rental)"
    ws['A12'].value = "Created Work Order"
    ws['A13'].value = "TopServe (FLEET Driver)"
    ws['A14'].value = "CARPLAN 95%"
    ws['A15'].value = "CARPLAN 5%"
    #data
    ws['B4'].value = orix
    ws['B5'].value = diamond
    ws['B6'].value = petron
    ws['B7'].value = shell
    ws['B8'].value = ""
    ws['B9'].value = ""
    ws['B10'].value = safari
    ws['B11'].value = tg
    ws['B12'].value = ""
    ws['B13'].value = ""
    ws['B14'].value = ""
    ws['B15'].value = ""

    ws['C4'].value = tp_orix
    ws['C5'].value = tp_diamond
    ws['C6'].value = pro_petron
    ws['C7'].value = pro_shell
    ws['C8'].value = ""
    ws['C9'].value = ""
    ws['C10'].value = tp_safari
    ws['C11'].value = tp_tg
    ws['C12'].value = ""
    ws['C13'].value = ""
    ws['C14'].value = ""
    ws['C15'].value = ""

   
    wb.save(output)
    return output
