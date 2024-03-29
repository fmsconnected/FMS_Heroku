from django.shortcuts import render, HttpResponseRedirect, HttpResponse, reverse
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
from django.core import serializers
import datetime
from django.utils import formats
from datetime import date, timedelta
from rest_framework import viewsets
from rest_framework.response import Response
from django.shortcuts import render
from django.db.models import Q, Count
from bootstrap_modal_forms.generic import BSModalDeleteView
# User
from django.core.exceptions import PermissionDenied
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, HttpResponseRedirect
from django.contrib.auth.models import User
from django.contrib.auth.mixins import LoginRequiredMixin
from vehicle_masterlist.models import VehicleMasterList

from django.views.generic import (
    ListView,
    CreateView,
    DetailView,
    UpdateView,
)

from . forms import (
    CS_form,
    CS_formupdate
)

from .serializers import (
    CS_log_serializer
)
from .models import (
    CS_log
)
from request.models import Vehicle_Repair
from registration.models import Registration

def in_group(user):
    if user.groups.filter(name="CustomerLog").exists():
        return True
    else:
        raise PermissionDenied


# class CSCreateView(CreateView):
#     # @method_decorator(user_passes_test(in_group))
#     def dispatch(self, *args, **kwargs):
#         return super().dispatch(*args, **kwargs)
#     model = CS_log
#     form_class = CS_form
#     template_name = 'CS/CS_create.html'

def CSCreateView(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    customer = CS_log.objects.all()
    return render(request, 'CS/CS_create.html',{'Title':'Customer Car Log','customer':customer})

def CSsubmit(request):
    
    if request.method == 'POST':

        Date_received = request.POST.get('date_received')
        Fleet_member = request.POST.get('Fleet_member')
        Client_name = request.POST.get('Client_name')
        Email = request.POST.get('Email')
        Mobile_no = request.POST.get('Mobile_no')
        Transaction_type = request.POST.get('Transaction_type')
        Plate_no = request.POST.get('Plate_no')
        Problem = request.POST.get('Problem')
        Date_resolved = request.POST.get('Date_resolved')
        Action_taken = request.POST.get('Action_taken')

        
        Date_resolved_inital = datetime.datetime.today() + timedelta(days=5)
        saveto_customer = CS_log(Date_received=Date_received, Fleet_member=Fleet_member, Client_name=Client_name,
                                            Email=Email, Mobile_no=Mobile_no, Transaction_type=Transaction_type, Plate_no=Plate_no, Problem=Problem,
                                            Date_resolved_inital=Date_resolved_inital, Action_taken=Action_taken)
        saveto_customer.save()
    return HttpResponseRedirect('/Customer/')


class CSDetails(DetailView):
    # @ method_decorator(user_passes_test(in_group))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    model = CS_log
    template_name = 'CS/CS_details.html'


def CSListView(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    dl2 = CS_log.objects.filter(Q(Ageing='') & Q(Date_received__date=datetime.datetime.today(
    ) - timedelta(days=4)) | Q(Date_received__date=datetime.datetime.today() - timedelta(days=5)) | Q(Date_received__date=datetime.datetime.today() - timedelta(days=3)) | Q(Date_received__date=datetime.datetime.today() - timedelta(days=2)))
    # dl1 = CS_log.objects.filter(
    #     Date_received__date=datetime.datetime.today() - timedelta(days=5))
    ccl_count = dl2.aggregate(counted=Count('id'))[
        'counted']

    object_list = CS_log.objects.all()
    return render(request, 'CS/CS_list.html', {'Title': 'Customer Care Log', 'object_list': object_list, 'ccl_count': ccl_count})


def CSpending(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    pending = CS_log.objects.filter(Ageing="")
    pendingcount = pending.aggregate(counted=Count('id'))[
        'counted']
    return render(request, 'CS/CS_pending.html', {'Title': 'Customer Care Log', 'pending': pending, 'pendingcount': pendingcount})

# @ user_passes_test(in_group)
def CSUpdate(request, pk):

    if request.method == 'POST':

        Date_received = request.POST.get('date_received')
        Fleet_member = request.POST.get('Fleet_member')
        Client_name = request.POST.get('Client_name')
        Email = request.POST.get('Email')
        Mobile_no = request.POST.get('Mobile_no')
        Transaction_type = request.POST.get('Transaction_type')
        Plate_no = request.POST.get('Plate_no')
        Problem = request.POST.get('Problem')
        Date_resolved = request.POST.get('Date_resolved')
        Action_taken = request.POST.get('Action_taken')

        if Date_resolved == '':
            Date_resolved = ''
            ageing = ''
        else:
            Date_received = datetime.datetime.strptime(
                Date_received, '%Y-%m-%d').date()
            Date_resolved = datetime.datetime.strptime(
                Date_resolved, '%Y-%m-%d').date()
            ageing = (Date_resolved - Date_received)

        CS_log.objects.filter(id=pk).update(Date_received=Date_received, Fleet_member=Fleet_member, Client_name=Client_name,
                                            Email=Email, Mobile_no=Mobile_no, Transaction_type=Transaction_type, Plate_no=Plate_no, Problem=Problem,
                                            Date_resolved=Date_resolved, Action_taken=Action_taken, Ageing=ageing)
    return HttpResponseRedirect('/Customer/')


class CSDeleteView(BSModalDeleteView):
    @ method_decorator(user_passes_test(in_group))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    model = CS_log
    template_name = 'CS/CS_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('CS_List')


def CS_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    dl = CS_log.objects.filter(
        Ageing='', Date_received__date=datetime.datetime.today() - timedelta(days=4))
    dl2 = CS_log.objects.filter(
        Ageing='', Date_received__date=datetime.datetime.today() - timedelta(days=5))
    dl3 = CS_log.objects.filter(
        Ageing='', Date_received__date=datetime.datetime.today() - timedelta(days=6))
    dl4 = CS_log.objects.filter(
        Ageing='', Date_received__date=datetime.datetime.today() - timedelta(days=7))
    dl5 = CS_log.objects.filter(
        Ageing='', Date_received__date=datetime.datetime.today() - timedelta(days=8))
    # dl6 = CS_log.objects.filter(
    #     Ageing='', Date_received__date=datetime.datetime.today())
    dl7 = CS_log.objects.filter(
        Ageing='', Date_received__date=datetime.datetime.today() - timedelta(days=3))
    dl8 = CS_log.objects.filter(
        Ageing='', Date_received__date=datetime.datetime.today() - timedelta(days=2))

    return render(request, 'CS/CS_deadline.html', {'title': 'Customer Care Log', 'dl': dl, 'dl2': dl2, 'dl3': dl3, 'dl4': dl4, 'dl5': dl5, 'dl7': dl7, 'dl8': dl8})


# @ user_passes_test(in_group)
def customer_log_excel(request):
    customer_queryset = CS_log.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Customer Care Log.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Customer Care Log'

    columns = [
        'Ticket Number',
        'Date Received',
        'Fleet Member',
        'Client Name',
        'Email',
        'Mobile Number',
        'Transaction Type',
        'Plate Number',
        'Problem',
        'Date Resolved',
        'Action Taken'
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for cc in customer_queryset:
        row_num += 1
        row = [
            cc.Activity_id,
            cc.Date_received,
            cc.Fleet_member,
            cc.Client_name,
            cc.Email,
            cc.Mobile_no,
            cc.Transaction_type,
            cc.Plate_no,
            cc.Problem,
            cc.Date_resolved,
            cc.Action_taken
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def request_email_log(request):
    email_log = Vehicle_Repair.objects.filter(sent_email="Yes")
    return render (request, 'email_log/request_email_log.html',{'email_log':email_log})

def registration_email_log(request):
    email_log = VehicleMasterList.objects.filter(Status="Yes")
    return render (request, 'email_log/registration_email_log.html',{'email_log':email_log})


