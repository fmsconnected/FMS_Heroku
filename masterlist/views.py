from django.shortcuts import render,HttpResponseRedirect,HttpResponse,reverse
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
# from django.core.urlresolvers import reverse_lazy
from django.views.generic import (
     ListView,
     CreateView,
     DetailView,
     UpdateView,
 )
from .models import (
    EmployeeMasterlist
    )
from . forms import (
    EmpMasterlistForm,
    )
from bootstrap_modal_forms.generic import BSModalDeleteView

from rest_framework import viewsets
from rest_framework.response import Response
# from .models import VehicleMasterList,
from .serializers import (
    EmployeeSerializer
    )

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render


# def Vmastertables(request):
#     return render(request, 'vehicleMasterlist/vehicleMasterlist.html')


# class vehicleViewSet(viewsets.ModelViewSet):
#     queryset = VehicleMasterList.objects.all().order_by('id')
#     serializer_class = vehicleSerializer


# def VmasterlistCreate(request):
#     if request.method == 'POST':
#         plate = request.POST.get('plate_no')
#         cs = request.POST.get('cs')
#         cr_name = request.POST.get('cr_name')
#         model = request.POST.get('model')
#         brand = request.POST.get('brand')
#         vmake = request.POST.get('vmake')
#         eng_no = request.POST.get('eng_no')
#         chassis_no = request.POST.get('chassis_no')
#         mvfile = request.POST.get('mvfile')
#         vtype = request.POST.get('vtype')
#         vcat = request.POST.get('vcat')
#         emp_id = request.POST.get('emp_id')
#         band = request.POST.get('band')
#         benefit = request.POST.get('benefit')
#         cost = request.POST.get('cost')
#         group = request.POST.get('group')
#         div = request.POST.get('div')
#         dept = request.POST.get('dept')
#         sec = request.POST.get('sec')
#         is_emp = request.POST.get('is_emp')
#         IS_NAME = request.POST.get('is_fname')
#         # is_lname = request.POST.get('is_fname')
#         loc = request.POST.get('loc')
#         aqui_date = request.POST.get('aqui_date')
#         aqui_cost = request.POST.get('aqui_cost')
#         asset = request.POST.get('asset')
#         po_no = request.POST.get('po_no')
#         sap_pr = request.POST.get('sap_pr')
#         ivn_no = request.POST.get('ivn_no')
#         mathdoc = request.POST.get('mathdoc')
#         eq_no = request.POST.get('eq_no')
#         dealer = request.POST.get('dealer')
#         dealer_name = request.POST.get('dealer_name')
#         plate_date = request.POST.get('plate_date')
#         or_date = request.POST.get('or_date')
#         remarks = request.POST.get('remarks')
#         status = request.POST.get('status')
#         l_remark = request.POST.get('l_remark')
#         l_remark = request.POST.get('l_remark')
#         vehicle_status = request.POST.get('vstatus')
        
#         if or_date == '':
#             or_date = None
#         else:
#             or_date = datetime.datetime.strptime(or_date,'%Y-%m-%d')

#         if aqui_date == '':
#             aqui_date = None
#         else:
#             aqui_date = datetime.datetime.strptime(aqui_date,'%Y-%m-%d')

#         if plate_date == '':
#             plate_date = None
#         else:
#             plate_date = datetime.datetime.strptime(plate_date,'%Y-%m-%d')


#         employee_list = EmployeeMasterlist.objects.all()
#         for e_id in employee_list:
#             if e_id.Employee_Id == emp_id:
#                 emp_save = e_id
            
#         if emp_id == '':
#             emp_save = None
            
#         reg = ''
#         endplate = ''
#         if plate != '':
#             endplate = int(plate[-1])
#             if endplate == 1:
#                 reg = 'JAN'
#             elif endplate == 2:
#                 reg = 'FEB'
#             elif endplate == 3:
#                 reg = 'MAR'
#             elif endplate == 4:
#                 reg = 'APR'
#             elif endplate == 5:
#                 reg = 'MAY'
#             elif endplate == 6:
#                 reg = 'JUN'
#             elif endplate == 7:
#                 reg = 'JUL'
#             elif endplate == 8:
#                 reg = 'AUG'
#             elif endplate == 9:
#                 reg = 'SEP'
#             elif endplate == 0:
#                 reg = 'OCT'

#         saveto_end = VehicleMasterList(PLATE_NO=plate, CS_NO=cs, CR_NAME=cr_name, MODEL=model, BRAND=brand,PLATE_ENDING=endplate, REGISTRATION_MONTH=reg,
#             VEHICLE_MAKE=vmake, ENGINE_NO=eng_no, CHASSIS_NO=chassis_no, MV_FILE_NO=mvfile, VEHICLE_TYPE=vtype, VEHICLE_CATEGORY=vcat,
#             Employee=emp_save, BAND_LEVEL=band, BENEFIT_GROUP=benefit, COST_CENTER=cost, GROUP=group, DIVISION=div,
#             DEPARTMENT=dept, SECTION=sec, IS_ID=is_emp, IS_NAME=IS_NAME, LOCATION=loc,
#             ACQ_DATE=aqui_date, ACQ_COST=aqui_cost, ASSET_NO=asset, PO_NO=po_no, PLATE_NUMBER_RELEASE_DATE=plate_date, ORIGINAL_OR_DATE=or_date,EQUIPMENT_NO=eq_no,
#             SAP_PR=sap_pr,Vehicle_IVN_no=ivn_no,Unit_MATDOC=mathdoc,dealer=dealer, dealer_name=dealer_name, Remarks=remarks, Status=status, leasing_remark=l_remark,
#             vehicle_status = vehicle_status
#             )
#         saveto_end.save()

#     return HttpResponseRedirect('/Masterlist/VehicleMasterlist/')

# def vehicle(request):
#     elist = EmployeeMasterlist.objects.all()
#     return render(request, 'vehicleMasterlist/vmasterlist.html', {'Title': 'Vehicle Masterlist','elist':elist})

# # class vehicleMasterListView(ListView):
# #     model = VehicleMasterList
# #     template_name = 'vehicleMasterlist/vm.html'
        
# class vehicleMasterDetails(DetailView):
#     model = VehicleMasterList
#     template_name = 'vehicleMasterlist/vehicleMasterlist_details.html'

# class vehicleMasterUpdate(UpdateView):
#     model = VehicleMasterList
#     form_class = Vmasterlist
#     template_name = 'vehicleMasterlist/vehicleMasterlist_form.html'

# class releaseUpdate(UpdateView):
#     model = VehicleMasterList
#     form_class = Vmaster
#     template_name = 'registration/release_date.html'

# class vehicleMasterlistDeleteView(BSModalDeleteView):
#     model = VehicleMasterList
#     template_name = 'vehicleMasterlist/vehicleMasterlist_delete.html'
#     success_message = 'Success: Item was deleted.'
#     success_url = reverse_lazy('vehicle-list')

# def vehicleMasterlistHistoryView(request):
#     if request.method == "GET":
#        obj = VehicleMasterList.history.all()

#        return render(request, 'vehicleMasterlist/vehicleMasterlist_history.html', context={'object': obj})

# def vehicle_masterlist_active(request):
#     context = {
#             'vehicle_list_active': VehicleMasterList.objects.filter(vehicle_status__contains='Active')
#         }

#     return render(request, 'vehicleMasterlist/vehicle_masterlist_active.html', context)

# def vehicle_masterlist_solved(request):
#     context = {
#             'vehicle_list_solved': VehicleMasterList.objects.filter(vehicle_status__contains='Sold')
#         }

#     return render(request, 'vehicleMasterlist/vehicle_masterlist_solved.html', context)

# def vehicle_masterlist_trans(request):
#     context = {
#             'vehicle_list_trans': VehicleMasterList.objects.filter(vehicle_status__contains='Transferred')
#         }

#     return render(request, 'vehicleMasterlist/vehicle_masterlist_trans.html', context)



class employeeCreateView(CreateView):
    model = EmployeeMasterlist
    form_class = EmpMasterlistForm
    template_name = 'employeeMasterlist/employeeMasterlist_form.html'

class employeeViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMasterlist.objects.all().order_by('id')
    serializer_class = EmployeeSerializer

# class employeeListView(ListView):
#     model = EmployeeMasterlist
#     template_name = 'employeeMasterlist/employeeMasterlist.html'

def empmastertables(request):
    return render(request, 'employeeMasterlist/emp.html')

class employeeDetailView(DetailView):
    model = EmployeeMasterlist
    template_name = 'employeeMasterlist/employeeMasterlist_details.html'

class employeeUpdateView(UpdateView):
    model = EmployeeMasterlist
    form_class = EmpMasterlistForm
    template_name = 'employeeMasterlist/employeeMasterlist_form.html'

class employeeMasterlistDeleteView(BSModalDeleteView):
    model = EmployeeMasterlist
    template_name = 'employeeMasterlist/employeeMasterlist_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('employee-list')

def employeeMasterlistHistoryView(request):
    if request.method == "GET":
       obj = EmployeeMasterlist.history.all()

       return render(request, 'employeeMasterlist/employeeMasterlist_history.html', context={'object': obj})


 #######################################  
##############  Leasing  ################
 #######################################


# def leasing(request):
#     return render(request, 'leasing/leasing_list.html')


# class leasingViewSet(viewsets.ModelViewSet):
#     queryset = Leasing.objects.all().order_by('id')
#     serializer_class = leasingSerializer


# class leasingCreateView(CreateView):
#     model = Leasing
#     form_class = leasing_form
#     template_name = 'leasing/leasing_form.html'


# class leasingUpdateView(UpdateView):
#     model = Leasing
#     form_class = leasing_form
#     template_name = 'leasing/leasing_form.html'


# class leasingDetailView(DetailView):
#     model = Leasing
#     template_name = 'leasing/leasing_details.html'


# def leasingHistoryView(request):
#     if request.method == "GET":
#        obj = Leasing.history.all()

#        return render(request, 'leasing/leasing_history.html', context={'object': obj})


# class vreg_details(DetailView):
#     model = VehicleMasterList
#     template_name = 'vehicleMasterlist/vreg_details.html'
    

# def vehicle_bayan(request):
#     context = {
#             'bayan_list': VehicleMasterList.objects.filter(CR_NAME__contains="BAYANTEL")
#         }

#     return render(request, 'vehicleMasterlist/vehicle_bayan.html', context)
# def vehicle_telicphil(request):
#     context = {
#             'teli_list': VehicleMasterList.objects.filter(CR_NAME__contains="TELICPHIL")
#         }

#     return render(request, 'vehicleMasterlist/vehicle_telicphil.html', context)


# def vehicle_leasing(request):
#     context = {
#             'leasing_list': VehicleMasterList.objects.filter(leasing_remark__isnull=False)
#         }

#     return render(request, 'vehicleMasterlist/vehicle_leasing.html', context)


# ## Leasing ##
# def vehicle_leasing_active(request):
#     context = {
#             'leasing_list_active': Leasing.objects.filter(vleasing_status__contains='Active')
#         }

#     return render(request, 'leasing/leasing_active.html', context)

# def vehicle_leasing_solved(request):
#     context = {
#             'leasing_list_solved': Leasing.objects.filter(vleasing_status__contains='Sold')
#         }

#     return render(request, 'leasing/leasing_solved.html', context)

# def vehicle_leasing_trans(request):
#     context = {
#             'leasing_list_trans': Leasing.objects.filter(vleasing_status__contains='Transferred')
#         }

#     return render(request, 'leasing/leasing_trans.html', context)

# def vehicle_excel_leasing(request):
#     v_queryset = VehicleMasterList.objects.filter(leasing_remark__isnull=False)   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Leasing.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist Leasing'

#     columns = [
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'VEHICLE_CATEGORY',
#             'Employee',
#             'BAND_LEVEL',
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_FIRST_NAME',
#             'IS_LAST_NAME',
#             'LOCATION',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'PO_NO',
#             'EQUIPMENT_NO',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'REGISTRATION_MONTH',
#             'ORIGINAL_OR_DATE',
#             'PLATE_NUMBER_RELEASE_DATE',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
#         # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.Employee,
#                 vehicle.BAND_LEVEL,
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_FIRST_NAME,
#                 vehicle.IS_LAST_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.PO_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response
# def leasing_export(request):
#     leasing_queryset = Leasing.objects.all()
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Leasing Masterlist'

#     columns = [

#             'Activity_Id',
#             'PLATE_NUMBER',
#             'CS_NO',
#             'COMPANY',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'VEHICLE_TYPE',
#             'LAST_NAME_ASSIGNEE',
#             'FIRST_NAME_ASSIGNEE',
#             'VEHICLE_CATEGORY',
#             'COST_CENTER',
#             'ID_NUMBER',
#             'BAND',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_EMPLOYEE_ID',
#             'IS_NAME',
#             'IS_FIRSTNAME',
#             'LOCATION',
#             'AREA',
#             'ACQUISITION_DATE',
#             'remarks',
#             'acquisition_cost',
#             'months_36',
#             'amount1',
#             'date_in_1',
#             'date_out_1',
#             'months_24',
#             'amount_Vat_EX',
#             'date_in_2',
#             'date_out_2',
#             'extension',
#             'amount2',
#             'date_in_3',
#             'date_out_3',
#             'chasis_no',
#             'engine_no',
#             'CONTRACT_NUMBER',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for leasing in leasing_queryset:
#         row_num += 1
#         row = [
#             leasing.Activity_Id,
#             leasing.PLATE_NUMBER,
#             leasing.CS_NO,
#             leasing.COMPANY,
#             leasing.MODEL,
#             leasing.BRAND,
#             leasing.VEHICLE_MAKE,
#             leasing.VEHICLE_TYPE,
#             leasing.LAST_NAME_ASSIGNEE,
#             leasing.FIRST_NAME_ASSIGNEE,
#             leasing.VEHICLE_CATEGORY,
#             leasing.COST_CENTER,
#             leasing.ID_NUMBER,
#             leasing.BAND,
#             leasing.GROUP,
#             leasing.DIVISION,
#             leasing.DEPARTMENT,
#             leasing.SECTION,
#             leasing.IS_EMPLOYEE_ID,
#             leasing.IS_LASTNAME,
#             leasing.IS_FIRSTNAME,
#             leasing.LOCATION,
#             leasing.AREA,
#             leasing.ACQUISITION_DATE,
#             leasing.remarks,
#             leasing.acquisition_cost,
#             leasing.months_36,
#             leasing.amount1,
#             leasing.date_in_1,
#             leasing.date_out_1,
#             leasing.months_24,
#             leasing.amount_Vat_EX,
#             leasing.date_in_2,
#             leasing.date_out_2,
#             leasing.extension,
#             leasing.amount2,
#             leasing.date_in_3,
#             leasing.date_out_3,
#             leasing.chasis_no,
#             leasing.engine_no,
#             leasing.CONTRACT_NUMBER,
#         ]

#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response

# def leasing_active_export(request):
#     leasing_queryset = Leasing.objects.filter(vleasing_status='Active')
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist Active.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Leasing Masterlist Active'

#     columns = [

#             'Activity_Id',
#             'PLATE_NUMBER',
#             'CS_NO',
#             'COMPANY',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'VEHICLE_TYPE',
#             'LAST_NAME_ASSIGNEE',
#             'FIRST_NAME_ASSIGNEE',
#             'VEHICLE_CATEGORY',
#             'COST_CENTER',
#             'ID_NUMBER',
#             'BAND',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_EMPLOYEE_ID',
#             'IS_NAME',
#             'IS_FIRSTNAME',
#             'LOCATION',
#             'AREA',
#             'ACQUISITION_DATE',
#             'remarks',
#             'acquisition_cost',
#             'months_36',
#             'amount1',
#             'date_in_1',
#             'date_out_1',
#             'months_24',
#             'amount_Vat_EX',
#             'date_in_2',
#             'date_out_2',
#             'extension',
#             'amount2',
#             'date_in_3',
#             'date_out_3',
#             'chasis_no',
#             'engine_no',
#             'CONTRACT_NUMBER',
#             'Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for leasing in leasing_queryset:
#         row_num += 1
#         row = [
#             leasing.Activity_Id,
#             leasing.PLATE_NUMBER,
#             leasing.CS_NO,
#             leasing.COMPANY,
#             leasing.MODEL,
#             leasing.BRAND,
#             leasing.VEHICLE_MAKE,
#             leasing.VEHICLE_TYPE,
#             leasing.LAST_NAME_ASSIGNEE,
#             leasing.FIRST_NAME_ASSIGNEE,
#             leasing.VEHICLE_CATEGORY,
#             leasing.COST_CENTER,
#             leasing.ID_NUMBER,
#             leasing.BAND,
#             leasing.GROUP,
#             leasing.DIVISION,
#             leasing.DEPARTMENT,
#             leasing.SECTION,
#             leasing.IS_EMPLOYEE_ID,
#             leasing.IS_LASTNAME,
#             leasing.IS_FIRSTNAME,
#             leasing.LOCATION,
#             leasing.AREA,
#             leasing.ACQUISITION_DATE,
#             leasing.remarks,
#             leasing.acquisition_cost,
#             leasing.months_36,
#             leasing.amount1,
#             leasing.date_in_1,
#             leasing.date_out_1,
#             leasing.months_24,
#             leasing.amount_Vat_EX,
#             leasing.date_in_2,
#             leasing.date_out_2,
#             leasing.extension,
#             leasing.amount2,
#             leasing.date_in_3,
#             leasing.date_out_3,
#             leasing.chasis_no,
#             leasing.engine_no,
#             leasing.CONTRACT_NUMBER,
#             leasing.vleasing_status,
#         ]

#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response

# def leasing_solved_export(request):
#     leasing_queryset = Leasing.objects.filter(vleasing_status='Sold')
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist Sold.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Leasing Masterlist Sold'

#     columns = [

#             'Activity_Id',
#             'PLATE_NUMBER',
#             'CS_NO',
#             'COMPANY',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'VEHICLE_TYPE',
#             'LAST_NAME_ASSIGNEE',
#             'FIRST_NAME_ASSIGNEE',
#             'VEHICLE_CATEGORY',
#             'COST_CENTER',
#             'ID_NUMBER',
#             'BAND',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_EMPLOYEE_ID',
#             'IS_NAME',
#             'IS_FIRSTNAME',
#             'LOCATION',
#             'AREA',
#             'ACQUISITION_DATE',
#             'remarks',
#             'acquisition_cost',
#             'months_36',
#             'amount1',
#             'date_in_1',
#             'date_out_1',
#             'months_24',
#             'amount_Vat_EX',
#             'date_in_2',
#             'date_out_2',
#             'extension',
#             'amount2',
#             'date_in_3',
#             'date_out_3',
#             'chasis_no',
#             'engine_no',
#             'CONTRACT_NUMBER',
#             'Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for leasing in leasing_queryset:
#         row_num += 1
#         row = [
#             leasing.Activity_Id,
#             leasing.PLATE_NUMBER,
#             leasing.CS_NO,
#             leasing.COMPANY,
#             leasing.MODEL,
#             leasing.BRAND,
#             leasing.VEHICLE_MAKE,
#             leasing.VEHICLE_TYPE,
#             leasing.LAST_NAME_ASSIGNEE,
#             leasing.FIRST_NAME_ASSIGNEE,
#             leasing.VEHICLE_CATEGORY,
#             leasing.COST_CENTER,
#             leasing.ID_NUMBER,
#             leasing.BAND,
#             leasing.GROUP,
#             leasing.DIVISION,
#             leasing.DEPARTMENT,
#             leasing.SECTION,
#             leasing.IS_EMPLOYEE_ID,
#             leasing.IS_LASTNAME,
#             leasing.IS_FIRSTNAME,
#             leasing.LOCATION,
#             leasing.AREA,
#             leasing.ACQUISITION_DATE,
#             leasing.remarks,
#             leasing.acquisition_cost,
#             leasing.months_36,
#             leasing.amount1,
#             leasing.date_in_1,
#             leasing.date_out_1,
#             leasing.months_24,
#             leasing.amount_Vat_EX,
#             leasing.date_in_2,
#             leasing.date_out_2,
#             leasing.extension,
#             leasing.amount2,
#             leasing.date_in_3,
#             leasing.date_out_3,
#             leasing.chasis_no,
#             leasing.engine_no,
#             leasing.CONTRACT_NUMBER,
#             leasing.vleasing_status,
#         ]

#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response

# def leasing_trans_export(request):
#     leasing_queryset = Leasing.objects.filter(vleasing_status='Transferred')
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist Transferred.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Leasing Masterlist Transferred'

#     columns = [

#             'Activity_Id',
#             'PLATE_NUMBER',
#             'CS_NO',
#             'COMPANY',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'VEHICLE_TYPE',
#             'LAST_NAME_ASSIGNEE',
#             'FIRST_NAME_ASSIGNEE',
#             'VEHICLE_CATEGORY',
#             'COST_CENTER',
#             'ID_NUMBER',
#             'BAND',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_EMPLOYEE_ID',
#             'IS_NAME',
#             'IS_FIRSTNAME',
#             'LOCATION',
#             'AREA',
#             'ACQUISITION_DATE',
#             'remarks',
#             'acquisition_cost',
#             'months_36',
#             'amount1',
#             'date_in_1',
#             'date_out_1',
#             'months_24',
#             'amount_Vat_EX',
#             'date_in_2',
#             'date_out_2',
#             'extension',
#             'amount2',
#             'date_in_3',
#             'date_out_3',
#             'chasis_no',
#             'engine_no',
#             'CONTRACT_NUMBER',
#             'Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for leasing in leasing_queryset:
#         row_num += 1
#         row = [
#             leasing.Activity_Id,
#             leasing.PLATE_NUMBER,
#             leasing.CS_NO,
#             leasing.COMPANY,
#             leasing.MODEL,
#             leasing.BRAND,
#             leasing.VEHICLE_MAKE,
#             leasing.VEHICLE_TYPE,
#             leasing.LAST_NAME_ASSIGNEE,
#             leasing.FIRST_NAME_ASSIGNEE,
#             leasing.VEHICLE_CATEGORY,
#             leasing.COST_CENTER,
#             leasing.ID_NUMBER,
#             leasing.BAND,
#             leasing.GROUP,
#             leasing.DIVISION,
#             leasing.DEPARTMENT,
#             leasing.SECTION,
#             leasing.IS_EMPLOYEE_ID,
#             leasing.IS_LASTNAME,
#             leasing.IS_FIRSTNAME,
#             leasing.LOCATION,
#             leasing.AREA,
#             leasing.ACQUISITION_DATE,
#             leasing.remarks,
#             leasing.acquisition_cost,
#             leasing.months_36,
#             leasing.amount1,
#             leasing.date_in_1,
#             leasing.date_out_1,
#             leasing.months_24,
#             leasing.amount_Vat_EX,
#             leasing.date_in_2,
#             leasing.date_out_2,
#             leasing.extension,
#             leasing.amount2,
#             leasing.date_in_3,
#             leasing.date_out_3,
#             leasing.chasis_no,
#             leasing.engine_no,
#             leasing.CONTRACT_NUMBER,
#             leasing.vleasing_status,
#         ]

#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response


# def vehicle_excel(request):
#     v_queryset = VehicleMasterList.objects.all()   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist'

#     columns = [
            
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'REGISTRATION_MONTH',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'Employee',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'VEHICLE_CATEGORY',
#             'BAND_LEVEL', 
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_NAME',
#             'LOCATION',
#             'ORIGINAL_OR_DATE',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'EQUIPMENT_NO',
#             'SAP_PR',
#             'Vehicle_IVN_no',
#             'Unit_MATDOC',
#             'dealer',
#             'dealer_name',
#             'PO_NO',
#             'CHECKLIST BY',
#             'PLATE_NUMBER_RELEASE_DATE',
#             'Last_Registration_Date',
#             'Smoke_Emission_Date',
#             'Smoke_due',
#             'COC_Date',
#             'Remarks',
#             'Status',
#             'Other Remarks',
#             'Vehicle Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
#         # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.Employee,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.BAND_LEVEL, 
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.SAP_PR,
#                 vehicle.Vehicle_IVN_no,
#                 vehicle.Unit_MATDOC,
#                 vehicle.dealer,
#                 vehicle.dealer_name,
#                 vehicle.PO_NO,
#                 vehicle.CHECKLIST_BY,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#                 vehicle.Last_Registration_Date,
#                 vehicle.Smoke_Emission_Date,
#                 vehicle.Smoke_due,
#                 vehicle.COC_Date,
#                 vehicle.Remarks,
#                 vehicle.Status,
#                 vehicle.leasing_remark,
#                 vehicle.vehicle_status,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response
    

# def vehicle_excel_active(request):
#     v_queryset = VehicleMasterList.objects.filter(vehicle_status='Active')   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Active.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist Active'

#     columns = [
            
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'REGISTRATION_MONTH',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'Employee',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'VEHICLE_CATEGORY',
#             'BAND_LEVEL', 
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_NAME',
#             'LOCATION',
#             'ORIGINAL_OR_DATE',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'EQUIPMENT_NO',
#             'SAP_PR',
#             'Vehicle_IVN_no',
#             'Unit_MATDOC',
#             'dealer',
#             'dealer_name',
#             'PO_NO',
#             'CHECKLIST BY',
#             'PLATE_NUMBER_RELEASE_DATE',
#             'Last_Registration_Date',
#             'Smoke_Emission_Date',
#             'Smoke_due',
#             'COC_Date',
#             'Remarks',
#             'Status',
#             'Other Remarks',
#             'Vehicle Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.Employee,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.BAND_LEVEL, 
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.SAP_PR,
#                 vehicle.Vehicle_IVN_no,
#                 vehicle.Unit_MATDOC,
#                 vehicle.dealer,
#                 vehicle.dealer_name,
#                 vehicle.PO_NO,
#                 vehicle.CHECKLIST_BY,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#                 vehicle.Last_Registration_Date,
#                 vehicle.Smoke_Emission_Date,
#                 vehicle.Smoke_due,
#                 vehicle.COC_Date,
#                 vehicle.Remarks,
#                 vehicle.Status,
#                 vehicle.leasing_remark,
#                 vehicle.vehicle_status,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response

# def vehicle_excel_solved(request):
#     v_queryset = VehicleMasterList.objects.filter(vehicle_status='Sold')   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Sold.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist Sold'

#     columns = [
            
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'REGISTRATION_MONTH',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'Employee',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'VEHICLE_CATEGORY',
#             'BAND_LEVEL', 
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_NAME',
#             'LOCATION',
#             'ORIGINAL_OR_DATE',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'EQUIPMENT_NO',
#             'SAP_PR',
#             'Vehicle_IVN_no',
#             'Unit_MATDOC',
#             'dealer',
#             'dealer_name',
#             'PO_NO',
#             'CHECKLIST BY',
#             'PLATE_NUMBER_RELEASE_DATE',
#             'Last_Registration_Date',
#             'Smoke_Emission_Date',
#             'Smoke_due',
#             'COC_Date',
#             'Remarks',
#             'Status',
#             'Other Remarks',
#             'Vehicle Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.Employee,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.BAND_LEVEL, 
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.SAP_PR,
#                 vehicle.Vehicle_IVN_no,
#                 vehicle.Unit_MATDOC,
#                 vehicle.dealer,
#                 vehicle.dealer_name,
#                 vehicle.PO_NO,
#                 vehicle.CHECKLIST_BY,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#                 vehicle.Last_Registration_Date,
#                 vehicle.Smoke_Emission_Date,
#                 vehicle.Smoke_due,
#                 vehicle.COC_Date,
#                 vehicle.Remarks,
#                 vehicle.Status,
#                 vehicle.leasing_remark,
#                 vehicle.vehicle_status,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response
    
# def vehicle_excel_trans(request):
#     v_queryset = VehicleMasterList.objects.filter(vehicle_status='Transferred')   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Transferred.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist Transferred'

#     columns = [
            
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'REGISTRATION_MONTH',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'Employee',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'VEHICLE_CATEGORY',
#             'BAND_LEVEL', 
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_NAME',
#             'LOCATION',
#             'ORIGINAL_OR_DATE',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'EQUIPMENT_NO',
#             'SAP_PR',
#             'Vehicle_IVN_no',
#             'Unit_MATDOC',
#             'dealer',
#             'dealer_name',
#             'PO_NO',
#             'CHECKLIST BY',
#             'PLATE_NUMBER_RELEASE_DATE',
#             'Last_Registration_Date',
#             'Smoke_Emission_Date',
#             'Smoke_due',
#             'COC_Date',
#             'Remarks',
#             'Status',
#             'Other Remarks',
#             'Vehicle Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.Employee,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.BAND_LEVEL, 
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.SAP_PR,
#                 vehicle.Vehicle_IVN_no,
#                 vehicle.Unit_MATDOC,
#                 vehicle.dealer,
#                 vehicle.dealer_name,
#                 vehicle.PO_NO,
#                 vehicle.CHECKLIST_BY,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#                 vehicle.Last_Registration_Date,
#                 vehicle.Smoke_Emission_Date,
#                 vehicle.Smoke_due,
#                 vehicle.COC_Date,
#                 vehicle.Remarks,
#                 vehicle.Status,
#                 vehicle.leasing_remark,
#                 vehicle.vehicle_status,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response


# def vehicle_excel_bayan(request):
#     v_queryset = VehicleMasterList.objects.filter(CR_NAME__contains="BAYANTEL")   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Bayan.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist Bayan'

#     columns = [
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'REGISTRATION_MONTH',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'Employee',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'VEHICLE_CATEGORY',
#             'BAND_LEVEL', 
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_LAST_NAME',
#             # 'IS_FIRST_NAME',
#             'LOCATION',
#             'ORIGINAL_OR_DATE',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'EQUIPMENT_NO',
#             'SAP_PR',
#             'Vehicle_IVN_no',
#             'Unit_MATDOC',
#             'dealer',
#             'dealer_name',
#             'PO_NO',
#             'PLATE_NUMBER_RELEASE_DATE',
#             'Last_Registration_Date',
#             'Smoke_Emission_Date',
#             'Smoke_due',
#             'COC_Date',
#             'Remarks',
#             'Status',
#             'Other Remarks',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
#         # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.Employee,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.BAND_LEVEL, 
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_NAME,
#                 # vehicle.IS_FIRST_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.SAP_PR,
#                 vehicle.Vehicle_IVN_no,
#                 vehicle.Unit_MATDOC,
#                 vehicle.dealer,
#                 vehicle.dealer_name,
#                 vehicle.PO_NO,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#                 vehicle.Last_Registration_Date,
#                 vehicle.Smoke_Emission_Date,
#                 vehicle.Smoke_due,
#                 vehicle.COC_Date,
#                 vehicle.Remarks,
#                 vehicle.Status,
#                 vehicle.leasing_remark,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response
    
# def vehicle_excel_teli(request):
#     v_queryset = VehicleMasterList.objects.filter(CR_NAME__contains="TELICPHIL")   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Telicphil.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist Telicphil'

#     columns = [
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'REGISTRATION_MONTH',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'Employee',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'VEHICLE_CATEGORY',
#             'BAND_LEVEL', 
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_NAME',
#             # 'IS_FIRST_NAME',
#             'LOCATION',
#             'ORIGINAL_OR_DATE',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'EQUIPMENT_NO',
#             'SAP_PR',
#             'Vehicle_IVN_no',
#             'Unit_MATDOC',
#             'dealer',
#             'dealer_name',
#             'PO_NO',
#             'PLATE_NUMBER_RELEASE_DATE',
#             'Last_Registration_Date',
#             'Smoke_Emission_Date',
#             'Smoke_due',
#             'COC_Date',
#             'Remarks',
#             'Status',
#             'Other Remarks',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
#         # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.Employee,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.BAND_LEVEL, 
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_NAME,
#                 # vehicle.IS_FIRST_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.SAP_PR,
#                 vehicle.Vehicle_IVN_no,
#                 vehicle.Unit_MATDOC,
#                 vehicle.dealer,
#                 vehicle.dealer_name,
#                 vehicle.PO_NO,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#                 vehicle.Last_Registration_Date,
#                 vehicle.Smoke_Emission_Date,
#                 vehicle.Smoke_due,
#                 vehicle.COC_Date,
#                 vehicle.Remarks,
#                 vehicle.Status,
#                 vehicle.leasing_remark,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response

def employee_excel(request):
    emp_queryset = EmployeeMasterlist.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Employee Masterlist.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Employee Masterlist'

    columns = [
            'Company',
            'Employee_Id',
            'Last_name',
            'First_name',
            'Middle_name',
            'Suffix',
            'Band',
            'Cost_center',
            'DIV_code',
            'Group',
            'Division',
            'Department',
            'Section',
            'Unit',
            'Sub_unit',
            'IS_ID',
            'IS_lastname',
            'IS_firstname',
            'Location',
            'Area',
            'Area2',
            'Benefit',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for emp in emp_queryset:
        row_num += 1

        row = [
            emp.Company,
            emp.Employee_Id,
            emp.Last_name,
            emp.First_name,
            emp.Middle_name,
            emp.Suffix,
            emp.Band,
            emp.Cost_center,
            emp.DIV_code,
            emp.Group,
            emp.Division,
            emp.Department,
            emp.Section,
            emp.Unit,
            emp.Sub_unit,
            emp.IS_ID,
            emp.IS_lastname,
            emp.IS_firstname,
            emp.Location,
            emp.Area,
            emp.Area2,
            emp.Benefit,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response





