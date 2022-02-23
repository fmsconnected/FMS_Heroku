import os
import numpy as np
import pandas as pd
import datetime
from django.shortcuts import render,HttpResponseRedirect,HttpResponse,redirect
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from .render import Render
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.core.mail import get_connection,send_mail
from django.core import mail
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.template import Context
from datetime import date, timedelta
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from .models import (
   	Registration,
    CarRegistration
)
from django import template
register = template.Library()
from django.template.defaultfilters import floatformat
from rest_framework import viewsets
from rest_framework.response import Response
from django.db.models import Q
from . forms import (
    reg_update_Form,
    
)
from vehicle_masterlist.models import(
    VehicleMasterList
    )
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                                View,
                				)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

from .serializers import (
    RegistrationSerializer
    )


def registration_new(request):
    return render(request, 'registration_form.html')

# def regUpdate(request, pk):
#     if request.method == 'POST':
#         Last_Registration_Date = request.POST.get('Last_Registration_Date')
#         Smoke_Emission_Date = request.POST.get('Smoke_Emission_Date')
#         COC_Date = request.POST.get('COC_Date')
#         Remarks = request.POST.get('Remarks')

	
#     VehicleMasterList.objects.filter(id=pk).update(Last_Registration_Date=Last_Registration_Date, Smoke_Emission_Date=Smoke_Emission_Date,
#      COC_Date=COC_Date, Remarks=Remarks)
#     return HttpResponseRedirect('/Registration/Details/{}'.format(pk))


class registrationViewSet(viewsets.ModelViewSet):
    queryset = VehicleMasterList.objects.all().order_by('id')
    serializer_class = RegistrationSerializer

class regUpdate(UpdateView):
    model = VehicleMasterList
    form_class = reg_update_Form
    template_name = 'regupdate.html'
    success_url = reverse_lazy('/Registration/January/')

class registrationDetails(DetailView):
    model = VehicleMasterList
    template_name = 'registration_details.html'

def registrationCreate(request):
    if request.method == 'POST':
        PLATE_NO = request.POST.get('plate_no')
        CS_NO = request.POST.get('cs')
        CR_NAME = request.POST.get('cr_name')
        MODEL = request.POST.get('model')
        BRAND = request.POST.get('brand')
        VEHICLE_MAKE = request.POST.get('vmake')
        ENGINE_NO = request.POST.get('eng_no')
        CHASSIS_NO = request.POST.get('chassis_no')
        MV_FILE_NO = request.POST.get('mvfile')
        COC = request.POST.get('coc')
        SMOKE_TPL = request.POST.get('smoke')
        REMARKS_REGISTERED = request.POST.get('remarks_registered')
        DATE_EMAILED = request.POST.get('demail')
        JUSTIFICATION_REMARKS = request.POST.get('remarks_justification')
        email_status = request.POST.get('email_status')
        email = request.POST.get('email')

        reg = ''
        endplate = ''
        # if PLATE_NO != ' ':
        if PLATE_NO != '':
            endplate = int(PLATE_NO[-1])
            if endplate == 1:
                reg = 'JAN'
            elif endplate == 2:
                reg = 'FEB'
            elif endplate == 3:
                reg = 'MAR'
            elif endplate == 4:
                reg = 'APR'
            elif endplate == 5:
                reg = 'MAY'
            elif endplate == 6:
                reg = 'JUN'
            elif endplate == 7:
                reg = 'JUL'
            elif endplate == 8:
                reg = 'AUG'
            elif endplate == 9:
                reg = 'SEP'
            elif endplate == 0:
                reg = 'OCT' 

        saveto_end = Registration(PLATE_NO=PLATE_NO,Plate_ending= endplate, CS_NO=CS_NO, CR_NAME=CR_NAME, MODEL=MODEL,BRAND=BRAND,
            VEHICLE_MAKE=VEHICLE_MAKE, ENGINE_NO=ENGINE_NO, CHASSIS_NO=CHASSIS_NO, MV_FILE_NO=MV_FILE_NO,
            COC=COC, SMOKE_TPL=SMOKE_TPL, REMARKS_REGISTERED=REMARKS_REGISTERED, DATE_EMAILED=DATE_EMAILED,
            JUSTIFICATION_REMARKS=JUSTIFICATION_REMARKS, Registration_month = reg, sent_email=email_status, email=email
            )
        saveto_end.save()

    return HttpResponseRedirect('/Registration/January')



class registrationDeleteView(BSModalDeleteView):
    model = Registration
    template_name = 'registration_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('vehicle-list')

year = datetime.datetime.now().year
year1 = datetime.datetime.now().year - 1
year2 = datetime.datetime.now().year - 2

def othersRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'other_list': VehicleMasterList.objects.filter(PLATE_NO__isnull=True).exclude(exc)
        }

    return render(request, 'reg_others.html', context)

def trailerRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'trailer_list': VehicleMasterList.objects.filter(BRAND__contains="TRAILER").exclude(exc)
    .exclude(exc)    
    }

    return render(request, 'reg_trailer.html', context)


def janRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__startswith="JAN").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__startswith="JAN").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }
    print(context)

    return render(request, 'regJan_monitoring.html', context)

def febRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }

    return render(request, 'regFeb_monitoring.html', context)

def marRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }

    return render(request, 'regMar_monitoring.html', context)

def aprRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)

        }

    return render(request, 'regApr_monitoring.html', context)

def mayRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }

    return render(request, 'regMay_monitoring.html', context)

def junRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }

    return render(request, 'regJun_monitoring.html', context)

def julRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }
    return render(request, 'regJul_monitoring.html', context)

def augRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }

    return render(request, 'regAug_monitoring.html', context)

def sepRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }

    return render(request, 'regSep_monitoring.html', context)

def octRegView(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
            'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT").filter(Q(Smoke_Emission_Date__startswith="No") | Q(COC_Date__startswith="No")).exclude(exc),
            'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT").filter(Q(Smoke_Emission_Date__startswith="") | Q(COC_Date__startswith="")).exclude(exc)
        }

    return render(request, 'regOct_monitoring.html', context)

def summary(request):
    exc = Q(ACQ_DATE__year=year) | Q(ACQ_DATE__year=year1) | Q(ACQ_DATE__year=year2)
    context = {
        'reg_report':VehicleMasterList.objects.all()
    }
    return render(request, 'summaryV2.html', context)


class registrationsPDFView(View):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    @register.filter
    def as_percentage_of(part, whole):
        try:
            return "%d%%" % (float(part) / whole * 100)
        except (ValueError, ZeroDivisionError):
            return ""
    def get(self, request, pk, *args, **kwargs):
        reg_report = Registration.objects.get(pk=pk)
        return Render.render('reg_summary_report.html', {'reg_report': reg_report})

def registration_excel(request):
    rental_queryset = Registration.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Registration.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Registration'

    columns = [
        'Plate No' ,
        'Plate Ending' ,
        'CS NO' ,
        'CR NAME' ,
        'MODEL' ,
        'BRAND' ,
        'VEHICLE MAKE' ,
        'ENGINE NO' ,
        'CHASSIS NO' ,
        'MV_FILE NO' ,
        'COC' ,
        'SMOKE TPL' ,
        'REMARKS REGISTERED' ,
        'DATE EMAILED' ,
        'JUSTIFICATION REMARKS' ,
        'Registration month' ,
        'Email' ,
        'Sent email' ,
        'Date email log' ,
        'Date Registered',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for car in rental_queryset:
        row_num += 1
        row = [
                car.PLATE_NO ,
                car.Plate_ending ,
                car.CS_NO ,
                car.CR_NAME ,
                car.MODEL ,
                car.BRAND ,
                car.VEHICLE_MAKE ,
                car.ENGINE_NO ,
                car.CHASSIS_NO ,
                car.MV_FILE_NO ,
                car.COC ,
                car.SMOKE_TPL ,
                car.REMARKS_REGISTERED ,
                car.DATE_EMAILED ,
                car.JUSTIFICATION_REMARKS ,
                car.Registration_month ,
                car.email ,
                car.sent_email ,
                car.Date_email_log ,
                car.Date_registered,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


class HomeView():
    month = datetime.datetime.now().month
    date_now = datetime.datetime.now().date()
    print(date_now)
    sent_status = Registration.objects.all()
    if month == 1:
        car_status = Registration.objects.filter(
            Registration_month="JAN", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.plate_no)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 2:
        car_status = Registration.objects.filter(
            Registration_month="FEB", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 3:
        car_status = Registration.objects.filter(
            Registration_month="MAR", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 4:
        car_status = Registration.objects.filter(
            Registration_month="APR", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 5:
        car_status = Registration.objects.filter(
            Registration_month="MAY", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 6:
        car_status = Registration.objects.filter(
            Registration_month="JUN", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 7:
        car_status = Registration.objects.filter(
            Registration_month="JUL", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 8:
        car_status = Registration.objects.filter(
            Registration_month="AUG", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 9:
        car_status = Registration.objects.filter(
            Registration_month="SEP", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    elif month == 0:
        car_status = Registration.objects.filter(
            Registration_month="OCT", sent_email="No")
        plate = ""
        for carreg in car_status:
                # print(carreg.PLATE_NO)
                plate = carreg.PLATE_NO
                print(plate)

        if plate != "":
            for item in car_status:
                subject = 'Fleet Management System Automated Email'
                html_message = render_to_string('email.html',{'content':item.PLATE_NO})
                plain_message = item.PLATE_NO
                recipient_list = [item.email]
                from_email = 'Fleet Management System <jxmtsi.fms@gmail.com>'
                mail.send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message, fail_silently=False)
                car_status.update(sent_email="Yes")
                car_status.update(Date_email_log= date_now)

    model = Registration
    context_object_name = 'car_list'
    template_name = 'account/account/base/login.html'

def registration_report_detail(request):
    today = datetime.datetime.now()
    month = today.month

    # Due_For_Regs = Registration.objects.filter(Plate_ending=month).count()
    
    Due_For_Regs1 = Registration.objects.filter(Registration_month="JAN").count()
    Completed1 = Registration.objects.filter(Registration_month="JAN").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs2 = Registration.objects.filter(Registration_month="FEB").count()
    Completed2 = Registration.objects.filter(Registration_month="FEB").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs3 = Registration.objects.filter(Registration_month="MAR").count()
    Completed3 = Registration.objects.filter(Registration_month="MAR").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs4 = Registration.objects.filter(Registration_month="APR").count()
    Completed4 = Registration.objects.filter(Registration_month="APR").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs5 = Registration.objects.filter(Registration_month="MAY").count()
    Completed5 = Registration.objects.filter(Registration_month="MAY").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs6 = Registration.objects.filter(Registration_month="JUN").count()
    Completed6 = Registration.objects.filter(Registration_month="JUN").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs7 = Registration.objects.filter(Registration_month="JUL").count()
    Completed7 = Registration.objects.filter(Registration_month="JUL").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs8 = Registration.objects.filter(Registration_month="AUG").count()
    Completed8 = Registration.objects.filter(Registration_month="AUG").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs9 = Registration.objects.filter(Registration_month="SEP").count()
    Completed9 = Registration.objects.filter(Registration_month="SEP").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs0 = Registration.objects.filter(Registration_month="OCT").count()
    Completed0 = Registration.objects.filter(Registration_month="OCT").exclude(Date_registered__isnull=True).count()

    return  render(request, 'registration_report_details.html',{'title':'Registration - Registration Monitoring', 'Due_For_Regs1':Due_For_Regs1, 'Completed1':Completed1,
     'Due_For_Regs2':Due_For_Regs2, 'Completed2':Completed2, 'Due_For_Regs3':Due_For_Regs3, 'Due_For_Regs4':Due_For_Regs4, 'Completed4':Completed4,
     'Due_For_Regs5':Due_For_Regs5, 'Completed5':Completed5, 'Due_For_Regs6':Due_For_Regs6, 'Completed6':Completed6, 'Due_For_Regs7':Due_For_Regs7, 'Completed7':Completed7,
     'Due_For_Regs8':Due_For_Regs8, 'Completed8':Completed8, 'Due_For_Regs9':Due_For_Regs9, 'Completed9':Completed9, 'Due_For_Regs0':Due_For_Regs0, 'Completed0':Completed0,'today':today})

# Registration Daily Report
    
def registration_report(request):
    today = datetime.datetime.now()
    month = today.month

    # Due_For_Regs = Registration.objects.filter(Registration_month=month).count()
    
    Due_For_Regs1 = Registration.objects.filter(Registration_month="JAN").count()
    Completed1 = Registration.objects.filter(Registration_month="JAN").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs2 = Registration.objects.filter(Registration_month="FEB").count()
    Completed2 = Registration.objects.filter(Registration_month="FEB").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs3 = Registration.objects.filter(Registration_month="MAR").count()
    Completed3 = Registration.objects.filter(Registration_month="MAR").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs4 = Registration.objects.filter(Registration_month="APR").count()
    Completed4 = Registration.objects.filter(Registration_month="APR").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs5 = Registration.objects.filter(Registration_month="MAY").count()
    Completed5 = Registration.objects.filter(Registration_month="MAY").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs6 = Registration.objects.filter(Registration_month="JUN").count()
    Completed6 = Registration.objects.filter(Registration_month="JUN").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs7 = Registration.objects.filter(Registration_month="JUL").count()
    Completed7 = Registration.objects.filter(Registration_month="JUL").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs8 = Registration.objects.filter(Registration_month="AUG").count()
    Completed8 = Registration.objects.filter(Registration_month="AUG").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs9 = Registration.objects.filter(Registration_month="SEP").count()
    Completed9 = Registration.objects.filter(Registration_month="SEP").exclude(Date_registered__isnull=True).count()
    
    Due_For_Regs0 = Registration.objects.filter(Registration_month="OCT").count()
    Completed0 = Registration.objects.filter(Registration_month="OCT").exclude(Date_registered__isnull=True).count()

    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Registration_Report.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active
    ws.title = "Registration Report"
    ws['A1'].value = "PERSONNEL:"
    ws['B1'].value = "Francis Jae Dela Cruz"
    ws['A3'].value = "OUTPUT"
    ws['A4'].value = ""

##### Plate Ending 1 #######
    ws['A5'].value = "Ending 1"
    ws.append(['', 'Total','Remarks'])
    ws['A7'].value = "Due For Regs"
    ws['A8'].value = "Registered"
    ws['A9'].value = "unRegistered"
    ws['A10'].value = "Uploading"
    ws['A11'].value = "Alarm"
    ws['A12'].value = "Completed"
    ws['A13'].value = ""
    ws['A14'].value = ""

    ws['B7'].value = Due_For_Regs1
    ws['B12'].value = Completed1
##### Plate Ending 2 #######

    ws['A15'].value = "Ending 2"
    ws['A16'].value = "Due For Regs"
    ws['A17'].value = "Registered"
    ws['A18'].value = "unRegistered"
    ws['A19'].value = "Uploading"
    ws['A20'].value = "Alarm"
    ws['A21'].value = "Completed"
    ws['A22'].value = ""
    ws['A23'].value = ""

    ws['B16'].value = Due_For_Regs2
    ws['B21'].value = Completed2
##### Plate Ending 3 #######

    ws['A24'].value = "Ending 3"
    ws['A25'].value = "Due For Regs"
    ws['A26'].value = "Registered"
    ws['A27'].value = "unRegistered"
    ws['A28'].value = "Uploading"
    ws['A29'].value = "Alarm"
    ws['A30'].value = "Completed"
    ws['A31'].value = ""
    ws['A32'].value = ""
    ws['B25'].value = Due_For_Regs3
    ws['B30'].value = Completed3
##### Plate Ending 4 #######

    ws['A33'].value = "Ending 4"
    ws['A34'].value = "Due For Regs"
    ws['A35'].value = "Registered"
    ws['A36'].value = "unRegistered"
    ws['A37'].value = "Uploading"
    ws['A38'].value = "Alarm"
    ws['A39'].value = "Completed"
    ws['A40'].value = ""
    ws['A41'].value = ""
    ws['B34'].value = Due_For_Regs4
    ws['B39'].value = Completed4
##### Plate Ending 5 #######

    ws['A42'].value = "Ending 5"
    ws['A43'].value = "Due For Regs"
    ws['A44'].value = "Registered"
    ws['A45'].value = "unRegistered"
    ws['A46'].value = "Uploading"
    ws['A47'].value = "Alarm"
    ws['A48'].value = "Completed"
    ws['A49'].value = ""
    ws['A50'].value = ""
    ws['B43'].value = Due_For_Regs5
    ws['B48'].value = Completed5

    ws['G6'].value = ""
    ws['H6'].value = "Total"
    ws['I6'].value = "Remarks"
##### Plate Ending 6 #######

    ws['G5'].value = "Ending 6"
    ws['G7'].value = "Due For Regs"
    ws['G8'].value = "Registered"
    ws['G9'].value = "unRegistered"
    ws['G10'].value = "Uploading"
    ws['G11'].value = "Alarm"
    ws['G12'].value = "Completed"
    ws['G13'].value = ""
    ws['G14'].value = ""
    ws['H7'].value = Due_For_Regs6
    ws['H12'].value = Completed6

##### Plate Ending 7 #######

    ws['G15'].value = "Ending 7"
    ws['G16'].value = "Due For Regs"
    ws['G17'].value = "Registered"
    ws['G18'].value = "unRegistered"
    ws['G19'].value = "Uploading"
    ws['G20'].value = "Alarm"
    ws['G21'].value = "Completed"
    ws['G22'].value = ""
    ws['G23'].value = ""
    ws['H16'].value = Due_For_Regs7
    ws['H21'].value = Completed7

##### Plate Ending 8 #######

    ws['G24'].value = "Ending 8"
    ws['G25'].value = "Due For Regs"
    ws['G26'].value = "Registered"
    ws['G27'].value = "unRegistered"
    ws['G28'].value = "Uploading"
    ws['G29'].value = "Alarm"
    ws['G30'].value = "Completed"
    ws['G31'].value = ""
    ws['G32'].value = ""
    ws['H25'].value = Due_For_Regs8
    ws['H30'].value = Completed8

##### Plate Ending 9 #######

    ws['G33'].value = "Ending 9"
    ws['G34'].value = "Due For Regs"
    ws['G35'].value = "Registered"
    ws['G36'].value = "unRegistered"
    ws['G37'].value = "Uploading"
    ws['G38'].value = "Alarm"
    ws['G39'].value = "Completed"
    ws['G40'].value = ""
    ws['G41'].value = ""
    ws['H34'].value = Due_For_Regs9
    ws['H39'].value = Completed9

##### Plate Ending 0 #######

    ws['G42'].value = "Ending 0"
    ws['G43'].value = "Due For Regs"
    ws['G44'].value = "Registered"
    ws['G45'].value = "unRegistered"
    ws['G46'].value = "Uploading"
    ws['G47'].value = "Alarm"
    ws['G48'].value = "Completed"
    ws['G49'].value = ""
    ws['G50'].value = ""
    ws['H43'].value = Due_For_Regs0
    ws['H48'].value = Completed0
    wb.save(output)
    return output
    # wb.save(f"/Users/{username}/Desktop/Registration_Report.xlsx")
    # return redirect('/Registration/January')

