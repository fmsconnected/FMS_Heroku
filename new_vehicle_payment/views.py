from django.shortcuts import render,HttpResponseRedirect,reverse,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
from datetime import date, timedelta
from .models import VehiclePayment
from masterlist.models import EmployeeMasterlist
from vehicle_masterlist.models import VehicleMasterList

from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from . forms import VehiclePaymentform
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )


def vehiclecreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    elist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'vehiclepayment_form.html',{'elist':elist,'vlist':vlist})

def vehicle_ongoing(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    ongoing = VehiclePayment.objects.filter(Status="Ongoing")
    return render(request, 'vehicle_ongoing.html',{'ongoing':ongoing})

def vehicle_completed(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    completed = VehiclePayment.objects.filter(Status="Completed")
    return render(request, 'vehicle_completed.html',{'completed':completed})

def vehicle_submit(request):
    if request.method == 'POST':
        a_emp_id = request.POST.get('a_emp_id')
        emp_fname = request.POST.get('emp_fname')
        emp_lname = request.POST.get('emp_lname')
        Delivery_Date = request.POST.get('Delivery_Date')
        Plate_Number = request.POST.get('Plate_Number')
        Model_Year = request.POST.get('Model_Year')
        Brand = request.POST.get('Brand')
        Make = request.POST.get('Make')
        Dealer = request.POST.get('Dealer')
        date_received = request.POST.get('Date_Received_LTO_Documents')
        Docs_plate_no = request.POST.get('Docs_plate_no')
        LTO_stickers = request.POST.get('LTO_stickers')
        Sticker_fields = request.POST.get('Sticker_fields')
        Date_initial = request.POST.get('Date_initial')
        First_payment = request.POST.get('First_payment')
        LTO_Charges = request.POST.get('LTO_Charges')
        Outstanding_Balance = request.POST.get('Outstanding_Balance')
        Date_final = request.POST.get('Date_final')
        Routing_Remarks = request.POST.get('Routing_Remarks')
        v_sla = request.POST.get('v_sla')
        PO_no = request.POST.get('rfp')
        invoice_number = request.POST.get('invno')
        equip_no = request.POST.get('equip_no')
        asset_no = request.POST.get('asset_no')
        sap_no = request.POST.get('sap_no')
        mat_no = request.POST.get('mat_no')
        Dealer_name = request.POST.get('Dealer_name')
        status = request.POST.get('status')
        
        saveto_v = VehiclePayment(A_employee_ID = a_emp_id,E_First_name = emp_fname,E_Last_name = emp_lname,V_deliverDate = Delivery_Date,
        Plate_no = Plate_Number,V_model = Model_Year,V_brand = Brand,V_make = Make,V_dealer = Dealer,LTO_documents = date_received,
        Docs_plate_no = Docs_plate_no,LTO_stickers = LTO_stickers,Sticker_fields = Sticker_fields,Date_initial = Date_initial,
        First_payment = First_payment,LTO_charges = LTO_Charges,Outstanding_balance = Outstanding_Balance,Date_final = Date_final,
        Routing_remarks = Routing_Remarks,V_SLA = v_sla,PO_no = PO_no,invoice_number = invoice_number, equip_no = equip_no,
        asset_no = asset_no, sap_no = sap_no, mat_no = mat_no, Dealer_name = Dealer_name, Status=status)
        saveto_v.save()

        return HttpResponseRedirect('/NVP/Vehicle/')

class VehicleListView(ListView):
    model = VehiclePayment
    template_name = 'vehicle_list.html'

class VehicleDetailView(DetailView):
    model = VehiclePayment
    template_name = 'vehiclepayment_summary.html'

class VehicleUpdateView(SuccessMessageMixin, UpdateView):
    model = VehiclePayment
    form_class = VehiclePaymentform
    template_name = 'vehiclepayment_new.html'

    def get_success_message(self, cleaned_data):
        print(cleaned_data)
        return "New Vehicle Payment's Update Successfully!"

class VehicleDeleteView(BSModalDeleteView):
    model = VehiclePayment
    template_name = 'vehicle_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('Vehicle_list')

def VehicleHistoryView(request):
    if request.method == "GET":
       obj = VehiclePayment.history.all()

       return render(request, 'vehicle_history.html', context={'object': obj})
def nvp_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = VehiclePayment.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = VehiclePayment.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = VehiclePayment.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = VehiclePayment.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = VehiclePayment.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = VehiclePayment.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'vehicledeadline.html',{'title':'Vehicle - Vehicle Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})

            

def vehicle_excel(request):
    vehicle_queryset = VehiclePayment.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Payment'

    columns = [
            'PO Number',
            'Employee Id',
            'First Name',
            'Last Name',
            'Delivery Date',
            'Plate No',
            'Model',
            'Brand',
            'Make',
            'Dealer',
            'LTO Documents',
            'Documents Plate No',
            'LTO Conduction Stickers',
            'Date Initial',
            'First Payment',
            'LTO Charges',
            'Outstanding Balance',
            'Date Final',
            'Remarks',
            'Date Initiated',
            'rfp_number',
            'invoice_number',
            'equip_no',
            'asset_no',
            'sap_no',
            'mat_no',
            'Dealer_name',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in vehicle_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                vehicle.rfp_number,
                vehicle.A_employee_ID,
                vehicle.E_First_name,
                vehicle.E_Last_name,
                vehicle.V_deliverDate,
                vehicle.Plate_no,
                vehicle.V_model,
                vehicle.V_brand,
                vehicle.V_make,
                vehicle.V_dealer,
                vehicle.LTO_documents,
                vehicle.Docs_plate_no,
                vehicle.LTO_stickers,
                vehicle.Date_initial,
                vehicle.First_payment,
                vehicle.LTO_charges,
                vehicle.Outstanding_balance,
                vehicle.Date_final,
                vehicle.Routing_remarks,
                vehicle.Date_initiated,
                vehicle.rfp_number,
                vehicle.invoice_number,
                vehicle.equip_no,
                vehicle.asset_no,
                vehicle.sap_no,
                vehicle.mat_no,
                vehicle.Dealer_name,
                
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

## Car rental Daily report Details
def car_report_details(request):
    date = datetime.datetime.today()
    carreport_SAFARI = CarRental.objects.filter(car_provider="SAFARI", sqa_number="").count()
    carreport_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND", sqa_number="").count()
    carreport_TG = CarRental.objects.filter(car_provider="Tiger City", sqa_number="").count()
    carreport_ORIX = CarRental.objects.filter(car_provider="ORIX", sqa_number="").count()
    
    vreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX", invoice_date="").count()
    vreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI", invoice_date="").count()
    vreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND", invoice_date="").count()
    vreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City", invoice_date="").count()
    
    processed_ORIX = CarRental.objects.filter(car_provider="ORIX").exclude(sqa_number='').count()
    processed_SAFARI = CarRental.objects.filter(car_provider="SAFARI").exclude(sqa_number='').count()
    processed_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND").exclude(sqa_number='').count()
    processed_TG = CarRental.objects.filter(car_provider="Tiger City").exclude(sqa_number='').count()

    proreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX").exclude(invoice_date="").count()
    proreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI").exclude(invoice_date="").count()
    proreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND").exclude(invoice_date="").count()
    proreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City").exclude(invoice_date="").count()
    
    shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL", SOA_billdate="").count()
    pro_shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL").exclude(SOA_billdate="").count()
    petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation", SOA_billdate="").count()
    pro_petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation").exclude(SOA_billdate="").count()

    orix = carreport_ORIX + vreport_ORIX
    safari = carreport_SAFARI + vreport_SAFARI
    diamond = carreport_DIAMOND + vreport_DIAMOND
    tg = carreport_TG + vreport_TG

    tp_orix = processed_ORIX + proreport_ORIX
    tp_safari = processed_SAFARI + proreport_SAFARI
    tp_diamond = processed_DIAMOND + proreport_DIAMOND
    tp_tg = processed_TG + proreport_TG

    return render(request, 'payment_report.html',{'title':'Payment','date':date,'orix':orix, 'safari':safari, 'diamond':diamond, 'tg':tg,'tp_orix':tp_orix, 'tp_safari':tp_safari, 'tp_diamond':tp_diamond, 'tp_tg':tp_tg})

# Car rental Daily Report
def car_report(request):
    date = datetime.datetime.today()
    yr = datetime.datetime.now().year
    months = ['zero','January','February','March','April','May','June','July','August','September','October','November','December']
    month = months[date.month]
    datem = datetime.datetime(date.year, date.month, 1)
    carreport_SAFARI = CarRental.objects.filter(car_provider="SAFARI", sqa_number="").count()
    carreport_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND", sqa_number="").count()
    carreport_TG = CarRental.objects.filter(car_provider="Tiger City", sqa_number="").count()
    carreport_ORIX = CarRental.objects.filter(car_provider="ORIX", sqa_number="").count()
    
    vreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX", invoice_date="").count()
    vreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI", invoice_date="").count()
    vreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND", invoice_date="").count()
    vreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City", invoice_date="").count()
    
    processed_ORIX = CarRental.objects.filter(car_provider="ORIX").exclude(sqa_number='').count()
    processed_SAFARI = CarRental.objects.filter(car_provider="SAFARI").exclude(sqa_number='').count()
    processed_DIAMOND = CarRental.objects.filter(car_provider="DIAMOND").exclude(sqa_number='').count()
    processed_TG = CarRental.objects.filter(car_provider="Tiger City").exclude(sqa_number='').count()

    proreport_ORIX = Vehicle_Repair_payment.objects.filter(dealership="ORIX").exclude(invoice_date="").count()
    proreport_SAFARI = Vehicle_Repair_payment.objects.filter(dealership="SAFARI").exclude(invoice_date="").count()
    proreport_DIAMOND = Vehicle_Repair_payment.objects.filter(dealership="DIAMOND").exclude(invoice_date="").count()
    proreport_TG = Vehicle_Repair_payment.objects.filter(dealership="Tiger City").exclude(invoice_date="").count()
    
    shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL", SOA_billdate="").count()
    pro_shell = Fuel_supplier.objects.filter(Fuel_provider="SHELL").exclude(SOA_billdate="").count()
    petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation", SOA_billdate="").count()
    pro_petron = Fuel_supplier.objects.filter(Fuel_provider="Petron Corporation").exclude(SOA_billdate="").count()

    orix = carreport_ORIX + vreport_ORIX
    safari = carreport_SAFARI + vreport_SAFARI
    diamond = carreport_DIAMOND + vreport_DIAMOND
    tg = carreport_TG + vreport_TG

    tp_orix = processed_ORIX + proreport_ORIX
    tp_safari = processed_SAFARI + proreport_SAFARI
    tp_diamond = processed_DIAMOND + proreport_DIAMOND
    tp_tg = processed_TG + proreport_TG


    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Insurance Daily Report.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Insurance Daily Report"
    ws['A1'].value = "Personel"
    ws['B1'].value = "Janine Manzo"
    ws['B2'].value = datem
    ws['A2'].value = "Date"
    ws.append(['Vendor','Total Unpaid SOA','Total Processed','For this Month Total SOA Processed'])
    ws['A4'].value = "ORIX"
    ws['A5'].value = "DIAMOND"
    ws['A6'].value = "PETRON"
    ws['A7'].value = "SHELL"
    ws['A8'].value = "GR8"
    ws['A9'].value = "JXM"
    ws['A10'].value = "SAFARI (Car Rental)"
    ws['A11'].value = "Tiger City (Car Rental)"
    ws['A12'].value = "Created Work Order"
    ws['A13'].value = "TopServe (FLEET Driver)"
    ws['A14'].value = "CARPLAN 95%"
    ws['A15'].value = "CARPLAN 5%"
    #data
    ws['B4'].value = orix
    ws['B5'].value = diamond
    ws['B6'].value = petron
    ws['B7'].value = shell
    ws['B8'].value = ""
    ws['B9'].value = ""
    ws['B10'].value = safari
    ws['B11'].value = tg
    ws['B12'].value = ""
    ws['B13'].value = ""
    ws['B14'].value = ""
    ws['B15'].value = ""

    ws['C4'].value = tp_orix
    ws['C5'].value = tp_diamond
    ws['C6'].value = pro_petron
    ws['C7'].value = pro_shell
    ws['C8'].value = ""
    ws['C9'].value = ""
    ws['C10'].value = tp_safari
    ws['C11'].value = tp_tg
    ws['C12'].value = ""
    ws['C13'].value = ""
    ws['C14'].value = ""
    ws['C15'].value = ""

   
    wb.save(output)
    return output
