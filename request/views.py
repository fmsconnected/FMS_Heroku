from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from django.views import generic
import schedule
import time
from django.core.mail import get_connection,send_mail
from django.core import mail
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.template import Context
from django.shortcuts import render,HttpResponseRedirect, get_list_or_404,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
import datetime
from datetime import date, timedelta
from django.db.models import Q, Count
from .models import (
        Vehicle_Repair,
        Gas_card
)
from masterlist.models import (
    EmployeeMasterlist,
    )
from vehicle_masterlist.models import VehicleMasterList
from leasingmasterlist.models import Leasing

from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from CustomerLog.models import (
    CS_log
)
from . forms import (
    repairform,
    gascardform

    )

from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )



                      #####################################  
                    #########################################
                   #####.       Gas Card Request         #####
                    #########################################
                     ######################################


def gascreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    elist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'gas_card/gascard_new.html',{'Title':'Gas - Gas Card Request', 'elist':elist, 'vlist':vlist})
    # def get_success_message(self, cleaned_data):
    #    print(cleaned_data)
    #    return "New Car Rental Request Has been Created!"

def gassubmit(request):
    if request.method == 'POST':
        date_app = request.POST.get('date_app')
        app_type = request.POST.get('app_type')
        fleet_card = request.POST.get('fleet_card')
        fleet_card_type = request.POST.get('fleet_card_type')
        fuel_amount = request.POST.get('fuel_amount')
        fuel_quantity = request.POST.get('fuel_quantity')
        products_restriction = request.POST.get('products_restriction')
        req_emp_id = request.POST.get('req_emp_id')
        r_fname = request.POST.get('r_fname')
        r_lname = request.POST.get('r_lname')
        r_costcenter = request.POST.get('r_costcenter')
        r_title = request.POST.get('r_title')
        atd_no = request.POST.get('atd_no')
        temp_atd = request.POST.get('temp_atd')
        new_empId = request.POST.get('new_empId')
        new_fname = request.POST.get('new_fname')
        new_lname = request.POST.get('new_lname')
        new_costcenter = request.POST.get('new_costcenter')
        new_tempATD = request.POST.get('new_tempATD')
        new_assignee = request.POST.get('new_assignee')
        cost_code = request.POST.get('cost_code')
        gcr_cancel = request.POST.get('gcr_cancel')
        plate_no = request.POST.get('plate_no')
        c_sticker = request.POST.get('c_sticker')
        gcr_sla = request.POST.get('gcr_sla')
        v_brand = request.POST.get('v_brand')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        v_fueltype = request.POST.get('v_fueltype')
        new_plate_no = request.POST.get('new_plate_no')
        new_cs = request.POST.get('new_cs')
        new_model = request.POST.get('new_model')
        new_brand = request.POST.get('new_brand')
        new_make = request.POST.get('new_make')
        new_fueltype = request.POST.get('new_fueltype')
        approved_by = request.POST.get('approved_by')
        date_sumitted_app = request.POST.get('date_sumitted_app')
        date_recieved_fleet = request.POST.get('date_recieved_fleet')
        fleet_card_no = request.POST.get('fleet_card_no')
        fleet_card_release = request.POST.get('fleet_card_release')
        person_release_card = request.POST.get('person_release_card')
        gcr_sla = request.POST.get('gcr_sla')

        saveto_gas = Gas_card(date_received = date_app, application_type=app_type,fleet_provider= fleet_card,fleetcard_type =fleet_card_type ,
            fuel_limit_amount=fuel_amount, fuel_limit_quantity = fuel_quantity, products_restriction = products_restriction,req_employee = req_emp_id,
            req_fname = r_fname,req_lname = r_lname ,req_title = r_title,req_cost_center = r_costcenter, atd_no=atd_no, temporary_atd=temp_atd,
            new_emp_id=new_empId, new_emp_fname=new_fname, new_emp_lname=new_lname, new_emp_cost = new_costcenter,new_temp_atd = new_tempATD,
            new_assignee = new_assignee,cost_center_code = cost_code,cancellation = gcr_cancel,plate_no = plate_no,con_sticker = c_sticker,
            model_year = v_model,brand = v_brand,make = v_make,fuel_type = v_fueltype,new_plate_no = new_plate_no,new_cond_sticker = new_cs,
            new_model_year = new_model,new_vbrand = new_brand,new_vmake = new_make,new_vfuel_type = new_fueltype,approved_by = approved_by,
            date_summitted =date_sumitted_app,fleet_received = date_recieved_fleet,fleet_card_no = fleet_card_no,fleet_date_release = fleet_card_release,
            person_release = person_release_card,GCR_SLA = gcr_sla)
        saveto_gas.save()

        return HttpResponseRedirect('/Request/Gas/')
    
class gasListView(ListView):
    model = Gas_card
    template_name = 'gas_card/gascard_list.html'

class gasUpdateView(SuccessMessageMixin, UpdateView):
    model = Gas_card
    form_class = gascardform
    template_name = 'gas_card/gascard_form.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Gas Card Details Update Successfully!"

class gasDetailView(DetailView):
    model = Gas_card
    template_name = 'gas_card/gascard_details.html'

class gasDeleteView(BSModalDeleteView):
    model = Gas_card
    template_name = 'gas_card/gascard_delete.html'
    success_message = 'Success: Gas Gard Request was deleted.'
    success_url = reverse_lazy('gascard_list')

def gasHistoryView(request):
    if request.method == "GET":
       obj = Gas_card.history.all()

       return render(request, 'gas_card/gascard_history.html', context={'object': obj})
def gcc_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = Gas_card.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = Gas_card.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = Gas_card.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = Gas_card.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = Gas_card.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = Gas_card.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'gas_card/gccdeadline.html',{'title':'Gas - Gas Card Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})


def gas_request_excel(request):
    rental_queryset = Gas_card.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Gas Card Request.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Gas Card Request'

    columns = [
            'Date Received' ,
            'Application Type' ,
            'Fleet Provider' ,
            'Fleet Card Type' ,
            'Fuel Limit Amount' ,
            'Fuel Limit Quantity' ,
            'Products Restriction' ,
            'Employee' ,
            'First Name' ,
            'Last Name' ,
            'Title' ,
            'Cost Center' ,
            'ATD No' ,
            'Temporary ATD' ,
            'New Employee ID' ,
            'New employee First Name' ,
            'New employee Last Name' ,
            'New employee Cost Center' ,
            'New employee ATD' ,
            'New Assignee' ,
            'Cost Center Code' ,
            'Cancellation' ,
            'Plate No' ,
            'Conduction Sticker' ,
            'Model Year' ,
            'Brand' ,
            'Make' ,
            'Fuel Type' ,
            'New Plate No' ,
            'New Conduction Sticker' ,
            'New Model Year' ,
            'New Vehicle Brand' ,
            'New Vehicle Make' ,
            'New Vehicle fuel type' ,
            'Approved By' ,
            'Date Summitted' ,
            'Fleet Received' ,
            'Fleet Card No' ,
            'Fleet Date Release' ,
            'Person Release' ,
            'Date Initiated' ,
            'SLA' ,
            'Deadline',

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for rental in rental_queryset:
        row_num += 1
        row = [
            rental.date_received ,
            rental.application_type ,
            rental.fleet_provider ,
            rental.fleetcard_type ,
            rental.fuel_limit_amount ,
            rental.fuel_limit_quantity ,
            rental.products_restriction ,
            rental.req_employee ,
            rental.req_fname ,
            rental.req_lname ,
            rental.req_title ,
            rental.req_cost_center ,
            rental.atd_no ,
            rental.temporary_atd ,
            rental.new_emp_id ,
            rental.new_emp_fname ,
            rental.new_emp_lname ,
            rental.new_emp_cost ,
            rental.new_temp_atd ,
            rental.new_assignee ,
            rental.cost_center_code ,
            rental.cancellation ,
            rental.plate_no ,
            rental.con_sticker ,
            rental.model_year ,
            rental.brand ,
            rental.make ,
            rental.fuel_type ,
            rental.new_plate_no ,
            rental.new_cond_sticker ,
            rental.new_model_year ,
            rental.new_vbrand ,
            rental.new_vmake ,
            rental.new_vfuel_type ,
            rental.approved_by ,
            rental.date_summitted ,
            rental.fleet_received ,
            rental.fleet_card_no ,
            rental.fleet_date_release ,
            rental.person_release ,
            rental.date_initiated ,
            rental.GCR_SLA ,
            rental.Deadline
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


                      #####################################  
                    #########################################
                   #####.     Vehicle Repair  Request    #####
                    #########################################
                     ######################################


def repairListView(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # model = Vehicle_Repair
    # template_name='vehicle_repair/repair_list.html'
    dl1 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl12 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl13 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl14 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl15 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl16 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today())
    
    dl1_count = dl1.aggregate(counted=Count('id'))['counted'] + dl12.aggregate(counted=Count('id'))['counted'] + dl13.aggregate(counted=Count('id'))['counted'] + dl14.aggregate(counted=Count('id'))['counted'] + dl15.aggregate(counted=Count('id'))['counted'] + dl16.aggregate(counted=Count('id'))['counted']  # number of records

    dl1_count = dl1.aggregate(counted=Count('id'))['counted'] + dl12.aggregate(counted=Count('id'))['counted'] + dl13.aggregate(counted=Count('id'))['counted'] + dl14.aggregate(counted=Count('id'))['counted'] + dl15.aggregate(counted=Count('id'))['counted'] + dl16.aggregate(counted=Count('id'))['counted']  # number of records

    object_list = Vehicle_Repair.objects.all()
    print(dl1_count)

    return render(request,'vehicle_repair/repair_list.html',{'Title':'Vehicle Repair List', 'object_list':object_list, 'dl1_count':dl1_count})

def repairCreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'vehicle_repair/repair_new.html',{'Title':'Vehicle - Vehicle Repair','emplist':emplist,'vlist':vlist})

def repairsubmit(request):
    if request.method == 'POST':
        request_date = request.POST.get('request_date')
        emp_id = request.POST.get('emp_id')
        cost_center = request.POST.get('cost_center')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        c_no = request.POST.get('c_no')
        company = request.POST.get('company')
        department = request.POST.get('department')
        group = request.POST.get('group')
        plate_no = request.POST.get('plate_no')
        v_brand = request.POST.get('v_brand')
        engine = request.POST.get('engine')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        chassis = request.POST.get('chassis')
        v_band = request.POST.get('v_band')
        cs_no = request.POST.get('cs_no')
        eq_no = request.POST.get('eq_no')
        fleet_area = request.POST.get('fleet_area')
        particulars = request.POST.get('particulars')
        category = request.POST.get('category')
        maintenance_type1 = request.POST.get('maintenance_type1')
        maintenance_type2 = request.POST.get('maintenance_type2')
        scope_work1 = request.POST.get('scope_work1')
        scope_work2 = request.POST.get('scope_work2')
        recomendations = request.POST.get('recomendations')
        service_reminder = request.POST.get('service_reminder')
        repair_verified_by = request.POST.get('repair_verified_by')
        work_order1 = request.POST.get('work_order1')
        work_order2 = request.POST.get('work_order2')
        work_order3 = request.POST.get('work_order3')
        date_work_created = request.POST.get('date_work_created')
        repair_shop = request.POST.get('repair_shop')
        memo_app = request.POST.get('memo_app')
        date_forward = request.POST.get('date_forward')
        estimate_no = request.POST.get('estimate_no')
        maintenance_amount = request.POST.get('maintenance_amount')
        less_discount = request.POST.get('less_discount')
        estimate_remark = request.POST.get('estimate_remark')
        estimate_attach = request.POST.get('estimate_attach')
        approved_by = request.POST.get('approved_by')
        kilo_reading = request.POST.get('kilo_reading')
        email = request.POST.get('email')
        email_status = request.POST.get('email_status')
        vrr_sla = request.POST.get('vrr_sla')

        # req_date = request_date.replace(day=30)
        # print(request_date)

        deadl = datetime.datetime.strptime(request_date, '%Y-%m-%d')
        dead_date = deadl + datetime.timedelta(days=30)
        # print(dead_date)

        saveto_repair = Vehicle_Repair(request_date=request_date, employee=emp_id, cost_center=cost_center, first_name=fname,
            last_name=lname, contact_no=c_no, company=company, department=department, group_section=group,
            plate_no=plate_no, v_brand=v_brand, engine=engine, v_make=v_make, v_model=v_model, chassis=chassis,
            band=v_band, cond_sticker=cs_no, equipment_no=eq_no, fleet_area=fleet_area, particulars=particulars,
            category=category, maintenance_type1=maintenance_type1, scope_work1=scope_work1, maintenance_type2=maintenance_type2,
            scope_work2=scope_work2, recommendations=recomendations, service_reminder=service_reminder, verified_by=repair_verified_by,
            work_order1=work_order1, work_order2=work_order2, work_order3=work_order3, datework_created=date_work_created,
            Shop_vendor=repair_shop, date_forwarded=date_forward, estimate_no=estimate_no, maintenance_amount=maintenance_amount,
            less_discount=less_discount, estimate_remarks=estimate_remark, estimate_attached=estimate_attach, approvedby=approved_by,
            meter_reading=kilo_reading, VRR_SLA=vrr_sla, memo_app=memo_app,email=email,sent_email=email_status,Deadline=dead_date
    )
        saveto_repair.save()

        return HttpResponseRedirect('/Request/Repair/')

class repairDetailView(DetailView):
    model = Vehicle_Repair
    template_name = 'vehicle_repair/repair_details.html'

class repairUpdateView(SuccessMessageMixin, UpdateView):
    model = Vehicle_Repair
    form_class = repairform
    template_name = 'vehicle_repair/repair_form.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Vehicle Repair Request Updated Successfully!"

class repairDeleteView(BSModalDeleteView):
    model = Vehicle_Repair
    template_name = 'vehicle_repair/repair_delete.html'
    success_message = 'Success: Vehicle Repair Request was deleted.'
    success_url = reverse_lazy('repair_list')

def repairHistoryView(request):
    if request.method == "GET":
       obj = Vehicle_Repair.history.all()

       return render(request, 'vehicle_repair/repair_history.html', context={'object': obj})
       
def vrp_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = Vehicle_Repair.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'vehicle_repair/vrpdeadline.html',{'title':'Vehicle - Vehicle Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})


def repair_request_excel(request):
    repair_queryset = Vehicle_Repair.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Preventive Maintenance.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Preventive Maintenance'

    columns = [
                'Request Date' ,
                'Employee' ,
                'Cost Center' ,
                'First Name' ,
                'Last Name' ,
                'Contact No' ,
                'Company' ,
                'Department' ,
                'Group Section' ,
                'Plate No' ,
                'Brand' ,
                'Engine' ,
                'Make' ,
                'Model' ,
                'Chassis' ,
                'Band' ,
                'Conduction Sticker' ,
                'Equipment No' ,
                'Fleet Area' ,
                'Particulars' ,
                'Category' ,
                'Maintenance Type 1' ,
                'Scope Work 1' ,
                'Maintenance Type 2' ,
                'Scope Work 2' ,
                'Recommendations' ,
                'Service Reminder' ,
                'Verified By' ,
                'Work Order 1' ,
                'Work Order 2' ,
                'Work Order 3' ,
                'Date Work Created' ,
                'Shop Vendor' ,
                'Memo App Number',
                'Date Forwarded' ,
                'Estimate No' ,
                'Maintenance Amount' ,
                'Less Discount' ,
                'Estimate Remarks' ,
                'Estimate Attached' ,
                'Approved By' ,
                'Meter Reading' ,
                'SLA' ,
                'Email',
                'Date Initiated' ,
                'Sent email',
                'Date email log',
                'Deadline',
                'Status',

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for repair in repair_queryset:
        row_num += 1
        row = [
                repair.request_date,
                repair.employee ,
                repair.cost_center ,
                repair.first_name,
                repair.last_name,
                repair.contact_no ,
                repair.company ,
                repair.department ,
                repair.group_section ,
                repair.plate_no ,
                repair.v_brand ,
                repair.engine ,
                repair.v_make ,
                repair.v_model ,
                repair.chassis ,
                repair.band ,
                repair.cond_sticker ,
                repair.equipment_no ,
                repair.fleet_area ,
                repair.particulars ,
                repair.category ,
                repair.maintenance_type1 ,
                repair.scope_work1 ,
                repair.maintenance_type2 ,
                repair.scope_work2 ,
                repair.recommendations ,
                repair.service_reminder ,
                repair.verified_by ,
                repair.work_order1 ,
                repair.work_order2 ,
                repair.work_order3 ,
                repair.datework_created ,
                repair.Shop_vendor ,
                repair.memo_app,
                repair.date_forwarded ,
                repair.estimate_no ,
                repair.maintenance_amount ,
                repair.less_discount ,
                repair.estimate_remarks ,
                repair.estimate_attached ,
                repair.approvedby ,
                repair.meter_reading ,
                repair.VRR_SLA ,
                repair.email ,
                repair.date_initiated ,
                repair.sent_email,
                repair.Date_email_log ,
                repair.Deadline ,
                repair.status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

##Daily report Details
def vehicle_maintenance_report_details(request):
    date = datetime.datetime.today()
    # notorized = Ownership.objects.filter(status = 'NOTARIZED').count()
    # routing_count = Ownership.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
    # tmg_schedule = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
    # tmg_appearance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
    
    return render(request, 'vehicle_repair/pms_report.html',{'Title: Vehicle Maintenance':'Vehicle Maintenance','date':date})

# Registration Daily Report
def vehicle_maintenance_report(request):
    date = datetime.datetime.today()

    # preventive_count = Vehicle_Repair.objects.filter(status = 'NOTARIZED').count()
    # corrective_count = Corrective.objects.filter(status = 'ON GOING ROUTING FOR APPROVAL').count()
    # tire_battery = Ownership.objects.filter(status = 'WITH_TMG SCHEDULE').count()
    # insurance = Ownership.objects.filter(status = 'FOR TMG APPEARANCE').count()
    
    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Vehicle Maitenance.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Vehicle Maintenance"
    ws['A1'].value = "SUBJECT:"
    ws['B1'].value = "Vehicle Maintenance"
    ws['A3'].value = "PERSONNEL:"
    ws['B3'].value = "Shane Santos"
    ws['A4'].value = "Date"
    ws['A5'].value = ""
    ws.append(['Vehicle Maitenance', 'Total','Remarks'])
    ws['A8'].value = "Preventive Maitenance"
    ws['A9'].value = "Corrective Maitenance"
    ws['A10'].value = "Tires and Battery"
    ws['A11'].value = ""
    ws['A12'].value = ""
    ws['A13'].value = "Insurance Claims"
    #data
    # ws['B4'].value = date
    # ws['B10'].value = routing_count
    # ws['B11'].value = notorized
    # ws['B14'].value = tmg_appearance
    # ws['B15'].value = tmg_schedule
    # ws['B17'].value = with_tmg_etching
    # ws['B19'].value = fleet_vismin
    # ws['B21'].value = lto_transfer
    # ws['B23'].value = total

    # style
    ws['A6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['B6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['C6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['D6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['E6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['F6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['G6'].fill = PatternFill("solid", fgColor="00FFCC99")
    #TMG color
    ws['A13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['B13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['C13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['D13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['E13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['F13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['G13'].fill = PatternFill("solid", fgColor="00FFCC99")
   
    wb.save(output)
    return output



