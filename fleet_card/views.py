from django.shortcuts import render, HttpResponseRedirect, HttpResponse, reverse
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
from django.core import serializers
import datetime
from django.utils import formats
from datetime import date, timedelta
from rest_framework import viewsets
from rest_framework.response import Response
from django.shortcuts import render
from django.db.models import Q, Count
from bootstrap_modal_forms.generic import BSModalDeleteView
# User
from django.core.exceptions import PermissionDenied
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, HttpResponseRedirect
from django.contrib.auth.models import User
from django.contrib.auth.mixins import LoginRequiredMixin


from django.views.generic import (
    ListView,
    CreateView,
    DetailView,
    UpdateView,
)

from . forms import (
    fleet_card_form
)

from .serializers import (
    fleet_card_serializer
)
from .models import (
    fleet_card
)

class fleet_card_list(ListView):
    date = datetime.datetime.today().date()
    print("Date:",date)

    model = fleet_card
    template_name = 'fcm_list.html'

class fleet_card_create(CreateView):
    model = fleet_card
    form_class = fleet_card_form
    template_name = 'fcm_create.html'

class fleet_card_update(UpdateView):
    model = fleet_card
    form_class = fleet_card_form
    template_name = 'fcm_create.html'
    
class fleet_card_delete(BSModalDeleteView):
    model = fleet_card
    template_name = 'fcm_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('Fcm_list')

class fleet_cardDetails(DetailView):
    model = fleet_card
    template_name = 'fcm_details.html'


def fcm_export(request):
    fcm_queryset = fleet_card.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Fleet Card Monitoring.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Fleet Card Monitoring'

    columns = [
        'STATUS',
        'RECEIVED_REQUEST',
        'DATE_VERIFIED',
        'DATE_RECEIVED',
        'DATE_ISSUED',
        'WORK_DAYS',
        'CARD_NUMBER', 
        'NAME',
        'COST_CENTER',
        'PLATE_NUMBER',
        'CARD_TYPE',
        'CABONILLA',
        'STATION'
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for fcm in fcm_queryset:
        row_num += 1
        row = [
            fcm.STATUS,
            fcm.RECEIVED_REQUEST,
            fcm.DATE_VERIFIED,
            fcm.DATE_RECEIVED,
            fcm.DATE_ISSUED,
            fcm.WORK_DAYS,
            fcm.CARD_NUMBER, 
            fcm.NAME,
            fcm.COST_CENTER,
            fcm.PLATE_NUMBER,
            fcm.CARD_TYPE,
            fcm.CABONILLA,
            fcm.STATION
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

# Daily Report
def fleet_summary(request):
    date = datetime.datetime.today().date()
    fleetCard_report = fleet_card.objects.filter(STATUS="ON PROCESS",RECEIVED_REQUEST=date).count()
    fleetCard_issued = fleet_card.objects.filter(DATE_ISSUED=date).count()
    return render(request, 'fcm_report.html',{'title':'Fleet Card','date':date,'fleetCard_report':fleetCard_report,'fleetCard_issued':fleetCard_issued})

def fleet_report(request):
    date = datetime.datetime.today().date()
    print("Date:",date)
    fleetCard_report = fleet_card.objects.filter(STATUS="ON PROCESS",RECEIVED_REQUEST=date).count()
    # fleetCard_cert = fleet_card.objects.filter(car_provider=vendor, sqa_number="N/A").count()
    # fleetCard_pickup = fleet_card.objects.filter(car_provider=vendor, sqa_number="N/A").count()
    # fleetCard_inbound = fleet_card.objects.filter(car_provider=vendor).count()
    fleetCard_issued = fleet_card.objects.filter(DATE_ISSUED=date).count()
    # fleetCard_billing = fleet_card.objects.filter(car_provider=vendor).count()
    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Fleet Card Daily Report.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active
    ws.title = "Fleet Card Daily Report"
    ws['A1'].value = ""
    ws['A2'].value = "Date"
    ws['B2'].value = date
    ws.append(['Fleet Cards', 'TOTAL'])
    ws['A4'].value = ""
    ws['A5'].value = "Fleet Card Processed"
    ws['A6'].value = "Prepared Certification"
    ws['A7'].value = "Prepared for DTD/Pick up Cards"
    ws['A8'].value = "Inbound Fleet Cards"
    ws['A9'].value = "Issued Fleet Cards"
    ws['A10'].value = "Billing Processed"
    ws['A11'].value = "Vehicle Repair Request"
    ws['A12'].value = "PMS/TIRES/BATTERY"
    ws['A13'].value = "CM"


    ws['B5'].value = fleetCard_report
    ws['B6'].value = "0"
    ws['B7'].value = "0"
    ws['B8'].value = "0"
    ws['B9'].value = fleetCard_issued
    ws['B10'].value = "0"
    ws['B11'].value = "0"
    ws['B12'].value = "0"
    ws['B13'].value = "0"


    wb.save(output)
    return output

