from django.shortcuts import render,HttpResponseRedirect,HttpResponse,reverse
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
# from django.core.urlresolvers import reverse_lazy
from django.views.generic import (
     ListView,
     CreateView,
     DetailView,
     UpdateView,
 )
from .models import (
    VehicleMasterList
    )

from masterlist.models import (
    EmployeeMasterlist
    )

from . forms import (
    Vmasterlist,
    Vmaster,
    )
from bootstrap_modal_forms.generic import BSModalDeleteView

from rest_framework import viewsets
from rest_framework.response import Response
# from .models import VehicleMasterList,
from .serializers import (
    vehicleSerializer
    )

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render


def Vmastertables(request):
    return render(request, 'vehicleMasterlist.html')


class vehicleViewSet(viewsets.ModelViewSet):
    queryset = VehicleMasterList.objects.all().order_by('id')
    serializer_class = vehicleSerializer

def VmasterlistCreate(request):
    if request.method == 'POST':
        plate = request.POST.get('plate_no')
        cs = request.POST.get('cs')
        cr_name = request.POST.get('cr_name')
        model = request.POST.get('model')
        brand = request.POST.get('brand')
        vmake = request.POST.get('vmake')
        eng_no = request.POST.get('eng_no')
        chassis_no = request.POST.get('chassis_no')
        mvfile = request.POST.get('mvfile')
        vtype = request.POST.get('vtype')
        vcat = request.POST.get('vcat')
        emp_id = request.POST.get('emp_id')
        band = request.POST.get('band')
        benefit = request.POST.get('benefit')
        cost = request.POST.get('cost')
        group = request.POST.get('group')
        div = request.POST.get('div')
        dept = request.POST.get('dept')
        sec = request.POST.get('sec')
        is_emp = request.POST.get('is_emp')
        IS_NAME = request.POST.get('is_fname')
        # is_lname = request.POST.get('is_fname')
        loc = request.POST.get('loc')
        aqui_date = request.POST.get('aqui_date')
        aqui_cost = request.POST.get('aqui_cost')
        asset = request.POST.get('asset')
        po_no = request.POST.get('po_no')
        sap_pr = request.POST.get('sap_pr')
        ivn_no = request.POST.get('ivn_no')
        mathdoc = request.POST.get('mathdoc')
        eq_no = request.POST.get('eq_no')
        dealer = request.POST.get('dealer')
        dealer_name = request.POST.get('dealer_name')
        plate_date = request.POST.get('plate_date')
        or_date = request.POST.get('or_date')
        remarks = request.POST.get('remarks')
        status = request.POST.get('status')
        COC_Date = request.POST.get('COC_Date')
        Smoke_Emission_Date = request.POST.get('Smoke_Emission_Date')
        Last_Registration_Date = request.POST.get('Last_Registration_Date')
        l_remark = request.POST.get('l_remark')
        email = request.POST.get('email')
        vehicle_status = request.POST.get('vstatus')
        email= request.POST.get('email')
        Status = request.POST.get('Status')
        confirmation = request.POST.get('confirmation')
        smoke = request.POST.get('smoke')
    
        if or_date == '':
            or_date = None
        else:
            or_date = datetime.datetime.strptime(or_date,'%Y-%m-%d')

        if aqui_date == '':
            aqui_date = None
        else:
            aqui_date = datetime.datetime.strptime(aqui_date,'%Y-%m-%d')

        if plate_date == '':
            plate_date = None
        else:
            plate_date = datetime.datetime.strptime(plate_date,'%Y-%m-%d')


        employee_list = ''
        emp_save = ''
        employee_list == EmployeeMasterlist.objects.all()
        for e_id in employee_list:
            if e_id.Employee_Id == emp_id:
                emp_save == e_id
            
        if emp_id == '':
            emp_save == None
            
        reg = ''
        endplate = ''
        if plate != '':
            endplate = int(plate[-1])
            if endplate == 1:
                reg = 'JAN'
            elif endplate == 2:
                reg = 'FEB'
            elif endplate == 3:
                reg = 'MAR'
            elif endplate == 4:
                reg = 'APR'
            elif endplate == 5:
                reg = 'MAY'
            elif endplate == 6:
                reg = 'JUN'
            elif endplate == 7:
                reg = 'JUL'
            elif endplate == 8:
                reg = 'AUG'
            elif endplate == 9:
                reg = 'SEP'
            elif endplate == 0:
                reg = 'OCT'
            
        saveto_end = VehicleMasterList(PLATE_NO=plate, CS_NO=cs, CR_NAME=cr_name, MODEL=model, BRAND=brand,PLATE_ENDING=endplate, REGISTRATION_MONTH=reg,
            VEHICLE_MAKE=vmake, ENGINE_NO=eng_no, CHASSIS_NO=chassis_no, MV_FILE_NO=mvfile, VEHICLE_TYPE=vtype, VEHICLE_CATEGORY=vcat,
            Employee=emp_save, BAND_LEVEL=band, BENEFIT_GROUP=benefit, COST_CENTER=cost, GROUP=group, DIVISION=div,
            DEPARTMENT=dept, SECTION=sec, IS_ID=is_emp, IS_NAME=IS_NAME, LOCATION=loc,
            ACQ_DATE=aqui_date, ACQ_COST=aqui_cost, ASSET_NO=asset, PO_NO=po_no, PLATE_NUMBER_RELEASE_DATE=plate_date, ORIGINAL_OR_DATE=or_date,EQUIPMENT_NO=eq_no,
            SAP_PR=sap_pr,Vehicle_IVN_no=ivn_no,Unit_MATDOC=mathdoc,dealer=dealer, dealer_name=dealer_name, Remarks=remarks, Status=status, leasing_remark=l_remark,
            vehicle_status = vehicle_status,EMAIL=email,COC_Date=COC_Date,Smoke_Emission_Date=Smoke_Emission_Date,Last_Registration_Date=Last_Registration_Date,
            confirmation=confirmation, smoke=smoke
            )
        saveto_end.save()

    return HttpResponseRedirect('/VehicleMasterlist/VehicleMasterlist/')
def vehicleupdate(request):
    elist = EmployeeMasterlist.objects.all()
    return render(request, 'vmasterlist_update.html', {'Title': 'Vehicle Masterlist','elist':elist})

def VmasterlistUpdate(request, pk):
    if request.method == 'POST':
        plate = request.POST.get('plate_no')
        cs = request.POST.get('cs')
        cr_name = request.POST.get('cr_name')
        model = request.POST.get('model')
        brand = request.POST.get('brand')
        vmake = request.POST.get('vmake')
        eng_no = request.POST.get('eng_no')
        chassis_no = request.POST.get('chassis_no')
        mvfile = request.POST.get('mvfile')
        vtype = request.POST.get('vtype')
        vcat = request.POST.get('vcat')
        emp_id = request.POST.get('emp_id')
        band = request.POST.get('band')
        benefit = request.POST.get('benefit')
        cost = request.POST.get('cost')
        group = request.POST.get('group')
        div = request.POST.get('div')
        dept = request.POST.get('dept')
        sec = request.POST.get('sec')
        is_emp = request.POST.get('is_emp')
        IS_NAME = request.POST.get('is_fname')
        loc = request.POST.get('loc')
        aqui_date = request.POST.get('aqui_date')
        aqui_cost = request.POST.get('aqui_cost')
        asset = request.POST.get('asset')
        po_no = request.POST.get('po_no')
        sap_pr = request.POST.get('sap_pr')
        ivn_no = request.POST.get('ivn_no')
        mathdoc = request.POST.get('mathdoc')
        eq_no = request.POST.get('eq_no')
        dealer = request.POST.get('dealer')
        dealer_name = request.POST.get('dealer_name')
        plate_date = request.POST.get('plate_date')
        or_date = request.POST.get('or_date')
        remarks = request.POST.get('remarks')
        l_remark = request.POST.get('l_remark')
        email = request.POST.get('email')
        vehicle_status = request.POST.get('vstatus')
        email = request.POST.get('email')
        COC_Date = request.POST.get('coc_date')
        Smoke_Emission_Date = request.POST.get('Smoke_Emission_Date')
        Last_Registration_Date = request.POST.get('Last_Registration_Date')
        confirmation = request.POST.get('confirmation')
        smoke = request.POST.get('smoke')
        
        if or_date == '':
            or_date = None
        else:
            or_date = datetime.datetime.strptime(or_date,'%Y-%m-%d')

        if aqui_date == '':
            aqui_date = None
        else:
            aqui_date = datetime.datetime.strptime(aqui_date,'%Y-%m-%d')

        if plate_date == '':
            plate_date = None
        else:
            plate_date = datetime.datetime.strptime(plate_date,'%Y-%m-%d')

            
        # if emp_id == '':
        #     emp_save = None
            
        reg = ''
        endplate = ''
        if plate != '':
            endplate = int(plate[-1])
            if endplate == 1:
                reg = 'JAN'
            elif endplate == 2:
                reg = 'FEB'
            elif endplate == 3:
                reg = 'MAR'
            elif endplate == 4:
                reg = 'APR'
            elif endplate == 5:
                reg = 'MAY'
            elif endplate == 6:
                reg = 'JUN'
            elif endplate == 7:
                reg = 'JUL'
            elif endplate == 8:
                reg = 'AUG'
            elif endplate == 9:
                reg = 'SEP'
            elif endplate == 0:
                reg = 'OCT'
        VehicleMasterList.objects.filter(id=pk).update(PLATE_NO=plate, CS_NO=cs, CR_NAME=cr_name, MODEL=model, BRAND=brand,PLATE_ENDING=endplate, REGISTRATION_MONTH=reg,
            VEHICLE_MAKE=vmake, ENGINE_NO=eng_no, CHASSIS_NO=chassis_no, MV_FILE_NO=mvfile, VEHICLE_TYPE=vtype, VEHICLE_CATEGORY=vcat,
            Employee=emp_id, BAND_LEVEL=band, BENEFIT_GROUP=benefit, COST_CENTER=cost, GROUP=group, DIVISION=div,
            DEPARTMENT=dept, SECTION=sec, IS_ID=is_emp, IS_NAME=IS_NAME, LOCATION=loc,
            ACQ_DATE=aqui_date, ACQ_COST=aqui_cost, ASSET_NO=asset, PO_NO=po_no, PLATE_NUMBER_RELEASE_DATE=plate_date, ORIGINAL_OR_DATE=or_date,EQUIPMENT_NO=eq_no,
            SAP_PR=sap_pr,Vehicle_IVN_no=ivn_no,Unit_MATDOC=mathdoc,dealer=dealer, dealer_name=dealer_name, Remarks=remarks, leasing_remark=l_remark,
            vehicle_status = vehicle_status,EMAIL=email,COC_Date=COC_Date,Smoke_Emission_Date=Smoke_Emission_Date,Last_Registration_Date=Last_Registration_Date,
            confirmation=confirmation, smoke=smoke
            )

    return HttpResponseRedirect('/VehicleMasterlist/VehicleMasterlist/')

def vehicle(request):
    elist = EmployeeMasterlist.objects.all()
    return render(request, 'vmasterlist.html', {'Title': 'Vehicle Masterlist','elist':elist})

class vehicleMasterDetails(DetailView):
    model = VehicleMasterList
    template_name = 'vehicleMasterlist_details.html'

class vehicleMasterUpdate(CreateView):
    model = VehicleMasterList
    fields = '__all__'
    template_name = 'vehicleMasterlist_form.html'
    # def get_form(self, *args, **kwargs):
    #     form = super(vehicleMasterUpdate,self).get_form(*args, **kwargs) #instantiate using parent
    #     form.fields['Employee'].queryset = VehicleMasterList.objects.filter(Employee=Employee)
    #     return form

    def form_valid(self, form):
        # This method is called when valid form data has been POSTed.
        if form.instance.PLATE_NO != '':
            endplate = int(form.instance.PLATE_NO[-1])
            if endplate == 1:
                form.instance.REGISTRATION_MONTH = 'JAN'
                form.instance.PLATE_ENDING = '1'
            elif endplate == 2:
                form.instance.REGISTRATION_MONTH = 'FEB'
                form.instance.PLATE_ENDING = '2'
            if endplate == 3:
                form.instance.REGISTRATION_MONTH = 'MAR'
                form.instance.PLATE_ENDING = '3'
            elif endplate == 4:
                form.instance.REGISTRATION_MONTH = 'APR'
                form.instance.PLATE_ENDING = '4'
            if endplate == 5:
                form.instance.REGISTRATION_MONTH = 'MAY'
                form.instance.PLATE_ENDING = '5'
            elif endplate == 6:
                form.instance.REGISTRATION_MONTH = 'JUN'
                form.instance.PLATE_ENDING = '6'
            if endplate == 7:
                form.instance.REGISTRATION_MONTH = 'JUL'
                form.instance.PLATE_ENDING = '7'
            elif endplate == 8:
                form.instance.REGISTRATION_MONTH = 'AUG'
                form.instance.PLATE_ENDING = '8'
            if endplate == 9:
                form.instance.REGISTRATION_MONTH = 'SEP'
                form.instance.PLATE_ENDING = '9'
            elif endplate == 0:
                form.instance.REGISTRATION_MONTH = 'OCT'
                form.instance.PLATE_ENDING = '0'

        return super().form_valid(form)


class confirmationDetails(DetailView):
    model = VehicleMasterList
    template_name = 'confirmation.html'

def vmUpdate(request,id):
    if request.method == 'POST':
        confirmation=request.POST.get('confirmation')
        smoke=request.POST.get('smoke')
    VehicleMasterList.objects.filter(id=pk).update(confirmation=confirmation,smoke=smoke)
    return HttpResponseRedirect('/VehicleMasterlist/VehicleMasterlist/')

class releaseUpdate(UpdateView):
    model = VehicleMasterList
    form_class = Vmaster
    template_name = 'registration/release_date.html'

class vehicleMasterlistDeleteView(BSModalDeleteView):
    model = VehicleMasterList
    template_name = 'vehicleMasterlist_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('vehicle-list')

def vehicleMasterlistHistoryView(request):
    if request.method == "GET":
       obj = VehicleMasterList.history.all()

       return render(request, 'vehicleMasterlist_history.html', context={'object': obj})

def vehicle_masterlist_active(request):
    # context = {
    #         'vehicle_list_active': VehicleMasterList.objects.filter(vehicle_status__contains='Active')[:100]
    #     }

    return render(request, 'vehicle_masterlist_active.html')

def vehicle_masterlist_solved(request):
    # context = {
    #         'vehicle_list_solved': VehicleMasterList.objects.prefetch_related('Sold').order_by('vehicle_status')[:100]
    #         # 'vehicle_list_solved':VehicleMasterList.objects.raw('SELECT * FROM VehicleMasterList')[0]
    #     }

    return render(request, 'vehicle_masterlist_solved.html')

def vehicle_masterlist_globe(request):

    return render(request, 'vehicle_masterlist_globe.html')

def vehicle_masterlist_trans(request):
    context = {
            'vehicle_list_trans': VehicleMasterList.objects.filter(vehicle_status__contains='Transferred')
        }

    return render(request, 'vehicle_masterlist_trans.html', context)



class vreg_details(DetailView):
    model = VehicleMasterList
    template_name = 'vreg_details.html'
    

def vehicle_bayan(request):
    context = {
            'bayan_list': VehicleMasterList.objects.filter(CR_NAME__contains="BAYANTEL")
        }

    return render(request, 'vehicle_bayan.html', context)
def vehicle_telicphil(request):
    context = {
            'teli_list': VehicleMasterList.objects.filter(CR_NAME__contains="TELICPHIL")
        }

    return render(request, 'vehicle_telicphil.html', context)

def vehicle_excel(request):
    v_queryset = VehicleMasterList.objects.all()  
    v_queryset_active = VehicleMasterList.objects.filter(vehicle_status='Active')   
    v_queryset_sold = VehicleMasterList.objects.filter(vehicle_status='Sold')
    v_queryset_trans = VehicleMasterList.objects.filter(vehicle_status='Transferred') 
    v_queryset_BAYANTEL = VehicleMasterList.objects.filter(CR_NAME__contains="BAYANTEL") 
    v_queryset_teli = VehicleMasterList.objects.filter(CR_NAME__contains="TELICPHIL") 

    from openpyxl import Workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Vehicle Masterlist.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active
    ws.title = "All data"

    columns = [
            'ID',
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'CHECKLIST BY',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
            'Vehicle Status',
            'confirmation',
            'smoke',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset:
        row_num += 1
        # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                vehicle.Activity_Id,
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.CHECKLIST_BY,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
                vehicle.vehicle_status,
                vehicle.confirmation,
                vehicle.smoke,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    ws1 = wb.create_sheet("Vehicle Active")
    ws1.title = "Vehicle Active"
    columns = [
            'ID',
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'CHECKLIST BY',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
            'Vehicle Status',
            'EMAIL',
            'Status',
            'Status_2',
            'Status_3',
            'Status_4',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws1.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset_active:
        row_num += 1
        row = [
                vehicle.Activity_Id,
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.CHECKLIST_BY,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
                vehicle.vehicle_status,
                vehicle.EMAIL,
                vehicle.Status,
                vehicle.Status_2,
                vehicle.Status_3,
                vehicle.Status_4,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = cell_value

    ws2 = wb.create_sheet("Vehicle Sold")
    ws2.title = "Vehicle Sold"

    columns = [
            'ID',
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'CHECKLIST BY',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Email',
            'Other Remarks',
            'Vehicle Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws2.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset_sold:
        row_num += 1
        row = [
                vehicle.Activity_Id,
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.CHECKLIST_BY,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.EMAIL,
                vehicle.leasing_remark,
                vehicle.vehicle_status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    ws3 = wb.create_sheet("Vehicle Transferred")
    ws3.title = "Vehicle Transferred"

    columns = [
            'ID',
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'CHECKLIST BY',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Email',
            'Other Remarks',
            'Vehicle Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws3.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset_trans:
        row_num += 1
        row = [
                vehicle.Activity_Id,
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.CHECKLIST_BY,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.EMAIL,
                vehicle.leasing_remark,
                vehicle.vehicle_status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws3.cell(row=row_num, column=col_num)
            cell.value = cell_value

      
    ws4 = wb.create_sheet("Vehicle BAYANTEL")
    ws4.title = "Vehicle BAYANTEL"

    columns = [
            'ID',
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_LAST_NAME',
            # 'IS_FIRST_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws4.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset_BAYANTEL:
        row_num += 1
        # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                vehicle.Activity_Id,
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                # vehicle.IS_FIRST_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws4.cell(row=row_num, column=col_num)
            cell.value = cell_value
      
    ws5 = wb.create_sheet("Vehicle TELICPHIL")
    ws5.title = "Vehicle TELICPHIL"

    columns = [
            'ID',
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            # 'IS_FIRST_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws5.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset_teli:
        row_num += 1
        # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                vehicle.Activity_Id,
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                # vehicle.IS_FIRST_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws5.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(output)
    return output





###############################

# def vehicle_excel(request):
#     v_queryset = VehicleMasterList.objects.all()   
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist.xlsx'
#     workbook = Workbook()

#     worksheet = workbook.active
#     worksheet.title = 'Vehicle Masterlist'

#     columns = [
            
#             'NO',
#             'PLATE_NO',
#             'CS_NO',
#             'CR_NAME',
#             'PLATE_ENDING',
#             'REGISTRATION_MONTH',
#             'MODEL',
#             'BRAND',
#             'VEHICLE_MAKE',
#             'ENGINE_NO',
#             'CHASSIS_NO',
#             'MV_FILE_NO',
#             'VEHICLE_TYPE',
#             'Employee',
#             'ASSIGNEE_LAST_NAME',
#             'ASSIGNEE_FIRST_NAME',
#             'VEHICLE_CATEGORY',
#             'BAND_LEVEL', 
#             'BENEFIT_GROUP',
#             'COST_CENTER',
#             'GROUP',
#             'DIVISION',
#             'DEPARTMENT',
#             'SECTION',
#             'IS_ID',
#             'IS_NAME',
#             'LOCATION',
#             'ORIGINAL_OR_DATE',
#             'ACQ_DATE',
#             'ACQ_COST',
#             'ASSET_NO',
#             'EQUIPMENT_NO',
#             'SAP_PR',
#             'Vehicle_IVN_no',
#             'Unit_MATDOC',
#             'dealer',
#             'dealer_name',
#             'PO_NO',
#             'CHECKLIST BY',
#             'PLATE_NUMBER_RELEASE_DATE',
#             'Last_Registration_Date',
#             'Smoke_Emission_Date',
#             'Smoke_due',
#             'COC_Date',
#             'Remarks',
#             'Status',
#             'Other Remarks',
#             'Vehicle Status',
#     ]
#     row_num = 1

#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title

#     for vehicle in v_queryset:
#         row_num += 1
#         # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
#         # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
#         row = [
#                 vehicle.NO,
#                 vehicle.PLATE_NO,
#                 vehicle.CS_NO,
#                 vehicle.CR_NAME,
#                 vehicle.PLATE_ENDING,
#                 vehicle.REGISTRATION_MONTH,
#                 vehicle.MODEL,
#                 vehicle.BRAND,
#                 vehicle.VEHICLE_MAKE,
#                 vehicle.ENGINE_NO,
#                 vehicle.CHASSIS_NO,
#                 vehicle.MV_FILE_NO,
#                 vehicle.VEHICLE_TYPE,
#                 vehicle.Employee,
#                 vehicle.ASSIGNEE_LAST_NAME,
#                 vehicle.ASSIGNEE_FIRST_NAME,
#                 vehicle.VEHICLE_CATEGORY,
#                 vehicle.BAND_LEVEL, 
#                 vehicle.BENEFIT_GROUP,
#                 vehicle.COST_CENTER,
#                 vehicle.GROUP,
#                 vehicle.DIVISION,
#                 vehicle.DEPARTMENT,
#                 vehicle.SECTION,
#                 vehicle.IS_ID,
#                 vehicle.IS_NAME,
#                 vehicle.LOCATION,
#                 vehicle.ORIGINAL_OR_DATE,
#                 vehicle.ACQ_DATE,
#                 vehicle.ACQ_COST,
#                 vehicle.ASSET_NO,
#                 vehicle.EQUIPMENT_NO,
#                 vehicle.SAP_PR,
#                 vehicle.Vehicle_IVN_no,
#                 vehicle.Unit_MATDOC,
#                 vehicle.dealer,
#                 vehicle.dealer_name,
#                 vehicle.PO_NO,
#                 vehicle.CHECKLIST_BY,
#                 vehicle.PLATE_NUMBER_RELEASE_DATE,
#                 vehicle.Last_Registration_Date,
#                 vehicle.Smoke_Emission_Date,
#                 vehicle.Smoke_due,
#                 vehicle.COC_Date,
#                 vehicle.Remarks,
#                 vehicle.Status,
#                 vehicle.leasing_remark,
#                 vehicle.vehicle_status,
#         ]
        
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value

#     workbook.save(response)
#     return response
    

def vehicle_excel_active(request):
    v_queryset = VehicleMasterList.objects.filter(vehicle_status='Active')   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Active.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Masterlist Active'

    columns = [
            
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'CHECKLIST BY',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
            'Vehicle Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset:
        row_num += 1
        row = [
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.CHECKLIST_BY,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
                vehicle.vehicle_status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def vehicle_excel_solved(request):
    v_queryset = VehicleMasterList.objects.filter(vehicle_status='Sold')   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Sold.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Masterlist Sold'

    columns = [
            
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'CHECKLIST BY',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
            'Vehicle Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset:
        row_num += 1
        row = [
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.CHECKLIST_BY,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
                vehicle.vehicle_status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
    
def vehicle_excel_trans(request):
    v_queryset = VehicleMasterList.objects.filter(vehicle_status='Transferred')   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Transferred.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Masterlist Transferred'

    columns = [
            
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'CHECKLIST BY',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
            'Vehicle Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset:
        row_num += 1
        row = [
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.CHECKLIST_BY,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
                vehicle.vehicle_status,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def vehicle_excel_bayan(request):
    v_queryset = VehicleMasterList.objects.filter(CR_NAME__contains="BAYANTEL")   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Bayan.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Masterlist Bayan'

    columns = [
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_LAST_NAME',
            # 'IS_FIRST_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset:
        row_num += 1
        # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                # vehicle.IS_FIRST_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
    
def vehicle_excel_teli(request):
    v_queryset = VehicleMasterList.objects.filter(CR_NAME__contains="TELICPHIL")   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Masterlist Telicphil.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Masterlist Telicphil'

    columns = [
            'NO',
            'PLATE_NO',
            'CS_NO',
            'CR_NAME',
            'PLATE_ENDING',
            'REGISTRATION_MONTH',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'ENGINE_NO',
            'CHASSIS_NO',
            'MV_FILE_NO',
            'VEHICLE_TYPE',
            'Employee',
            'ASSIGNEE_LAST_NAME',
            'ASSIGNEE_FIRST_NAME',
            'VEHICLE_CATEGORY',
            'BAND_LEVEL', 
            'BENEFIT_GROUP',
            'COST_CENTER',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_ID',
            'IS_NAME',
            # 'IS_FIRST_NAME',
            'LOCATION',
            'ORIGINAL_OR_DATE',
            'ACQ_DATE',
            'ACQ_COST',
            'ASSET_NO',
            'EQUIPMENT_NO',
            'SAP_PR',
            'Vehicle_IVN_no',
            'Unit_MATDOC',
            'dealer',
            'dealer_name',
            'PO_NO',
            'PLATE_NUMBER_RELEASE_DATE',
            'Last_Registration_Date',
            'Smoke_Emission_Date',
            'Smoke_due',
            'COC_Date',
            'Remarks',
            'Status',
            'Other Remarks',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in v_queryset:
        row_num += 1
        # ordate = vehicle.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = vehicle.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                vehicle.NO,
                vehicle.PLATE_NO,
                vehicle.CS_NO,
                vehicle.CR_NAME,
                vehicle.PLATE_ENDING,
                vehicle.REGISTRATION_MONTH,
                vehicle.MODEL,
                vehicle.BRAND,
                vehicle.VEHICLE_MAKE,
                vehicle.ENGINE_NO,
                vehicle.CHASSIS_NO,
                vehicle.MV_FILE_NO,
                vehicle.VEHICLE_TYPE,
                vehicle.Employee,
                vehicle.ASSIGNEE_LAST_NAME,
                vehicle.ASSIGNEE_FIRST_NAME,
                vehicle.VEHICLE_CATEGORY,
                vehicle.BAND_LEVEL, 
                vehicle.BENEFIT_GROUP,
                vehicle.COST_CENTER,
                vehicle.GROUP,
                vehicle.DIVISION,
                vehicle.DEPARTMENT,
                vehicle.SECTION,
                vehicle.IS_ID,
                vehicle.IS_NAME,
                # vehicle.IS_FIRST_NAME,
                vehicle.LOCATION,
                vehicle.ORIGINAL_OR_DATE,
                vehicle.ACQ_DATE,
                vehicle.ACQ_COST,
                vehicle.ASSET_NO,
                vehicle.EQUIPMENT_NO,
                vehicle.SAP_PR,
                vehicle.Vehicle_IVN_no,
                vehicle.Unit_MATDOC,
                vehicle.dealer,
                vehicle.dealer_name,
                vehicle.PO_NO,
                vehicle.PLATE_NUMBER_RELEASE_DATE,
                vehicle.Last_Registration_Date,
                vehicle.Smoke_Emission_Date,
                vehicle.Smoke_due,
                vehicle.COC_Date,
                vehicle.Remarks,
                vehicle.Status,
                vehicle.leasing_remark,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
