from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
# Create your views here.
from django.views import generic
from django.shortcuts import render,HttpResponseRedirect, get_list_or_404,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
import datetime
from datetime import date, timedelta
from .models import (
		battery,
)
from masterlist.models import (
    EmployeeMasterlist
    )
from vehicle_masterlist.models import VehicleMasterList
from leasingmasterlist.models import Leasing
from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
# from django.views.generic.edit import (
    # DeleteView,
# )
from . forms import batteryform
from bootstrap_modal_forms.generic import BSModalDeleteView

class batteryListView(ListView):
    model = battery
    template_name='battery_list.html'

def batterycreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'battery_create.html',{'Title':'Battery - Battery','emplist':emplist,'vlist':vlist})

def batterysubmit(request):
    if request.method == 'POST':
        request_date = request.POST.get('request_date')
        emp_id = request.POST.get('emp_id')
        cost_center = request.POST.get('cost_center')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        c_no = request.POST.get('c_no')
        company = request.POST.get('company')
        department = request.POST.get('department')
        group = request.POST.get('group')
        plate_no = request.POST.get('plate_no')
        v_brand = request.POST.get('v_brand')
        engine = request.POST.get('engine')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        chassis = request.POST.get('chassis')
        v_band = request.POST.get('v_band')
        cs_no = request.POST.get('cs_no')
        eq_no = request.POST.get('eq_no')
        fleet_area = request.POST.get('fleet_area')
        particulars = request.POST.get('particulars')
        category = request.POST.get('category')
        maintenance_type1 = request.POST.get('maintenance_type1')
        maintenance_type2 = request.POST.get('maintenance_type2')
        scope_work1 = request.POST.get('scope_work1')
        scope_work2 = request.POST.get('scope_work2')
        recomendations = request.POST.get('recomendations')
        service_reminder = request.POST.get('service_reminder')
        repair_verified_by = request.POST.get('repair_verified_by')
        work_order1 = request.POST.get('work_order1')
        work_order2 = request.POST.get('work_order2')
        work_order3 = request.POST.get('work_order3')
        date_work_created = request.POST.get('date_work_created')
        repair_shop = request.POST.get('repair_shop')
        memo_app = request.POST.get('memo_app')
        date_forward = request.POST.get('date_forward')
        estimate_no = request.POST.get('estimate_no')
        maintenance_amount = request.POST.get('maintenance_amount')
        less_discount = request.POST.get('less_discount')
        estimate_remark = request.POST.get('estimate_remark')
        estimate_attach = request.POST.get('estimate_attach')
        approved_by = request.POST.get('approved_by')
        kilo_reading = request.POST.get('kilo_reading')
        email = request.POST.get('email')
        email_status = request.POST.get('email_status')
        vrr_sla = request.POST.get('vrr_sla')

        deadl = datetime.datetime.strptime(request_date, '%Y-%m-%d')
        dead_date = deadl + datetime.timedelta(days=30)
        # print(dead_date)

        saveto_batterry = battery(request_date=request_date, employee=emp_id, cost_center=cost_center, first_name=fname,
            last_name=lname, contact_no=c_no, company=company, department=department, group_section=group,
            plate_no=plate_no, v_brand=v_brand, engine=engine, v_make=v_make, v_model=v_model, chassis=chassis,
            band=v_band, cond_sticker=cs_no, equipment_no=eq_no, fleet_area=fleet_area, particulars=particulars,
            category=category, maintenance_type1=maintenance_type1, scope_work1=scope_work1, maintenance_type2=maintenance_type2,
            scope_work2=scope_work2, recommendations=recomendations, service_reminder=service_reminder, verified_by=repair_verified_by,
            work_order1=work_order1, work_order2=work_order2, work_order3=work_order3, datework_created=date_work_created,
            Shop_vendor=repair_shop, date_forwarded=date_forward, estimate_no=estimate_no, maintenance_amount=maintenance_amount,
            less_discount=less_discount, estimate_remarks=estimate_remark, estimate_attached=estimate_attach, approvedby=approved_by,
            meter_reading=kilo_reading, VRR_SLA=vrr_sla, memo_app=memo_app,email=email,sent_email=email_status,Deadline=dead_date
        )
        saveto_batterry.save()

        return HttpResponseRedirect('/Battery/List')

class batteryDelete(BSModalDeleteView):
    model = battery
    template_name = 'battery_delete.html'
    success_message = 'Success: Battery Maintenance was deleted.'
    success_url = reverse_lazy('battery_List')


class batteryUpdateView(SuccessMessageMixin, UpdateView):
    model = battery
    form_class = batteryform
    template_name = 'battery_update.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Battery Maintenance Updated Successfully!"

class batteryDetailView(DetailView):
    model = battery
    template_name = 'battery_details.html'

def battery_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = battery.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = battery.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = battery.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=3))
    dl4 = battery.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=4))
    dl5 = battery.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=5))
    dl6 = battery.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'battery_deadline.html',{'title':'Battery - Battery Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3, 'dl4':dl4, 'dl5':dl5, 'dl6':dl6})



def battery_excel(request):
    repair_queryset = battery.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Battery Maintenance.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Battery Maintenance'

    columns = [
                'Request Date' ,
                'Employee' ,
                'Cost Center' ,
                'First Name' ,
                'Last Name' ,
                'Contact No' ,
                'Company' ,
                'Department' ,
                'Group Section' ,
                'Plate No' ,
                'Brand' ,
                'Engine' ,
                'Make' ,
                'Model' ,
                'Chassis' ,
                'Band' ,
                'Conduction Sticker' ,
                'Equipment No' ,
                'Fleet Area' ,
                'Particulars' ,
                'Category' ,
                'Maintenance Type 1' ,
                'Scope Work 1' ,
                'Maintenance Type 2' ,
                'Scope Work 2' ,
                'Recommendations' ,
                'Service Reminder' ,
                'Verified By' ,
                'Work Order 1' ,
                'Work Order 2' ,
                'Work Order 3' ,
                'Date Work Created' ,
                'Shop Vendor' ,
                'Memo App Number',
                'Date Forwarded' ,
                'Estimate No' ,
                'Maintenance Amount' ,
                'Less Discount' ,
                'Estimate Remarks' ,
                'Estimate Attached' ,
                'Approved By' ,
                'Meter Reading' ,
                'SLA' ,
                'Email',
                'Date Initiated' ,
                'Sent email',
                'Date email log',
                'Deadline',

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for repair in repair_queryset:
        row_num += 1
        row = [
                repair.request_date,
                repair.employee ,
                repair.cost_center ,
                repair.first_name,
                repair.last_name,
                repair.contact_no ,
                repair.company ,
                repair.department ,
                repair.group_section ,
                repair.plate_no ,
                repair.v_brand ,
                repair.engine ,
                repair.v_make ,
                repair.v_model ,
                repair.chassis ,
                repair.band ,
                repair.cond_sticker ,
                repair.equipment_no ,
                repair.fleet_area ,
                repair.particulars ,
                repair.category ,
                repair.maintenance_type1 ,
                repair.scope_work1 ,
                repair.maintenance_type2 ,
                repair.scope_work2 ,
                repair.recommendations ,
                repair.service_reminder ,
                repair.verified_by ,
                repair.work_order1 ,
                repair.work_order2 ,
                repair.work_order3 ,
                repair.datework_created ,
                repair.Shop_vendor ,
                repair.memo_app,
                repair.date_forwarded ,
                repair.estimate_no ,
                repair.maintenance_amount ,
                repair.less_discount ,
                repair.estimate_remarks ,
                repair.estimate_attached ,
                repair.approvedby ,
                repair.meter_reading ,
                repair.VRR_SLA ,
                repair.email ,
                repair.date_initiated ,
                repair.sent_email,
                repair.Date_email_log ,
                repair.Deadline ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def batteryHistoryView(request):
    if request.method == "GET":
       obj = battery.history.all()

       return render(request, 'battery_history.html', context={'object': obj})

# Registration Daily Report
def battery_report(request):
    date = datetime.datetime.today()
    from openpyxl import Workbook, load_workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Vehicle Maitenance.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active

    #header
    ws.title = "Vehicle Maintenance"
    ws['A1'].value = "SUBJECT:"
    ws['B1'].value = "Vehicle Maintenance"
    ws['A3'].value = "PERSONNEL:"
    ws['B3'].value = "Shane Santos"
    ws['A4'].value = "Date"
    ws['A5'].value = ""
    ws.append(['Vehicle Maitenance', 'Total','Remarks'])
    ws['A8'].value = "Preventive Maitenance"
    ws['A9'].value = "Corrective Maitenance"
    ws['A10'].value = "Tires and Battery"
    ws['A11'].value = ""
    ws['A12'].value = ""
    ws['A13'].value = "Insurance Claims"
    #data
    # ws['B4'].value = date
    # ws['B10'].value = routing_count
    # ws['B11'].value = notorized
    # ws['B14'].value = tmg_appearance
    # ws['B15'].value = tmg_schedule
    # ws['B17'].value = with_tmg_etching
    # ws['B19'].value = fleet_vismin
    # ws['B21'].value = lto_transfer
    # ws['B23'].value = total

    # style
    ws['A6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['B6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['C6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['D6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['E6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['F6'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['G6'].fill = PatternFill("solid", fgColor="00FFCC99")
    #TMG color
    ws['A13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['B13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['C13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['D13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['E13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['F13'].fill = PatternFill("solid", fgColor="00FFCC99")
    ws['G13'].fill = PatternFill("solid", fgColor="00FFCC99")
   
    wb.save(output)
    return output


