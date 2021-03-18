from django.shortcuts import render,HttpResponseRedirect,HttpResponse,reverse
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
# from django.core.urlresolvers import reverse_lazy
from django.views.generic import (
     ListView,
     CreateView,
     DetailView,
     UpdateView,
 )
from .models import (
    Leasing
    )
from . forms import (
    leasing_form
    )
from bootstrap_modal_forms.generic import BSModalDeleteView

from rest_framework import viewsets
from rest_framework.response import Response
# from .models import VehicleMasterList,
from .serializers import (
    leasingSerializer
    )

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render


 #######################################  
##############  Leasing  ################
 #######################################


def leasing(request):
    return render(request, 'leasing_list.html')


class leasingViewSet(viewsets.ModelViewSet):
    queryset = Leasing.objects.all().order_by('id')
    serializer_class = leasingSerializer


class leasingCreateView(CreateView):
    model = Leasing
    form_class = leasing_form
    template_name = 'leasing_form.html'


class leasingUpdateView(UpdateView):
    model = Leasing
    form_class = leasing_form
    template_name = 'leasing_form.html'


class leasingDetailView(DetailView):
    model = Leasing
    template_name = 'leasing_details.html'


def leasingHistoryView(request):
    if request.method == "GET":
       obj = Leasing.history.all()

       return render(request, 'leasing_history.html', context={'object': obj})



## Leasing ##
def vehicle_leasing_active(request):
    context = {
            'leasing_list_active': Leasing.objects.filter(vleasing_status__contains='Active')
        }

    return render(request, 'leasing_active.html', context)

def vehicle_leasing_solved(request):
    context = {
            'leasing_list_solved': Leasing.objects.filter(vleasing_status__contains='Sold')
        }

    return render(request, 'leasing_solved.html', context)

def vehicle_leasing_trans(request):
    context = {
            'leasing_list_trans': Leasing.objects.filter(vleasing_status__contains='Transferred')
        }

    return render(request, 'leasing_trans.html', context)

def leasing_export(request):
    leasing_queryset = Leasing.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Leasing Masterlist'

    columns = [

            'Activity_Id',
            'PLATE_NUMBER',
            'CS_NO',
            'COMPANY',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'VEHICLE_TYPE',
            'LAST_NAME_ASSIGNEE',
            'FIRST_NAME_ASSIGNEE',
            'VEHICLE_CATEGORY',
            'COST_CENTER',
            'ID_NUMBER',
            'BAND',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_EMPLOYEE_ID',
            'IS_NAME',
            'IS_FIRSTNAME',
            'LOCATION',
            'AREA',
            'ACQUISITION_DATE',
            'remarks',
            'acquisition_cost',
            'months_36',
            'amount1',
            'date_in_1',
            'date_out_1',
            'months_24',
            'amount_Vat_EX',
            'date_in_2',
            'date_out_2',
            'extension',
            'amount2',
            'date_in_3',
            'date_out_3',
            'chasis_no',
            'engine_no',
            'CONTRACT_NUMBER',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for leasing in leasing_queryset:
        row_num += 1
        row = [
            leasing.Activity_Id,
            leasing.PLATE_NUMBER,
            leasing.CS_NO,
            leasing.COMPANY,
            leasing.MODEL,
            leasing.BRAND,
            leasing.VEHICLE_MAKE,
            leasing.VEHICLE_TYPE,
            leasing.LAST_NAME_ASSIGNEE,
            leasing.FIRST_NAME_ASSIGNEE,
            leasing.VEHICLE_CATEGORY,
            leasing.COST_CENTER,
            leasing.ID_NUMBER,
            leasing.BAND,
            leasing.GROUP,
            leasing.DIVISION,
            leasing.DEPARTMENT,
            leasing.SECTION,
            leasing.IS_EMPLOYEE_ID,
            leasing.IS_LASTNAME,
            leasing.IS_FIRSTNAME,
            leasing.LOCATION,
            leasing.AREA,
            leasing.ACQUISITION_DATE,
            leasing.remarks,
            leasing.acquisition_cost,
            leasing.months_36,
            leasing.amount1,
            leasing.date_in_1,
            leasing.date_out_1,
            leasing.months_24,
            leasing.amount_Vat_EX,
            leasing.date_in_2,
            leasing.date_out_2,
            leasing.extension,
            leasing.amount2,
            leasing.date_in_3,
            leasing.date_out_3,
            leasing.chasis_no,
            leasing.engine_no,
            leasing.CONTRACT_NUMBER,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def leasing_active_export(request):
    leasing_queryset = Leasing.objects.filter(vleasing_status='Active')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist Active.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Leasing Masterlist Active'

    columns = [

            'Activity_Id',
            'PLATE_NUMBER',
            'CS_NO',
            'COMPANY',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'VEHICLE_TYPE',
            'LAST_NAME_ASSIGNEE',
            'FIRST_NAME_ASSIGNEE',
            'VEHICLE_CATEGORY',
            'COST_CENTER',
            'ID_NUMBER',
            'BAND',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_EMPLOYEE_ID',
            'IS_NAME',
            'IS_FIRSTNAME',
            'LOCATION',
            'AREA',
            'ACQUISITION_DATE',
            'remarks',
            'acquisition_cost',
            'months_36',
            'amount1',
            'date_in_1',
            'date_out_1',
            'months_24',
            'amount_Vat_EX',
            'date_in_2',
            'date_out_2',
            'extension',
            'amount2',
            'date_in_3',
            'date_out_3',
            'chasis_no',
            'engine_no',
            'CONTRACT_NUMBER',
            'Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for leasing in leasing_queryset:
        row_num += 1
        row = [
            leasing.Activity_Id,
            leasing.PLATE_NUMBER,
            leasing.CS_NO,
            leasing.COMPANY,
            leasing.MODEL,
            leasing.BRAND,
            leasing.VEHICLE_MAKE,
            leasing.VEHICLE_TYPE,
            leasing.LAST_NAME_ASSIGNEE,
            leasing.FIRST_NAME_ASSIGNEE,
            leasing.VEHICLE_CATEGORY,
            leasing.COST_CENTER,
            leasing.ID_NUMBER,
            leasing.BAND,
            leasing.GROUP,
            leasing.DIVISION,
            leasing.DEPARTMENT,
            leasing.SECTION,
            leasing.IS_EMPLOYEE_ID,
            leasing.IS_LASTNAME,
            leasing.IS_FIRSTNAME,
            leasing.LOCATION,
            leasing.AREA,
            leasing.ACQUISITION_DATE,
            leasing.remarks,
            leasing.acquisition_cost,
            leasing.months_36,
            leasing.amount1,
            leasing.date_in_1,
            leasing.date_out_1,
            leasing.months_24,
            leasing.amount_Vat_EX,
            leasing.date_in_2,
            leasing.date_out_2,
            leasing.extension,
            leasing.amount2,
            leasing.date_in_3,
            leasing.date_out_3,
            leasing.chasis_no,
            leasing.engine_no,
            leasing.CONTRACT_NUMBER,
            leasing.vleasing_status,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def leasing_solved_export(request):
    leasing_queryset = Leasing.objects.filter(vleasing_status='Sold')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist Sold.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Leasing Masterlist Sold'

    columns = [

            'Activity_Id',
            'PLATE_NUMBER',
            'CS_NO',
            'COMPANY',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'VEHICLE_TYPE',
            'LAST_NAME_ASSIGNEE',
            'FIRST_NAME_ASSIGNEE',
            'VEHICLE_CATEGORY',
            'COST_CENTER',
            'ID_NUMBER',
            'BAND',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_EMPLOYEE_ID',
            'IS_NAME',
            'IS_FIRSTNAME',
            'LOCATION',
            'AREA',
            'ACQUISITION_DATE',
            'remarks',
            'acquisition_cost',
            'months_36',
            'amount1',
            'date_in_1',
            'date_out_1',
            'months_24',
            'amount_Vat_EX',
            'date_in_2',
            'date_out_2',
            'extension',
            'amount2',
            'date_in_3',
            'date_out_3',
            'chasis_no',
            'engine_no',
            'CONTRACT_NUMBER',
            'Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for leasing in leasing_queryset:
        row_num += 1
        row = [
            leasing.Activity_Id,
            leasing.PLATE_NUMBER,
            leasing.CS_NO,
            leasing.COMPANY,
            leasing.MODEL,
            leasing.BRAND,
            leasing.VEHICLE_MAKE,
            leasing.VEHICLE_TYPE,
            leasing.LAST_NAME_ASSIGNEE,
            leasing.FIRST_NAME_ASSIGNEE,
            leasing.VEHICLE_CATEGORY,
            leasing.COST_CENTER,
            leasing.ID_NUMBER,
            leasing.BAND,
            leasing.GROUP,
            leasing.DIVISION,
            leasing.DEPARTMENT,
            leasing.SECTION,
            leasing.IS_EMPLOYEE_ID,
            leasing.IS_LASTNAME,
            leasing.IS_FIRSTNAME,
            leasing.LOCATION,
            leasing.AREA,
            leasing.ACQUISITION_DATE,
            leasing.remarks,
            leasing.acquisition_cost,
            leasing.months_36,
            leasing.amount1,
            leasing.date_in_1,
            leasing.date_out_1,
            leasing.months_24,
            leasing.amount_Vat_EX,
            leasing.date_in_2,
            leasing.date_out_2,
            leasing.extension,
            leasing.amount2,
            leasing.date_in_3,
            leasing.date_out_3,
            leasing.chasis_no,
            leasing.engine_no,
            leasing.CONTRACT_NUMBER,
            leasing.vleasing_status,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def leasing_trans_export(request):
    leasing_queryset = Leasing.objects.filter(vleasing_status='Transferred')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Leasing Masterlist Transferred.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Leasing Masterlist Transferred'

    columns = [

            'Activity_Id',
            'PLATE_NUMBER',
            'CS_NO',
            'COMPANY',
            'MODEL',
            'BRAND',
            'VEHICLE_MAKE',
            'VEHICLE_TYPE',
            'LAST_NAME_ASSIGNEE',
            'FIRST_NAME_ASSIGNEE',
            'VEHICLE_CATEGORY',
            'COST_CENTER',
            'ID_NUMBER',
            'BAND',
            'GROUP',
            'DIVISION',
            'DEPARTMENT',
            'SECTION',
            'IS_EMPLOYEE_ID',
            'IS_NAME',
            'IS_FIRSTNAME',
            'LOCATION',
            'AREA',
            'ACQUISITION_DATE',
            'remarks',
            'acquisition_cost',
            'months_36',
            'amount1',
            'date_in_1',
            'date_out_1',
            'months_24',
            'amount_Vat_EX',
            'date_in_2',
            'date_out_2',
            'extension',
            'amount2',
            'date_in_3',
            'date_out_3',
            'chasis_no',
            'engine_no',
            'CONTRACT_NUMBER',
            'Status',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for leasing in leasing_queryset:
        row_num += 1
        row = [
            leasing.Activity_Id,
            leasing.PLATE_NUMBER,
            leasing.CS_NO,
            leasing.COMPANY,
            leasing.MODEL,
            leasing.BRAND,
            leasing.VEHICLE_MAKE,
            leasing.VEHICLE_TYPE,
            leasing.LAST_NAME_ASSIGNEE,
            leasing.FIRST_NAME_ASSIGNEE,
            leasing.VEHICLE_CATEGORY,
            leasing.COST_CENTER,
            leasing.ID_NUMBER,
            leasing.BAND,
            leasing.GROUP,
            leasing.DIVISION,
            leasing.DEPARTMENT,
            leasing.SECTION,
            leasing.IS_EMPLOYEE_ID,
            leasing.IS_LASTNAME,
            leasing.IS_FIRSTNAME,
            leasing.LOCATION,
            leasing.AREA,
            leasing.ACQUISITION_DATE,
            leasing.remarks,
            leasing.acquisition_cost,
            leasing.months_36,
            leasing.amount1,
            leasing.date_in_1,
            leasing.date_out_1,
            leasing.months_24,
            leasing.amount_Vat_EX,
            leasing.date_in_2,
            leasing.date_out_2,
            leasing.extension,
            leasing.amount2,
            leasing.date_in_3,
            leasing.date_out_3,
            leasing.chasis_no,
            leasing.engine_no,
            leasing.CONTRACT_NUMBER,
            leasing.vleasing_status,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

