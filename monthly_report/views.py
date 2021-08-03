from django.shortcuts import render,HttpResponseRedirect,HttpResponse, redirect
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
import datetime
from datetime import date, timedelta
from django.urls import reverse_lazy

from .serializers import (
    petron_report_Serializer
    )
from . models import Petron_report
from rest_framework import viewsets
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                                DeleteView,
                				)
from django.db.models import Sum
from .serializers import petron_report_Serializer
from . forms import petron_form


def monthly_report_jan(request):
	jan_data = Petron_report.objects.all()
	return render(request, 'monthly_report/petron_report.html',{'title':'Report','jan_data':jan_data})


class petronViewSet(viewsets.ModelViewSet):
    queryset = Petron_report.objects.all().order_by('id')
    serializer_class = petron_report_Serializer

class petron_create(CreateView):
    model = Petron_report
    form_class = petron_form
    template_name = 'monthly_report/petron_create.html'

class petron_update(UpdateView):
    model = Petron_report
    form_class = petron_form
    template_name = 'monthly_report/petron_create.html'

class petron_details(DetailView):
    model = Petron_report
    template_name = 'monthly_report/petron_detail.html'

class petron_delete(DeleteView):
    model = Petron_report
    success_url = reverse_lazy('jan_monthly_report')
    template_name = 'monthly_report/petron_delete.html'

def monthly_report_jan_summary(request):

    date = datetime.datetime.today()
    months = ['zero','January','February','March','April','May','June','July','August','September','October','November','December']
    month = months[date.month]
    ## ProductAmount ####
    BB14_B1 = Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B10").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B11= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B11").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B7= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B7").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B8= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B8").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG12_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG12_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_B1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C5_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C5_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_E1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_F2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_F3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG3-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_B15= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B15").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-F").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG5_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG5_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG6_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG6_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG7_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG7-F").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG8_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG8-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CRA6_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CRA6-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C10").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C9= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C9").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN9_C0201= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0201").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN9_C0301= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0301").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN9_C0305= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0305").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_H= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-H").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_I= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-I").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_K= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-K").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_L= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-L").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG05_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG09_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GRTM-C503").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    ISG18_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="ISG18-C2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT11_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT11-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-B6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_C5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D01_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D01-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D08_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D08-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D2_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D2_05= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-05").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-04").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-06").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D4_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D4-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D5_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D5_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D6_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6-06").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D03_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D03-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D7_01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D7_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D7_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_G2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_G4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_A01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-A01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_B02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_B06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B06").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_C01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-C01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D02A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D02B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D2-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_E01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_E03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_E04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E04").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_F01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_F02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_F03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_H01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_H03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_I01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_I03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_J01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_J03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT6-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    OP12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP12-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    OP20_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP20-F1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_O= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-O").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_P= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-P").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_Q= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-Q").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_R= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-R").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_U= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_T1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-T1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0

    ## ProductAmount end ####

    ##Discount_Amount

    dis_BB14_B1 = Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B10").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B11= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B11").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B7= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B7").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B8= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B8").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG12_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG12_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_B1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C5_C= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C5_D= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_E1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_F2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_F3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG3-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_B15= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B15").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-F").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG5_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG5_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG6_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG6_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG7_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG7-F").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG8_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG8-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CRA6_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CRA6-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C10").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C9= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C9").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN9_C0201= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0201").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN9_C0301= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0301").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN9_C0305= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0305").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_H= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-H").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_I= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-I").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_K= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-K").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_L= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-L").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG05_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG09_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GRTM-C503").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_ISG18_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="ISG18-C2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT11_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT11-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-B6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_C5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D01_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D01-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D08_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D08-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D2_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D2_05= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-05").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-04").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-06").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D4_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D4-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D5_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D5_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D6_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6-06").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D03_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D03_B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D7_01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D7_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D7_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_G2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_G4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_A01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-A01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_B02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_B06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B06").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_C01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-C01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D02A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D02B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D2-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_E01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_E03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_E04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E04").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_F01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_F02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_F03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_H01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_H03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_I01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_I03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_J01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_J03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT6-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_OP12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP12-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_OP20_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP20-F1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_O= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-O").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_P= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-P").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_Q= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-Q").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_R= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-R").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_U= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_T1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-T1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    
    ## Discount_Amount end ######

    ### Net Amount #####
    net_BB14_B1 = Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B10").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B11= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B11").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B7= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B7").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B8= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B8").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG12_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG12_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_B1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C5_D= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C5_C= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_E1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_F2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_F3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG3-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_B15= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B15").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-F").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG5_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG5_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG6_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG6_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG7_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG7-F").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG8_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG8-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CRA6_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CRA6-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C10").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C9= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C9").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN9_C0201= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0201").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN9_C0301= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0301").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN9_C0305= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0305").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_H= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-H").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_I= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-I").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_K= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-K").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_L= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-L").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG05_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG09_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GRTM-C503").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_ISG18_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="ISG18-C2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT11_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT11-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_C5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-B6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D01_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D01-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D03_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D03-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D08_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D08-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D2_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D2_05= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-05").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-04").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-06").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D4_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D4-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D5_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D5_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D6_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6-06").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D7_01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D7_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D7_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_G2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_G4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_A01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-A01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_B02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_B06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B06").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_C01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-C01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D02A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D02B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D2-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_E01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_E03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_E04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E04").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_F01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_F02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_F03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_H01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_H03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_I01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_I03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_J01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_J03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT6_C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_OP12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP12-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_OP20_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP20-F1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_O= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-O").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_P= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-P").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_Q= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-Q").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_R= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-R").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_U= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_T1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-T1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    ### Net Amount End#####
    grand_total = BB14_B1+BB14_B10+BB14_B11+BB14_B2+BB14_B3+BB14_B4+BB14_B5
    +BB14_B6+BB14_B7+BB14_B8+BB14_C+BB14_E+CMG12_A+CMG12_C
    +CMG12_D+CMG2_B1+CMG2_B2+CMG2_B3+CMG2_C1+CMG2_C2+CMG2_C3
    +CMG2_C4+CMG2_C5_C+CMG2_C5_D+CMG2_D1+CMG2_D2+CMG2_D3+CMG2_E1
    +CMG2_E2+CMG2_E3+CMG2_F1+CMG2_F2+CMG2_F3+CMG3_A+CMG4_B
    +CMG4_B15+CMG4_B2+CMG4_E+CMG4_F+CMG4_G+CMG5_B+CMG5_E+CMG5_G
    +CMG6_B+CMG6_C+CMG6_E+CMG7_F+CMG8_B+CRA6_A+FIN23_C1+FIN23_C10
    +FIN23_C4+FIN23_C6+FIN23_C9+FIN9_C0201+FIN9_C0301+FIN9_C0305
    +GEG02_B+GEG02_C+GEG02_D+GEG02_E+GEG02_G+GEG02_H+GEG02_I+GEG02_K
    +GEG02_L+GEG04_B3+GEG05_E+GEG09_D+GRTM_C503+ISG18_C2+NTT11_E
    +NTT2_B6+NTT2_C3+NTT2_C5+NTT2_C6+NTT2_D01_A+NTT2_D08_C+NTT2_D1
    +NTT2_D2_02+NTT2_D2_03+NTT2_D2_05+NTT2_D3_02+NTT2_D3_03+NTT2_D3_04
    +NTT2_D3_06+NTT2_D4_03+NTT2_D5_02+NTT2_D5_03+NTT2_D6+NTT2_D6_06
    +NTT2_D03_B+NTT2_D7_01+NTT2_D7_02+NTT2_D7_03+NTT2_E2+NTT2_E3
    +NTT2_F3A+NTT2_F3B+NTT2_F3D+NTT2_F3E+NTT2_G2+NTT2_G4+NTT3_A
    +NTT3_B5+NTT3_C+NTT3_D2+NTT3_D3+NTT3_D4+NTT3_D5+NTT3_B3
    +NTT5_A01+NTT5_B02+NTT5_B06+NTT5_C01+NTT5_D01+NTT5_D02A
    +NTT5_D02B+NTT5_D2_03+NTT5_E01+NTT5_E03+NTT5_E04+NTT5_F01
    +NTT5_F02+NTT5_F03+NTT5_G+NTT5_G01+NTT5_G02+NTT5_G03+NTT5_H01
    +NTT5_H03+NTT5_I01+NTT5_I03+NTT5_J01+NTT5_J03+NTT6_C+OP12_A
    +OP20_F1+SG02_O+SG02_P+SG02_Q+SG02_R+SG02_U+GEG02_F+GRTM_C503
    +SG02_T1+GEG04_B3

    dis_grand_total = dis_BB14_B1+dis_BB14_B10+dis_BB14_B11+dis_BB14_B2+dis_BB14_B3+dis_BB14_B4+dis_BB14_B5
    +dis_BB14_B6+dis_BB14_B7+dis_BB14_B8+dis_BB14_C+dis_BB14_E+dis_CMG12_A+dis_CMG12_C
    +dis_CMG12_D+dis_CMG2_B1+dis_CMG2_B2+dis_CMG2_B3+dis_CMG2_C1+dis_CMG2_C2+dis_CMG2_C3
    +dis_CMG2_C4+dis_CMG2_C5_C+dis_CMG2_C5_D+dis_CMG2_D1+dis_CMG2_D2+dis_CMG2_D3+dis_CMG2_E1
    +dis_CMG2_E2+dis_CMG2_E3+dis_CMG2_F1+dis_CMG2_F2+dis_CMG2_F3+dis_CMG3_A+dis_CMG4_B
    +dis_CMG4_B15+dis_CMG4_B2+dis_CMG4_E+dis_CMG4_F+dis_CMG4_G+dis_CMG5_B+dis_CMG5_E+dis_CMG5_G
    +dis_CMG6_B+dis_CMG6_C+dis_CMG6_E+dis_CMG7_F+dis_CMG8_B+dis_CRA6_A+dis_FIN23_C1+dis_FIN23_C10
    +dis_FIN23_C4+dis_FIN23_C6+dis_FIN23_C9+dis_FIN9_C0201+dis_FIN9_C0301+dis_FIN9_C0305
    +dis_GEG02_B+dis_GEG02_C+dis_GEG02_D+dis_GEG02_E+dis_GEG02_G+dis_GEG02_H+dis_GEG02_I+dis_GEG02_K
    +dis_GEG02_L+dis_GEG04_B3+dis_GEG05_E+dis_GEG09_D+dis_GRTM_C503+dis_ISG18_C2+dis_NTT11_E
    +dis_NTT2_B6+dis_NTT2_C3+dis_NTT2_C5+dis_NTT2_C6+dis_NTT2_D01_A+dis_NTT2_D08_C+dis_NTT2_D1
    +dis_NTT2_D2_02+dis_NTT2_D2_03+dis_NTT2_D2_05+dis_NTT2_D3_02+dis_NTT2_D3_03+dis_NTT2_D3_04
    +dis_NTT2_D3_06+dis_NTT2_D4_03+dis_NTT2_D5_02+dis_NTT2_D5_03+dis_NTT2_D6+dis_NTT2_D6_06
    +dis_NTT2_D03_B+dis_NTT2_D7_01+dis_NTT2_D7_02+dis_NTT2_D7_03+dis_NTT2_E2+dis_NTT2_E3
    +dis_NTT2_F3A+dis_NTT2_F3B+dis_NTT2_F3D+dis_NTT2_F3E+dis_NTT2_G2+dis_NTT2_G4+dis_NTT3_A
    +dis_NTT3_B5+dis_NTT3_C+dis_NTT3_D2+dis_NTT3_D3+dis_NTT3_D4+dis_NTT3_D5+dis_NTT3_B3
    +dis_NTT5_A01+dis_NTT5_B02+dis_NTT5_B06+dis_NTT5_C01+dis_NTT5_D01+dis_NTT5_D02A
    +dis_NTT5_D02B+dis_NTT5_D2_03+dis_NTT5_E01+dis_NTT5_E03+dis_NTT5_E04+dis_NTT5_F01
    +dis_NTT5_F02+dis_NTT5_F03+dis_NTT5_G+dis_NTT5_G01+dis_NTT5_G02+dis_NTT5_G03+dis_NTT5_H01
    +dis_NTT5_H03+dis_NTT5_I01+dis_NTT5_I03+dis_NTT5_J01+dis_NTT5_J03+dis_NTT6_C+dis_OP12_A
    +dis_OP20_F1+dis_SG02_O+dis_SG02_P+dis_SG02_Q+dis_SG02_R+dis_SG02_U+dis_GEG02_F+dis_GRTM_C503
    +dis_SG02_T1+dis_GEG04_B3

    net_grand_total = net_BB14_B1+net_BB14_B10+net_BB14_B11+net_BB14_B2+net_BB14_B3+net_BB14_B4+net_BB14_B5
    +net_BB14_B6+net_BB14_B7+net_BB14_B8+net_BB14_C+net_BB14_E+net_CMG12_A+net_CMG12_C
    +net_CMG12_D+net_CMG2_B1+net_CMG2_B2+net_CMG2_B3+net_CMG2_C1+net_CMG2_C2+net_CMG2_C3
    +net_CMG2_C4+net_CMG2_C5_C+net_CMG2_C5_D+net_CMG2_D1+net_CMG2_D2+net_CMG2_D3+net_CMG2_E1
    +net_CMG2_E2+net_CMG2_E3+net_CMG2_F1+net_CMG2_F2+net_CMG2_F3+net_CMG3_A+net_CMG4_B
    +net_CMG4_B15+net_CMG4_B2+net_CMG4_E+net_CMG4_F+net_CMG4_G+net_CMG5_B+net_CMG5_E+net_CMG5_G
    +net_CMG6_B+net_CMG6_C+net_CMG6_E+net_CMG7_F+net_CMG8_B+net_CRA6_A+net_FIN23_C1+net_FIN23_C10
    +net_FIN23_C4+net_FIN23_C6+net_FIN23_C9+net_FIN9_C0201+net_FIN9_C0301+net_FIN9_C0305
    +net_GEG02_B+net_GEG02_C+net_GEG02_D+net_GEG02_E+net_GEG02_G+net_GEG02_H+net_GEG02_I+net_GEG02_K
    +net_GEG02_L+net_GEG04_B3+net_GEG05_E+net_GEG09_D+net_GRTM_C503+net_ISG18_C2+net_NTT11_E
    +net_NTT2_B6+net_NTT2_C3+net_NTT2_C5+net_NTT2_C6+net_NTT2_D01_A+net_NTT2_D08_C+net_NTT2_D1
    +net_NTT2_D2_02+net_NTT2_D2_03+net_NTT2_D2_05+net_NTT2_D3_02+net_NTT2_D3_03+net_NTT2_D3_04
    +net_NTT2_D3_06+net_NTT2_D4_03+net_NTT2_D5_02+net_NTT2_D5_03+net_NTT2_D6+net_NTT2_D6_06
    +net_NTT2_D03_B+net_NTT2_D7_01+net_NTT2_D7_02+net_NTT2_D7_03+net_NTT2_E2+net_NTT2_E3
    +net_NTT2_F3A+net_NTT2_F3B+net_NTT2_F3D+net_NTT2_F3E+net_NTT2_G2+net_NTT2_G4+net_NTT3_A
    +net_NTT3_B5+net_NTT3_C+net_NTT3_D2+net_NTT3_D3+net_NTT3_D4+net_NTT3_D5+net_NTT3_B3
    +net_NTT5_A01+net_NTT5_B02+net_NTT5_B06+net_NTT5_C01+net_NTT5_D01+net_NTT5_D02A
    +net_NTT5_D02B+net_NTT5_D2_03+net_NTT5_E01+net_NTT5_E03+net_NTT5_E04+net_NTT5_F01
    +net_NTT5_F02+net_NTT5_F03+net_NTT5_G+net_NTT5_G01+net_NTT5_G02+net_NTT5_G03+net_NTT5_H01
    +net_NTT5_H03+net_NTT5_I01+net_NTT5_I03+net_NTT5_J01+net_NTT5_J03+net_NTT6_C+net_OP12_A
    +net_OP20_F1+net_SG02_O+net_SG02_P+net_SG02_Q+net_SG02_R+net_SG02_U+net_GEG02_F+net_GRTM_C503
    +net_SG02_T1+net_GEG04_B3
    return render(request,'monthly_report/petron_report_summary.html',{'title':'Petron Data', 'BB14_B1':BB14_B1,
    'BB14_B10':BB14_B10, 'BB14_B11':BB14_B11, 'BB14_B2':BB14_B2, 'BB14_B3':BB14_B3, 'BB14_B4':BB14_B4, 'BB14_B5':BB14_B5, 'BB14_B6':BB14_B6,
    'BB14_B7':BB14_B7, 'BB14_B8':BB14_B8, 'BB14_C':BB14_C, 'BB14_E':BB14_E, 'CMG12_C':CMG12_C, 
    'CMG12_D':CMG12_D, 'CMG2_B1':CMG2_B1, 'CMG2_B2':CMG2_B2, 'CMG2_B3':CMG2_B3, 'CMG2_C1':CMG2_C1, 'CMG2_C2':CMG2_C2, 
    'CMG2_C3':CMG2_C3, 'CMG2_C4':CMG2_C4, 'CMG2_C5_C':CMG2_C5_C, 'CMG2_D1':CMG2_D1, 'CMG2_D2':CMG2_D2, 'CMG2_D3':CMG2_D3, 'CMG2_E1':CMG2_E1, 
    'CMG2_E2':CMG2_E2, 'CMG2_E3':CMG2_E3, 'CMG2_F1':CMG2_F1, 'CMG2_F2':CMG2_F2, 'CMG2_F3':CMG2_F3, 'CMG3_A':CMG3_A, 'CMG4_B':CMG4_B, 
    'CMG4_B15':CMG4_B15, 'CMG4_B2':CMG4_B2, 'CMG4_E':CMG4_E, 'CMG4_F':CMG4_F, 'CMG4_G':CMG4_G, 'CMG5_B':CMG5_B, 'CMG5_E':CMG5_E, 
    'CMG5_G':CMG5_G, 'CMG6_B':CMG6_B, 'CMG6_C':CMG6_C, 'CMG6_E':CMG6_E, 'CMG7_F':CMG7_F, 'CMG8_B':CMG8_B, 'CRA6_A':CRA6_A, 'FIN23_C1':FIN23_C1, 
    'FIN23_C10':FIN23_C10, 'FIN23_C4':FIN23_C4, 'FIN23_C9':FIN23_C9, 'FIN9_C0201':FIN9_C0201, 'FIN9_C0301':FIN9_C0301, 'GEG02_B':GEG02_B, 
    'GEG02_C':GEG02_C, 'GEG02_D':GEG02_D, 'GEG02_E':GEG02_E, 'GEG02_G':GEG02_G, 'GEG02_H':GEG02_H, 'GEG02_I':GEG02_I, 'GEG02_K':GEG02_K, 
    'GEG02_L':GEG02_L, 'GEG04_B3':GEG04_B3, 'GEG05_E':GEG05_E, 'GEG09_D':GEG09_D, 'GRTM_C503':GRTM_C503, 'ISG18_C2':ISG18_C2, 'NTT11_E':NTT11_E, 
    'NTT2_C3':NTT2_C3, 'NTT2_C5':NTT2_C5, 'NTT2_C6':NTT2_C6, 'NTT2_D01_A':NTT2_D01_A, 'NTT2_D08_C':NTT2_D08_C, 'NTT2_D1':NTT2_D1, 'NTT2_D2_02':NTT2_D2_02, 
    'NTT2_D2_03':NTT2_D2_03, 'NTT2_D2_05':NTT2_D2_05, 'NTT2_D3_02':NTT2_D3_02, 'NTT2_D3_03':NTT2_D3_03, 'NTT2_D3_04':NTT2_D3_04, 'NTT2_D3_06':NTT2_D3_06, 
    'NTT2_D4_03':NTT2_D4_03, 'NTT2_D5_02':NTT2_D5_02, 'NTT2_D5_03':NTT2_D5_03, 'NTT2_D6':NTT2_D6, 'NTT2_D6_06':NTT2_D6_06, 'NTT2_D7_01':NTT2_D7_01, 
    'NTT2_D7_02':NTT2_D7_02, 'NTT2_D7_03':NTT2_D7_03, 'NTT2_E2':NTT2_E2, 'NTT2_E3':NTT2_E3, 'NTT2_F3A':NTT2_F3A, 'NTT2_F3B':NTT2_F3B, 'NTT2_F3D':NTT2_F3D, 
    'NTT2_F3E':NTT2_F3E, 'NTT2_G2':NTT2_G2, 'NTT2_G4':NTT2_G4, 'NTT3_A':NTT3_A, 'NTT3_B5':NTT3_B5, 'NTT3_C':NTT3_C, 'NTT3_D2':NTT3_D2, 'NTT3_D3':NTT3_D3, 
    'NTT3_D4':NTT3_D4, 'NTT3_D5':NTT3_D5, 'NTT5_A01':NTT5_A01, 'NTT5_B02':NTT5_B02, 'NTT5_B06':NTT5_B06, 'NTT5_C01':NTT5_C01, 'NTT5_D01':NTT5_D01, 
    'NTT5_D02A':NTT5_D02A, 'NTT5_D02B':NTT5_D02B, 'NTT5_E03':NTT5_E03, 'NTT5_E04':NTT5_E04, 'NTT5_F01':NTT5_F01, 'NTT5_F02':NTT5_F02, 'NTT5_F03':NTT5_F03, 
    'NTT5_G':NTT5_G, 'NTT5_G01':NTT5_G01, 'NTT5_G02':NTT5_G02, 'NTT5_G03':NTT5_G03, 'NTT5_H01':NTT5_H01, 'NTT5_H03':NTT5_H03, 'NTT5_I01':NTT5_I01, 
    'NTT5_I03':NTT5_I03, 'NTT5_J01':NTT5_J01, 'NTT5_J03':NTT5_J03, 'OP12_A':OP12_A, 'OP20_F1':OP20_F1, 'SG02_O':SG02_O, 'SG02_P':SG02_P, 'SG02_Q':SG02_Q, 
    'SG02_R':SG02_R, 'SG02_U':SG02_U, 'dis_BB14_B1':dis_BB14_B1,
    'dis_BB14_B10':dis_BB14_B10, 'dis_BB14_B11':dis_BB14_B11, 'dis_BB14_B2':dis_BB14_B2, 'dis_BB14_B3':dis_BB14_B3, 'dis_BB14_B4':dis_BB14_B4, 'dis_BB14_B5':dis_BB14_B5, 'dis_BB14_B6':dis_BB14_B6,
    'dis_BB14_B7':dis_BB14_B7, 'dis_BB14_B8':dis_BB14_B8, 'dis_BB14_C':dis_BB14_C, 'dis_BB14_E':dis_BB14_E, 'dis_CMG12_C':dis_CMG12_C, 
    'dis_CMG12_D':dis_CMG12_D, 'dis_CMG2_B1':dis_CMG2_B1, 'dis_CMG2_B2':dis_CMG2_B2, 'dis_CMG2_B3':dis_CMG2_B3, 'dis_CMG2_C1':dis_CMG2_C1, 'dis_CMG2_C2':dis_CMG2_C2, 
    'dis_CMG2_C3':dis_CMG2_C3, 'dis_CMG2_C4':dis_CMG2_C4, 'dis_CMG2_C5_C':dis_CMG2_C5_C, 'dis_CMG2_D1':dis_CMG2_D1, 'dis_CMG2_D2':dis_CMG2_D2, 'dis_CMG2_D3':dis_CMG2_D3, 'dis_CMG2_E1':dis_CMG2_E1, 
    'dis_CMG2_E2':dis_CMG2_E2, 'dis_CMG2_E3':dis_CMG2_E3, 'dis_CMG2_F1':dis_CMG2_F1, 'dis_CMG2_F2':dis_CMG2_F2, 'dis_CMG2_F3':dis_CMG2_F3, 'dis_CMG3_A':dis_CMG3_A, 'dis_CMG4_B':dis_CMG4_B, 
    'dis_CMG4_B15':dis_CMG4_B15, 'dis_CMG4_B2':dis_CMG4_B2, 'dis_CMG4_E':dis_CMG4_E, 'dis_CMG4_F':dis_CMG4_F, 'dis_CMG4_G':dis_CMG4_G, 'dis_CMG5_B':dis_CMG5_B, 'dis_CMG5_E':dis_CMG5_E, 
    'dis_CMG5_G':dis_CMG5_G, 'dis_CMG6_B':dis_CMG6_B, 'dis_CMG6_C':dis_CMG6_C, 'dis_CMG6_E':dis_CMG6_E, 'dis_CMG7_F':dis_CMG7_F, 'dis_CMG8_B':dis_CMG8_B, 'dis_CRA6_A':dis_CRA6_A, 'dis_FIN23_C1':dis_FIN23_C1, 
    'dis_FIN23_C10':dis_FIN23_C10, 'dis_FIN23_C4':dis_FIN23_C4, 'dis_FIN23_C9':dis_FIN23_C9, 'dis_FIN9_C0201':dis_FIN9_C0201, 'dis_FIN9_C0301':dis_FIN9_C0301, 'dis_GEG02_B':dis_GEG02_B, 
    'dis_GEG02_C':dis_GEG02_C, 'dis_GEG02_D':dis_GEG02_D, 'dis_GEG02_E':dis_GEG02_E, 'dis_GEG02_G':dis_GEG02_G, 'dis_GEG02_H':dis_GEG02_H, 'dis_GEG02_I':dis_GEG02_I, 'dis_GEG02_K':dis_GEG02_K, 
    'dis_GEG02_L':dis_GEG02_L, 'dis_GEG04_B3':dis_GEG04_B3, 'dis_GEG05_E':dis_GEG05_E, 'dis_GEG09_D':dis_GEG09_D, 'dis_GRTM_C503':dis_GRTM_C503, 'dis_ISG18_C2':dis_ISG18_C2, 'dis_NTT11_E':dis_NTT11_E, 
    'dis_NTT2_C3':dis_NTT2_C3, 'dis_NTT2_C5':dis_NTT2_C5, 'dis_NTT2_C6':dis_NTT2_C6, 'dis_NTT2_D01_A':dis_NTT2_D01_A, 'dis_NTT2_D08_C':dis_NTT2_D08_C, 'dis_NTT2_D1':dis_NTT2_D1, 'dis_NTT2_D2_02':dis_NTT2_D2_02, 
    'dis_NTT2_D2_03':dis_NTT2_D2_03, 'dis_NTT2_D2_05':dis_NTT2_D2_05, 'dis_NTT2_D3_02':dis_NTT2_D3_02, 'dis_NTT2_D3_03':dis_NTT2_D3_03, 'dis_NTT2_D3_04':dis_NTT2_D3_04, 'dis_NTT2_D3_06':dis_NTT2_D3_06, 
    'dis_NTT2_D4_03':dis_NTT2_D4_03, 'dis_NTT2_D5_02':dis_NTT2_D5_02, 'dis_NTT2_D5_03':dis_NTT2_D5_03, 'dis_NTT2_D6':dis_NTT2_D6, 'dis_NTT2_D6_06':dis_NTT2_D6_06, 'dis_NTT2_D7_01':dis_NTT2_D7_01, 
    'dis_NTT2_D7_02':dis_NTT2_D7_02, 'dis_NTT2_D7_03':dis_NTT2_D7_03, 'dis_NTT2_E2':dis_NTT2_E2, 'dis_NTT2_E3':dis_NTT2_E3, 'dis_NTT2_F3A':dis_NTT2_F3A, 'dis_NTT2_F3B':dis_NTT2_F3B, 'dis_NTT2_F3D':dis_NTT2_F3D, 
    'dis_NTT2_F3E':dis_NTT2_F3E, 'dis_NTT2_G2':dis_NTT2_G2, 'dis_NTT2_G4':dis_NTT2_G4, 'dis_NTT3_A':dis_NTT3_A, 'dis_NTT3_B5':dis_NTT3_B5, 'dis_NTT3_C':dis_NTT3_C, 'dis_NTT3_D2':dis_NTT3_D2, 'dis_NTT3_D3':dis_NTT3_D3, 
    'dis_NTT3_D4':dis_NTT3_D4, 'dis_NTT3_D5':dis_NTT3_D5, 'dis_NTT5_A01':dis_NTT5_A01, 'dis_NTT5_B02':dis_NTT5_B02, 'dis_NTT5_B06':dis_NTT5_B06, 'dis_NTT5_C01':dis_NTT5_C01, 'dis_NTT5_D01':dis_NTT5_D01, 
    'dis_NTT5_D02A':dis_NTT5_D02A, 'dis_NTT5_D02B':dis_NTT5_D02B, 'dis_NTT5_E03':dis_NTT5_E03, 'dis_NTT5_E04':dis_NTT5_E04, 'dis_NTT5_F01':dis_NTT5_F01, 'dis_NTT5_F02':dis_NTT5_F02, 'dis_NTT5_F03':dis_NTT5_F03, 
    'dis_NTT5_G':dis_NTT5_G, 'dis_NTT5_G01':dis_NTT5_G01, 'dis_NTT5_G02':dis_NTT5_G02, 'dis_NTT5_G03':dis_NTT5_G03, 'dis_NTT5_H01':dis_NTT5_H01, 'dis_NTT5_H03':dis_NTT5_H03, 'dis_NTT5_I01':dis_NTT5_I01, 
    'dis_NTT5_I03':dis_NTT5_I03, 'dis_NTT5_J01':dis_NTT5_J01, 'dis_NTT5_J03':dis_NTT5_J03, 'dis_OP12_A':dis_OP12_A, 'dis_OP20_F1':dis_OP20_F1, 'dis_SG02_O':dis_SG02_O, 'dis_SG02_P':dis_SG02_P, 'dis_SG02_Q':dis_SG02_Q, 
    'dis_SG02_R':dis_SG02_R, 'dis_SG02_U':dis_SG02_U, 'net_BB14_B1':net_BB14_B1,
    'net_BB14_B10':net_BB14_B10, 'net_BB14_B11':net_BB14_B11, 'net_BB14_B2':net_BB14_B2, 'net_BB14_B3':net_BB14_B3, 'net_BB14_B4':net_BB14_B4, 'net_BB14_B5':net_BB14_B5, 'net_BB14_B6':net_BB14_B6,
    'net_BB14_B7':net_BB14_B7, 'net_BB14_B8':net_BB14_B8, 'net_BB14_C':net_BB14_C, 'net_BB14_E':net_BB14_E, 'net_CMG12_C':net_CMG12_C, 
    'net_CMG12_D':net_CMG12_D, 'net_CMG2_B1':net_CMG2_B1, 'net_CMG2_B2':net_CMG2_B2, 'net_CMG2_B3':net_CMG2_B3, 'net_CMG2_C1':net_CMG2_C1, 'net_CMG2_C2':net_CMG2_C2, 
    'net_CMG2_C3':net_CMG2_C3, 'net_CMG2_C4':net_CMG2_C4, 'net_CMG2_C5_C':net_CMG2_C5_C, 'net_CMG2_D1':net_CMG2_D1, 'net_CMG2_D2':net_CMG2_D2, 'net_CMG2_D3':net_CMG2_D3, 'net_CMG2_E1':net_CMG2_E1, 
    'net_CMG2_E2':net_CMG2_E2, 'net_CMG2_E3':net_CMG2_E3, 'net_CMG2_F1':net_CMG2_F1, 'net_CMG2_F2':net_CMG2_F2, 'net_CMG2_F3':net_CMG2_F3, 'net_CMG3_A':net_CMG3_A, 'net_CMG4_B':net_CMG4_B, 
    'net_CMG4_B15':net_CMG4_B15, 'net_CMG4_B2':net_CMG4_B2, 'net_CMG4_E':net_CMG4_E, 'net_CMG4_F':net_CMG4_F, 'net_CMG4_G':net_CMG4_G, 'net_CMG5_B':net_CMG5_B, 'net_CMG5_E':net_CMG5_E, 
    'net_CMG5_G':net_CMG5_G, 'net_CMG6_B':net_CMG6_B, 'net_CMG6_C':net_CMG6_C, 'net_CMG6_E':net_CMG6_E, 'net_CMG7_F':net_CMG7_F, 'net_CMG8_B':net_CMG8_B, 'net_CRA6_A':net_CRA6_A, 'net_FIN23_C1':net_FIN23_C1, 
    'net_FIN23_C10':net_FIN23_C10, 'net_FIN23_C4':net_FIN23_C4, 'net_FIN23_C9':net_FIN23_C9, 'net_FIN9_C0201':net_FIN9_C0201, 'net_FIN9_C0301':net_FIN9_C0301, 'net_GEG02_B':net_GEG02_B, 
    'net_GEG02_C':net_GEG02_C, 'net_GEG02_D':net_GEG02_D, 'net_GEG02_E':net_GEG02_E, 'net_GEG02_G':net_GEG02_G, 'net_GEG02_H':net_GEG02_H, 'net_GEG02_I':net_GEG02_I, 'net_GEG02_K':net_GEG02_K, 
    'net_GEG02_L':net_GEG02_L, 'net_GEG04_B3':net_GEG04_B3, 'net_GEG05_E':net_GEG05_E, 'net_GEG09_D':net_GEG09_D, 'net_GRTM_C503':net_GRTM_C503, 'net_ISG18_C2':net_ISG18_C2, 'net_NTT11_E':net_NTT11_E, 
    'net_NTT2_C3':net_NTT2_C3, 'net_NTT2_C5':net_NTT2_C5, 'net_NTT2_C6':net_NTT2_C6, 'net_NTT2_D01_A':net_NTT2_D01_A, 'net_NTT2_D08_C':net_NTT2_D08_C, 'net_NTT2_D1':net_NTT2_D1, 'net_NTT2_D2_02':net_NTT2_D2_02, 
    'net_NTT2_D2_03':net_NTT2_D2_03, 'net_NTT2_D2_05':net_NTT2_D2_05, 'net_NTT2_D3_02':net_NTT2_D3_02, 'net_NTT2_D3_03':net_NTT2_D3_03, 'net_NTT2_D3_04':net_NTT2_D3_04, 'net_NTT2_D3_06':net_NTT2_D3_06, 
    'net_NTT2_D4_03':net_NTT2_D4_03, 'net_NTT2_D5_02':net_NTT2_D5_02, 'net_NTT2_D5_03':net_NTT2_D5_03, 'net_NTT2_D6':net_NTT2_D6, 'net_NTT2_D6_06':net_NTT2_D6_06, 'net_NTT2_D7_01':net_NTT2_D7_01, 
    'net_NTT2_D7_02':net_NTT2_D7_02, 'net_NTT2_D7_03':net_NTT2_D7_03, 'net_NTT2_E2':net_NTT2_E2, 'net_NTT2_E3':net_NTT2_E3, 'net_NTT2_F3A':net_NTT2_F3A, 'net_NTT2_F3B':net_NTT2_F3B, 'net_NTT2_F3D':net_NTT2_F3D, 
    'net_NTT2_F3E':net_NTT2_F3E, 'net_NTT2_G2':net_NTT2_G2, 'net_NTT2_G4':net_NTT2_G4, 'net_NTT3_A':net_NTT3_A, 'net_NTT3_B5':net_NTT3_B5, 'net_NTT3_C':net_NTT3_C, 'net_NTT3_D2':net_NTT3_D2, 'net_NTT3_D3':net_NTT3_D3, 
    'net_NTT3_D4':net_NTT3_D4, 'net_NTT3_D5':net_NTT3_D5, 'net_NTT5_A01':net_NTT5_A01, 'net_NTT5_B02':net_NTT5_B02, 'net_NTT5_B06':net_NTT5_B06, 'net_NTT5_C01':net_NTT5_C01, 'net_NTT5_D01':net_NTT5_D01, 
    'net_NTT5_D02A':net_NTT5_D02A, 'net_NTT5_D02B':net_NTT5_D02B, 'net_NTT5_E03':net_NTT5_E03, 'net_NTT5_E04':net_NTT5_E04, 'net_NTT5_F01':net_NTT5_F01, 'net_NTT5_F02':net_NTT5_F02, 'net_NTT5_F03':net_NTT5_F03, 
    'net_NTT5_G':net_NTT5_G, 'net_NTT5_G01':net_NTT5_G01, 'net_NTT5_G02':net_NTT5_G02, 'net_NTT5_G03':net_NTT5_G03, 'net_NTT5_H01':net_NTT5_H01, 'net_NTT5_H03':net_NTT5_H03, 'net_NTT5_I01':net_NTT5_I01, 
    'net_NTT5_I03':net_NTT5_I03, 'net_NTT5_J01':net_NTT5_J01, 'net_NTT5_J03':net_NTT5_J03, 'net_OP12_A':net_OP12_A, 'net_OP20_F1':net_OP20_F1, 'net_SG02_O':net_SG02_O, 'net_SG02_P':net_SG02_P, 'net_SG02_Q':net_SG02_Q, 
    'net_SG02_R':net_SG02_R, 'net_SG02_U':net_SG02_U,'net_GEG02_F':net_GEG02_F, 'dis_GEG02_F':dis_GEG02_F, 'GEG02_F':GEG02_F, 'GEG04_B3':GEG04_B3, 'GRTM_C503':GRTM_C503, 'dis_GEG04_B3':dis_GEG04_B3, 'dis_GRTM_C503':dis_GRTM_C503,
    'net_GEG04_B3':net_GEG04_B3, 'net_GRTM_C503':net_GRTM_C503,'grand_total':grand_total,'dis_grand_total':dis_grand_total,'net_grand_total':net_grand_total,'month':month})

def petron_report(request):
    
    date = datetime.datetime.today()
    report = Petron_report.objects.filter(StatementDate__year=date.year, StatementDate__month=date.month)
    ## ProductAmount ####
    BB14_B1 = Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B10").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B11= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B11").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B7= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B7").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_B8= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B8").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    BB14_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG12_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG12_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_B1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C5_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_C5_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_E1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_F2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG2_F3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG3-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_B15= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B15").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-F").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG4_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG5_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG5_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG6_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG6_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG7_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG7-F").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CMG8_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG8-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    CRA6_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CRA6-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C10").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN23_C9= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C9").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN9_C0201= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0201").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN9_C0301= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0301").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    FIN9_C0305= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0305").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_H= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-H").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_I= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-I").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_K= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-K").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_L= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-L").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG05_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG09_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GRTM-C503").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    ISG18_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="ISG18-C2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT11_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT11-E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-B6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_C5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D01_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D01-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D08_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D08-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D2_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D2_05= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-05").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-04").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D3_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-06").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D4_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D4-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D5_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D5_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D6_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6-06").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D03_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D03-B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D7_01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D7_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_D7_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3D").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_F3E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3E").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_G2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT2_G4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D2").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D4").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_D5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D5").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT3_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B3").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_A01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-A01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_B02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_B06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B06").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_C01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-C01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D02A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D02B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02B").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D2-03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_E01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_E03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_E04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E04").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_F01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_F02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_F03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G02").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_G03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_H01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_H03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_I01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_I03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_J01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J01").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT5_J03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J03").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    NTT6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT6-C").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    OP12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP12-A").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    OP20_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP20-F1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_O= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-O").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_P= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-P").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_Q= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-Q").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_R= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-R").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_U= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG02_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    SG02_T1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-T1").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('ProductAmount')).get('ProductAmount__sum', 0.00) or 0
    
    grand_total = BB14_B1+BB14_B10+BB14_B11+BB14_B2+BB14_B3+BB14_B4+BB14_B5
    +BB14_B6+BB14_B7+BB14_B8+BB14_C+BB14_E+CMG12_A+CMG12_C
    +CMG12_D+CMG2_B1+CMG2_B2+CMG2_B3+CMG2_C1+CMG2_C2+CMG2_C3
    +CMG2_C4+CMG2_C5_C+CMG2_C5_D+CMG2_D1+CMG2_D2+CMG2_D3+CMG2_E1
    +CMG2_E2+CMG2_E3+CMG2_F1+CMG2_F2+CMG2_F3+CMG3_A+CMG4_B
    +CMG4_B15+CMG4_B2+CMG4_E+CMG4_F+CMG4_G+CMG5_B+CMG5_E+CMG5_G
    +CMG6_B+CMG6_C+CMG6_E+CMG7_F+CMG8_B+CRA6_A+FIN23_C1+FIN23_C10
    +FIN23_C4+FIN23_C6+FIN23_C9+FIN9_C0201+FIN9_C0301+FIN9_C0305
    +GEG02_B+GEG02_C+GEG02_D+GEG02_E+GEG02_G+GEG02_H+GEG02_I+GEG02_K
    +GEG02_L+GEG04_B3+GEG05_E+GEG09_D+GRTM_C503+ISG18_C2+NTT11_E
    +NTT2_B6+NTT2_C3+NTT2_C5+NTT2_C6+NTT2_D01_A+NTT2_D08_C+NTT2_D1
    +NTT2_D2_02+NTT2_D2_03+NTT2_D2_05+NTT2_D3_02+NTT2_D3_03+NTT2_D3_04
    +NTT2_D3_06+NTT2_D4_03+NTT2_D5_02+NTT2_D5_03+NTT2_D6+NTT2_D6_06
    +NTT2_D03_B+NTT2_D7_01+NTT2_D7_02+NTT2_D7_03+NTT2_E2+NTT2_E3
    +NTT2_F3A+NTT2_F3B+NTT2_F3D+NTT2_F3E+NTT2_G2+NTT2_G4+NTT3_A
    +NTT3_B5+NTT3_C+NTT3_D2+NTT3_D3+NTT3_D4+NTT3_D5+NTT3_B3
    +NTT5_A01+NTT5_B02+NTT5_B06+NTT5_C01+NTT5_D01+NTT5_D02A
    +NTT5_D02B+NTT5_D2_03+NTT5_E01+NTT5_E03+NTT5_E04+NTT5_F01
    +NTT5_F02+NTT5_F03+NTT5_G+NTT5_G01+NTT5_G02+NTT5_G03+NTT5_H01
    +NTT5_H03+NTT5_I01+NTT5_I03+NTT5_J01+NTT5_J03+NTT6_C+OP12_A
    +OP20_F1+SG02_O+SG02_P+SG02_Q+SG02_R+SG02_U+GEG02_F+GRTM_C503
    +SG02_T1+GEG04_B3
    ## ProductAmount end ####

    ##Discount_Amount

    dis_BB14_B1 = Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B10").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B11= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B11").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B7= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B7").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_B8= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B8").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_BB14_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG12_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG12_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_B1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C5_C= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_C5_D= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_E1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_F2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG2_F3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG3-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_B15= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B15").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-F").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG4_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG5_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG5_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG6_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG6_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG7_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG7-F").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CMG8_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG8-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_CRA6_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CRA6-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C10").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN23_C9= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C9").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN9_C0201= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0201").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN9_C0301= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0301").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_FIN9_C0305= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0305").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_H= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-H").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_I= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-I").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_K= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-K").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_L= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-L").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG05_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG09_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GRTM-C503").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_ISG18_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="ISG18-C2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT11_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT11-E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-B6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_C5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D01_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D01-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D08_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D08-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D2_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D2_05= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-05").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-04").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D3_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-06").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D4_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D4-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D5_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D5_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D6_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6-06").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D03_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D03_B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D7_01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D7_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_D7_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3D").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_F3E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3E").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_G2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT2_G4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D2").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D3").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D4").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT3_D5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D5").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_A01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-A01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_B02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_B06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B06").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_C01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-C01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D02A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D02B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02B").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D2-03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_E01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_E03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_E04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E04").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_F01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_F02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_F03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G02").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_G03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_H01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_H03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_I01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_I03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_J01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J01").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT5_J03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J03").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_NTT6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT6-C").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_OP12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP12-A").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_OP20_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP20-F1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_O= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-O").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_P= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-P").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_Q= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-Q").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_R= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-R").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_U= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG02_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_SG02_T1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-T1").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    dis_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('DiscountAmount')).get('DiscountAmount__sum', 0.00) or 0
    
    dis_grand_total = dis_BB14_B1+dis_BB14_B10+dis_BB14_B11+dis_BB14_B2+dis_BB14_B3+dis_BB14_B4+dis_BB14_B5
    +dis_BB14_B6+dis_BB14_B7+dis_BB14_B8+dis_BB14_C+dis_BB14_E+dis_CMG12_A+dis_CMG12_C
    +dis_CMG12_D+dis_CMG2_B1+dis_CMG2_B2+dis_CMG2_B3+dis_CMG2_C1+dis_CMG2_C2+dis_CMG2_C3
    +dis_CMG2_C4+dis_CMG2_C5_C+dis_CMG2_C5_D+dis_CMG2_D1+dis_CMG2_D2+dis_CMG2_D3+dis_CMG2_E1
    +dis_CMG2_E2+dis_CMG2_E3+dis_CMG2_F1+dis_CMG2_F2+dis_CMG2_F3+dis_CMG3_A+dis_CMG4_B
    +dis_CMG4_B15+dis_CMG4_B2+dis_CMG4_E+dis_CMG4_F+dis_CMG4_G+dis_CMG5_B+dis_CMG5_E+dis_CMG5_G
    +dis_CMG6_B+dis_CMG6_C+dis_CMG6_E+dis_CMG7_F+dis_CMG8_B+dis_CRA6_A+dis_FIN23_C1+dis_FIN23_C10
    +dis_FIN23_C4+dis_FIN23_C6+dis_FIN23_C9+dis_FIN9_C0201+dis_FIN9_C0301+dis_FIN9_C0305
    +dis_GEG02_B+dis_GEG02_C+dis_GEG02_D+dis_GEG02_E+dis_GEG02_G+dis_GEG02_H+dis_GEG02_I+dis_GEG02_K
    +dis_GEG02_L+dis_GEG04_B3+dis_GEG05_E+dis_GEG09_D+dis_GRTM_C503+dis_ISG18_C2+dis_NTT11_E
    +dis_NTT2_B6+dis_NTT2_C3+dis_NTT2_C5+dis_NTT2_C6+dis_NTT2_D01_A+dis_NTT2_D08_C+dis_NTT2_D1
    +dis_NTT2_D2_02+dis_NTT2_D2_03+dis_NTT2_D2_05+dis_NTT2_D3_02+dis_NTT2_D3_03+dis_NTT2_D3_04
    +dis_NTT2_D3_06+dis_NTT2_D4_03+dis_NTT2_D5_02+dis_NTT2_D5_03+dis_NTT2_D6+dis_NTT2_D6_06
    +dis_NTT2_D03_B+dis_NTT2_D7_01+dis_NTT2_D7_02+dis_NTT2_D7_03+dis_NTT2_E2+dis_NTT2_E3
    +dis_NTT2_F3A+dis_NTT2_F3B+dis_NTT2_F3D+dis_NTT2_F3E+dis_NTT2_G2+dis_NTT2_G4+dis_NTT3_A
    +dis_NTT3_B5+dis_NTT3_C+dis_NTT3_D2+dis_NTT3_D3+dis_NTT3_D4+dis_NTT3_D5+dis_NTT3_B3
    +dis_NTT5_A01+dis_NTT5_B02+dis_NTT5_B06+dis_NTT5_C01+dis_NTT5_D01+dis_NTT5_D02A
    +dis_NTT5_D02B+dis_NTT5_D2_03+dis_NTT5_E01+dis_NTT5_E03+dis_NTT5_E04+dis_NTT5_F01
    +dis_NTT5_F02+dis_NTT5_F03+dis_NTT5_G+dis_NTT5_G01+dis_NTT5_G02+dis_NTT5_G03+dis_NTT5_H01
    +dis_NTT5_H03+dis_NTT5_I01+dis_NTT5_I03+dis_NTT5_J01+dis_NTT5_J03+dis_NTT6_C+dis_OP12_A
    +dis_OP20_F1+dis_SG02_O+dis_SG02_P+dis_SG02_Q+dis_SG02_R+dis_SG02_U+dis_GEG02_F+dis_GRTM_C503
    +dis_SG02_T1+dis_GEG04_B3
    ## Discount_Amount end ######

    ### Net Amount #####
    net_BB14_B1 = Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B10").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B11= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B11").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B7= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B7").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_B8= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-B8").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_BB14_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="BB14-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG12_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG12_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG12-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_B1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C5_D= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_C5_C= Petron_report.objects.filter(StatementDate__year=date.year, 
       StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-C5-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-D3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_E1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-E3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_F2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG2_F3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG2-F3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG3-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_B15= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B15").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_B2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-B2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-F").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG4_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG4-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG5_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG5_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG5-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG6_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG6_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG6-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG7_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG7-F").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CMG8_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CMG8-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_CRA6_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="CRA6-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C10= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C10").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN23_C9= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN23-C9").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN9_C0201= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0201").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN9_C0301= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0301").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_FIN9_C0305= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="FIN9-C0305").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_H= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-H").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_I= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-I").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_K= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-K").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_L= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-L").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG05_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG09_D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GEG02-D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="GRTM-C503").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_ISG18_C2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="ISG18-C2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT11_E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT11-E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_C3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_C5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_C6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-C6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_B6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-B6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D01_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D01-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D03_B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D03-B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D08_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D08-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D2_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D2_05= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D2-05").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-04").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D3_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D3-06").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D4_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D4-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D5_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D5_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D5-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D6= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D6_06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D6-06").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D7_01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D7_02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_D7_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-D7-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_E2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_E3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-E3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3D= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3D").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_F3E= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-F3E").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_G2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT2_G4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT2-G4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_B5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-B3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D2= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D2").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D3").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D4= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D4").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT3_D5= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT3-D5").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_A01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-A01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_B02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_B06= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-B06").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_C01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-C01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D02A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D02B= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D02B").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_D2_03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-D2-03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_E01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_E03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_E04= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-E04").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_F01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_F02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_F03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-F03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G02= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G02").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_G03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-G03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_H01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_H03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-H03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_I01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_I03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-I03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_J01= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J01").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT5_J03= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT5-J03").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_NTT6_C= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="NTT6_C").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_OP12_A= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP12-A").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_OP20_F1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="OP20-F1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_O= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-O").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_P= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-P").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_Q= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-Q").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_R= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-R").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_U= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG02_F= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GRTM_C503= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_SG02_T1= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-T1").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_GEG04_B3= Petron_report.objects.filter(StatementDate__year=date.year, 
        StatementDate__month=date.month, Supplier='Petron', ChargingDepartment="SG02-U").aggregate(Sum('NetAmount')).get('NetAmount__sum', 0.00) or 0
    net_grand_total = net_BB14_B1+net_BB14_B10+net_BB14_B11+net_BB14_B2+net_BB14_B3+net_BB14_B4+net_BB14_B5
    +net_BB14_B6+net_BB14_B7+net_BB14_B8+net_BB14_C+net_BB14_E+net_CMG12_A+net_CMG12_C
    +net_CMG12_D+net_CMG2_B1+net_CMG2_B2+net_CMG2_B3+net_CMG2_C1+net_CMG2_C2+net_CMG2_C3
    +net_CMG2_C4+net_CMG2_C5_C+net_CMG2_C5_D+net_CMG2_D1+net_CMG2_D2+net_CMG2_D3+net_CMG2_E1
    +net_CMG2_E2+net_CMG2_E3+net_CMG2_F1+net_CMG2_F2+net_CMG2_F3+net_CMG3_A+net_CMG4_B
    +net_CMG4_B15+net_CMG4_B2+net_CMG4_E+net_CMG4_F+net_CMG4_G+net_CMG5_B+net_CMG5_E+net_CMG5_G
    +net_CMG6_B+net_CMG6_C+net_CMG6_E+net_CMG7_F+net_CMG8_B+net_CRA6_A+net_FIN23_C1+net_FIN23_C10
    +net_FIN23_C4+net_FIN23_C6+net_FIN23_C9+net_FIN9_C0201+net_FIN9_C0301+net_FIN9_C0305
    +net_GEG02_B+net_GEG02_C+net_GEG02_D+net_GEG02_E+net_GEG02_G+net_GEG02_H+net_GEG02_I+net_GEG02_K
    +net_GEG02_L+net_GEG04_B3+net_GEG05_E+net_GEG09_D+net_GRTM_C503+net_ISG18_C2+net_NTT11_E
    +net_NTT2_B6+net_NTT2_C3+net_NTT2_C5+net_NTT2_C6+net_NTT2_D01_A+net_NTT2_D08_C+net_NTT2_D1
    +net_NTT2_D2_02+net_NTT2_D2_03+net_NTT2_D2_05+net_NTT2_D3_02+net_NTT2_D3_03+net_NTT2_D3_04
    +net_NTT2_D3_06+net_NTT2_D4_03+net_NTT2_D5_02+net_NTT2_D5_03+net_NTT2_D6+net_NTT2_D6_06
    +net_NTT2_D03_B+net_NTT2_D7_01+net_NTT2_D7_02+net_NTT2_D7_03+net_NTT2_E2+net_NTT2_E3
    +net_NTT2_F3A+net_NTT2_F3B+net_NTT2_F3D+net_NTT2_F3E+net_NTT2_G2+net_NTT2_G4+net_NTT3_A
    +net_NTT3_B5+net_NTT3_C+net_NTT3_D2+net_NTT3_D3+net_NTT3_D4+net_NTT3_D5+net_NTT3_B3
    +net_NTT5_A01+net_NTT5_B02+net_NTT5_B06+net_NTT5_C01+net_NTT5_D01+net_NTT5_D02A
    +net_NTT5_D02B+net_NTT5_D2_03+net_NTT5_E01+net_NTT5_E03+net_NTT5_E04+net_NTT5_F01
    +net_NTT5_F02+net_NTT5_F03+net_NTT5_G+net_NTT5_G01+net_NTT5_G02+net_NTT5_G03+net_NTT5_H01
    +net_NTT5_H03+net_NTT5_I01+net_NTT5_I03+net_NTT5_J01+net_NTT5_J03+net_NTT6_C+net_OP12_A
    +net_OP20_F1+net_SG02_O+net_SG02_P+net_SG02_Q+net_SG02_R+net_SG02_U+net_GEG02_F+net_GRTM_C503
    +net_SG02_T1+net_GEG04_B3

    from openpyxl.workbook import Workbook
    output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "Petron.xlsx"
    output['Content-Disposition'] = 'attachment; filename='+ file_name
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws['A1'].value = ""
    ws['A2'].value = ""
    ws.append(['Charging Department', 'Sum of Product Amount','Sum of Discount Amount','Sum of Net Amount'])
    ws['A4'].value = "BB14-B1"
    ws['A5'].value = "BB14-B10"
    ws['A6'].value = "BB14-B11"
    ws['A7'].value = "BB14-B2"
    ws['A8'].value = "BB14-B3"
    ws['A9'].value = "BB14-B4"
    ws['A10'].value = "BB14-B5"
    ws['A11'].value = "BB14-B6"
    ws['A12'].value = "BB14-B7"
    ws['A13'].value = "BB14-B8"
    ws['A14'].value = "BB14-C"
    ws['A15'].value = "BB14-E"
    ws['A16'].value = "CMG12-A"
    ws['A17'].value = "CMG12-C"
    ws['A18'].value = "CMG12-D"
    ws['A19'].value = "CMG2-B1"
    ws['A20'].value = "CMG2-B2"
    ws['A21'].value = "CMG2-B3"
    ws['A22'].value = "CMG2-C1"
    ws['A23'].value = "CMG2-C2"
    ws['A24'].value = "CMG2-C3"
    ws['A25'].value = "CMG2-C4"
    ws['A26'].value = "CMG2-C5-C"
    ws['A27'].value = "CMG2-C5-D"
    ws['A28'].value = "CMG2-D1"
    ws['A29'].value = "CMG2-D2"
    ws['A30'].value = "CMG2-D3"
    ws['A31'].value = "CMG2-E1"
    ws['A32'].value = "CMG2-E2"
    ws['A33'].value = "CMG2-E3"
    ws['A34'].value = "CMG2-F1"
    ws['A35'].value = "CMG2-F2"
    ws['A36'].value = "CMG2-F3"
    ws['A37'].value = "CMG3-A"
    ws['A38'].value = "CMG4-B"
    ws['A39'].value = "CMG4-B15"
    ws['A40'].value = "CMG4-B2"
    ws['A41'].value = "CMG4-E"
    ws['A42'].value = "CMG4-F"
    ws['A43'].value = "CMG4-G"
    ws['A44'].value = "CMG5-B"
    ws['A45'].value = "CMG5-E"
    ws['A46'].value = "CMG5-G"
    ws['A47'].value = "CMG6-B"
    ws['A48'].value = "CMG6-C"
    ws['A49'].value = "CMG6-E"
    ws['A50'].value = "CMG7-F"
    ws['A51'].value = "CMG8-B"
    ws['A52'].value = "CRA6-A"
    ws['A53'].value = "FIN23-C1"
    ws['A54'].value = "FIN23-C4"
    ws['A55'].value = "FIN23-C6"
    ws['A56'].value = "FIN9-C0201"
    ws['A57'].value = "FIN9-C0301"
    ws['A58'].value = "FIN9-C0305"
    ws['A59'].value = "GEG02-B"
    ws['A60'].value = "GEG02-C"
    ws['A61'].value = "GEG02-D"
    ws['A62'].value = "GEG02-E"
    ws['A63'].value = "GEG02-F"
    ws['A64'].value = "GEG02-G"
    ws['A65'].value = "GEG02-H"
    ws['A66'].value = "GEG02-I"
    ws['A67'].value = "GEG02-K"
    ws['A68'].value = "GEG02-L"
    ws['A69'].value = "GEG04-B3"
    ws['A70'].value = "GEG05-E"
    ws['A71'].value = "GEG09-D"
    ws['A72'].value = "GRTM-C503"
    ws['A73'].value = "ISG18-C2"
    ws['A74'].value = "NTT11-E"
    ws['A75'].value = "NTT2-B6"
    ws['A76'].value = "NTT2-C3"
    ws['A77'].value = "NTT2-C5"
    ws['A78'].value = "NTT2-C6"
    ws['A79'].value = "NTT2-D03-B"
    ws['A80'].value = "NTT2-D08-C"
    ws['A81'].value = "NTT2-D1"
    ws['A82'].value = "NTT2-D2-02"
    ws['A83'].value = "NTT2-D2-03"
    ws['A84'].value = "NTT2-D2-05"
    ws['A85'].value = "NTT2-D3-02"
    ws['A86'].value = "NTT2-D3-03"
    ws['A87'].value = "NTT2-D3-04"
    ws['A88'].value = "NTT2-D4-03"
    ws['A89'].value = "NTT2-D5-03"
    ws['A90'].value = "NTT2-D6"
    ws['A91'].value = "NTT2-D6-06"
    ws['A92'].value = "NTT2-D7-01"
    ws['A93'].value = "NTT2-D7-02"
    ws['A94'].value = "NTT2-D7-03"
    ws['A95'].value = "NTT2-E2"
    ws['A96'].value = "NTT2-E3"
    ws['A97'].value = "NTT2-F3A"
    ws['A98'].value = "NTT2-F3B"
    ws['A99'].value = "NTT2-F3D"
    ws['A100'].value = "NTT2-F3E"
    ws['A101'].value = "NTT2-G2"
    ws['A102'].value = "NTT2-G4"
    ws['A103'].value = "NTT3-B3"
    ws['A104'].value = "NTT3-B5"
    ws['A105'].value = "NTT3-C"
    ws['A106'].value = "NTT3-D2"
    ws['A107'].value = "NTT3-D3"
    ws['A108'].value = "NTT3-D4"
    ws['A109'].value = "NTT3-D5"
    ws['A110'].value = "NTT5-A01"
    ws['A111'].value = "NTT5-B02"
    ws['A112'].value = "NTT5-B06"
    ws['A113'].value = "NTT5-C01"
    ws['A114'].value = "NTT5-D01"
    ws['A115'].value = "NTT5-D02A"
    ws['A116'].value = "NTT5-D02B"
    ws['A117'].value = "NTT5-D2-03"
    ws['A118'].value = "NTT5-E01"
    ws['A119'].value = "NTT5-E03"
    ws['A120'].value = "NTT5-E04"
    ws['A121'].value = "NTT5-F01"
    ws['A122'].value = "NTT5-F02"
    ws['A123'].value = "NTT5-F03"
    ws['A124'].value = "NTT5-G"
    ws['A125'].value = "NTT5-G01"
    ws['A126'].value = "NTT5-G02"
    ws['A127'].value = "NTT5-G03"
    ws['A128'].value = "NTT5-H01"
    ws['A129'].value = "NTT5-H03"
    ws['A130'].value = "NTT5-I01"
    ws['A131'].value = "NTT5-I03"
    ws['A132'].value = "NTT5-J01"
    ws['A133'].value = "NTT5-J03"
    ws['A134'].value = "NTT6-C"
    ws['A135'].value = "OP12-A"
    ws['A136'].value = "OP20-F1"
    ws['A137'].value = "SG02-O"
    ws['A138'].value = "SG02-P"
    ws['A139'].value = "SG02-Q"
    ws['A140'].value = "SG02-R"
    ws['A141'].value = "SG02-T1"
    ws['A142'].value = "SG02-U"
    ws['A143'].value = "Grand Total"

    ###Sum of Product Amount ###
    ws['B4'].value = BB14_B1
    ws['B5'].value = BB14_B10
    ws['B6'].value = BB14_B11
    ws['B7'].value = BB14_B2
    ws['B8'].value = BB14_B3
    ws['B9'].value = BB14_B4
    ws['B10'].value = BB14_B5
    ws['B11'].value = BB14_B6
    ws['B12'].value = BB14_B7
    ws['B13'].value = BB14_B8
    ws['B14'].value = BB14_C
    ws['B15'].value = BB14_E
    ws['B16'].value = CMG12_A
    ws['B17'].value = CMG12_C
    ws['B18'].value = CMG12_D
    ws['B19'].value = CMG2_B1
    ws['B20'].value = CMG2_B2
    ws['B21'].value = CMG2_B3
    ws['B22'].value = CMG2_C1
    ws['B23'].value = CMG2_C2
    ws['B24'].value = CMG2_C3
    ws['B25'].value = CMG2_C4
    ws['B26'].value = CMG2_C5_C
    ws['B27'].value = CMG2_C5_D
    ws['B28'].value = CMG2_D1
    ws['B29'].value = CMG2_D2
    ws['B30'].value = CMG2_D3
    ws['B31'].value = CMG2_E1
    ws['B32'].value = CMG2_E2
    ws['B33'].value = CMG2_E3
    ws['B34'].value = CMG2_F1
    ws['B35'].value = CMG2_F2
    ws['B36'].value = CMG2_F3
    ws['B37'].value = CMG3_A
    ws['B38'].value = CMG4_B
    ws['B39'].value = CMG4_B15
    ws['B40'].value = CMG4_B2
    ws['B41'].value = CMG4_E
    ws['B42'].value = CMG4_F
    ws['B43'].value = CMG4_G
    ws['B44'].value = CMG5_B
    ws['B45'].value = CMG5_E
    ws['B46'].value = CMG5_G
    ws['B47'].value = CMG6_B
    ws['B48'].value = CMG6_C
    ws['B49'].value = CMG6_E
    ws['B50'].value = CMG7_F
    ws['B51'].value = CMG8_B
    ws['B52'].value = CRA6_A
    ws['B53'].value = FIN23_C1
    ws['B54'].value = FIN23_C4
    ws['B55'].value = FIN23_C6
    ws['B56'].value = FIN9_C0201
    ws['B57'].value = FIN9_C0301
    ws['B58'].value = FIN9_C0305
    ws['B59'].value = GEG02_B
    ws['B60'].value = GEG02_C
    ws['B61'].value = GEG02_D
    ws['B62'].value = GEG02_E
    ws['B63'].value = GEG02_F
    ws['B64'].value = GEG02_G
    ws['B65'].value = GEG02_H
    ws['B66'].value = GEG02_I
    ws['B67'].value = GEG02_K
    ws['B68'].value = GEG02_L
    ws['B69'].value = GEG04_B3
    ws['B70'].value = GEG05_E
    ws['B71'].value = GEG09_D
    ws['B72'].value = GRTM_C503
    ws['B73'].value = ISG18_C2
    ws['B74'].value = NTT11_E
    ws['B75'].value = NTT2_B6
    ws['B76'].value = NTT2_C3
    ws['B77'].value = NTT2_C5
    ws['B78'].value = NTT2_C6
    ws['B79'].value = NTT2_D03_B
    ws['B80'].value = NTT2_D08_C
    ws['B81'].value = NTT2_D1
    ws['B82'].value = NTT2_D2_02
    ws['B83'].value = NTT2_D2_03
    ws['B84'].value = NTT2_D2_05
    ws['B85'].value = NTT2_D3_02
    ws['B86'].value = NTT2_D3_03
    ws['B87'].value = NTT2_D3_04
    ws['B88'].value = NTT2_D4_03
    ws['B89'].value = NTT2_D5_03
    ws['B90'].value = NTT2_D6
    ws['B91'].value = NTT2_D6_06
    ws['B92'].value = NTT2_D7_01
    ws['B93'].value = NTT2_D7_02
    ws['B94'].value = NTT2_D7_03
    ws['B95'].value = NTT2_E2
    ws['B96'].value = NTT2_E3
    ws['B97'].value = NTT2_F3A
    ws['B98'].value = NTT2_F3B
    ws['B99'].value = NTT2_F3D
    ws['B100'].value = NTT2_F3E
    ws['B101'].value = NTT2_G2
    ws['B102'].value = NTT2_G4
    ws['B103'].value = NTT3_B3
    ws['B104'].value = NTT3_B5
    ws['B105'].value = NTT3_C
    ws['B106'].value = NTT3_D2
    ws['B107'].value = NTT3_D3
    ws['B108'].value = NTT3_D4
    ws['B109'].value = NTT3_D5
    ws['B110'].value = NTT5_A01
    ws['B111'].value = NTT5_B02
    ws['B112'].value = NTT5_B06
    ws['B113'].value = NTT5_C01
    ws['B114'].value = NTT5_D01
    ws['B115'].value = NTT5_D02A
    ws['B116'].value = NTT5_D02B
    ws['B117'].value = NTT5_D2_03
    ws['B118'].value = NTT5_E01
    ws['B119'].value = NTT5_E03
    ws['B120'].value = NTT5_E04
    ws['B121'].value = NTT5_F01
    ws['B122'].value = NTT5_F02
    ws['B123'].value = NTT5_F03
    ws['B124'].value = NTT5_G
    ws['B125'].value = NTT5_G01
    ws['B126'].value = NTT5_G02
    ws['B127'].value = NTT5_G03
    ws['B128'].value = NTT5_H01
    ws['B129'].value = NTT5_H03
    ws['B130'].value = NTT5_I01
    ws['B131'].value = NTT5_I03
    ws['B132'].value = NTT5_J01
    ws['B133'].value = NTT5_J03
    ws['B134'].value = NTT6_C
    ws['B135'].value = OP12_A
    ws['B136'].value = OP20_F1
    ws['B137'].value = SG02_O
    ws['B138'].value = SG02_P
    ws['B139'].value = SG02_Q
    ws['B140'].value = SG02_R
    ws['B141'].value = SG02_T1
    ws['B142'].value = SG02_U
    ws['B143'].value = grand_total

    #####Sum of Discount####
    ws['C4'].value = dis_BB14_B1
    ws['C5'].value = dis_BB14_B10
    ws['C6'].value = dis_BB14_B11
    ws['C7'].value = dis_BB14_B2
    ws['C8'].value = dis_BB14_B3
    ws['C9'].value = dis_BB14_B4
    ws['C10'].value = dis_BB14_B5
    ws['C11'].value = dis_BB14_B6
    ws['C12'].value = dis_BB14_B7
    ws['C13'].value = dis_BB14_B8
    ws['C14'].value = dis_BB14_C
    ws['C15'].value = dis_BB14_E
    ws['C16'].value = dis_CMG12_A
    ws['C17'].value = dis_CMG12_C
    ws['C18'].value = dis_CMG12_D
    ws['C19'].value = dis_CMG2_B1
    ws['C20'].value = dis_CMG2_B2
    ws['C21'].value = dis_CMG2_B3
    ws['C22'].value = dis_CMG2_C1
    ws['C23'].value = dis_CMG2_C2
    ws['C24'].value = dis_CMG2_C3
    ws['C25'].value = dis_CMG2_C4
    ws['C26'].value = dis_CMG2_C5_C
    ws['C27'].value = dis_CMG2_C5_D
    ws['C28'].value = dis_CMG2_D1
    ws['C29'].value = dis_CMG2_D2
    ws['C30'].value = dis_CMG2_D3
    ws['C31'].value = dis_CMG2_E1
    ws['C32'].value = dis_CMG2_E2
    ws['C33'].value = dis_CMG2_E3
    ws['C34'].value = dis_CMG2_F1
    ws['C35'].value = dis_CMG2_F2
    ws['C36'].value = dis_CMG2_F3
    ws['C37'].value = dis_CMG3_A
    ws['C38'].value = dis_CMG4_B
    ws['C39'].value = dis_CMG4_B15
    ws['C40'].value = dis_CMG4_B2
    ws['C41'].value = dis_CMG4_E
    ws['C42'].value = dis_CMG4_F
    ws['C43'].value = dis_CMG4_G
    ws['C44'].value = dis_CMG5_B
    ws['C45'].value = dis_CMG5_E
    ws['C46'].value = dis_CMG5_G
    ws['C47'].value = dis_CMG6_B
    ws['C48'].value = dis_CMG6_C
    ws['C49'].value = dis_CMG6_E
    ws['C50'].value = dis_CMG7_F
    ws['C51'].value = dis_CMG8_B
    ws['C52'].value = dis_CRA6_A
    ws['C53'].value = dis_FIN23_C1
    ws['C54'].value = dis_FIN23_C4
    ws['C55'].value = dis_FIN23_C6
    ws['C56'].value = dis_FIN9_C0201
    ws['C57'].value = dis_FIN9_C0301
    ws['C58'].value = dis_FIN9_C0305
    ws['C59'].value = dis_GEG02_B
    ws['C60'].value = dis_GEG02_C
    ws['C61'].value = dis_GEG02_D
    ws['C62'].value = dis_GEG02_E
    ws['C63'].value = dis_GEG02_F
    ws['C64'].value = dis_GEG02_G
    ws['C65'].value = dis_GEG02_H
    ws['C66'].value = dis_GEG02_I
    ws['C67'].value = dis_GEG02_K
    ws['C68'].value = dis_GEG02_L
    ws['C69'].value = dis_GEG04_B3
    ws['C70'].value = dis_GEG05_E
    ws['C71'].value = dis_GEG09_D
    ws['C72'].value = dis_GRTM_C503
    ws['C73'].value = dis_ISG18_C2
    ws['C74'].value = dis_NTT11_E
    ws['C75'].value = dis_NTT2_B6
    ws['C76'].value = dis_NTT2_C3
    ws['C77'].value = dis_NTT2_C5
    ws['C78'].value = dis_NTT2_C6
    ws['C79'].value = dis_NTT2_D03_B
    ws['C80'].value = dis_NTT2_D08_C
    ws['C81'].value = dis_NTT2_D1
    ws['C82'].value = dis_NTT2_D2_02
    ws['C83'].value = dis_NTT2_D2_03
    ws['C84'].value = dis_NTT2_D2_05
    ws['C85'].value = dis_NTT2_D3_02
    ws['C86'].value = dis_NTT2_D3_03
    ws['C87'].value = dis_NTT2_D3_04
    ws['C88'].value = dis_NTT2_D4_03
    ws['C89'].value = dis_NTT2_D5_03
    ws['C90'].value = dis_NTT2_D6
    ws['C91'].value = dis_NTT2_D6_06
    ws['C92'].value = dis_NTT2_D7_01
    ws['C93'].value = dis_NTT2_D7_02
    ws['C94'].value = dis_NTT2_D7_03
    ws['C95'].value = dis_NTT2_E2
    ws['C96'].value = dis_NTT2_E3
    ws['C97'].value = dis_NTT2_F3A
    ws['C98'].value = dis_NTT2_F3B
    ws['C99'].value = dis_NTT2_F3D
    ws['C100'].value = dis_NTT2_F3E
    ws['C101'].value = dis_NTT2_G2
    ws['C102'].value = dis_NTT2_G4
    ws['C103'].value = dis_NTT3_B3
    ws['C104'].value = dis_NTT3_B5
    ws['C105'].value = dis_NTT3_C
    ws['C106'].value = dis_NTT3_D2
    ws['C107'].value = dis_NTT3_D3
    ws['C108'].value = dis_NTT3_D4
    ws['C109'].value = dis_NTT3_D5
    ws['C110'].value = dis_NTT5_A01
    ws['C111'].value = dis_NTT5_B02
    ws['C112'].value = dis_NTT5_B06
    ws['C113'].value = dis_NTT5_C01
    ws['C114'].value = dis_NTT5_D01
    ws['C115'].value = dis_NTT5_D02A
    ws['C116'].value = dis_NTT5_D02B
    ws['C117'].value = dis_NTT5_D2_03
    ws['C118'].value = dis_NTT5_E01
    ws['C119'].value = dis_NTT5_E03
    ws['C120'].value = dis_NTT5_E04
    ws['C121'].value = dis_NTT5_F01
    ws['C122'].value = dis_NTT5_F02
    ws['C123'].value = dis_NTT5_F03
    ws['C124'].value = dis_NTT5_G
    ws['C125'].value = dis_NTT5_G01
    ws['C126'].value = dis_NTT5_G02
    ws['C127'].value = dis_NTT5_G03
    ws['C128'].value = dis_NTT5_H01
    ws['C129'].value = dis_NTT5_H03
    ws['C130'].value = dis_NTT5_I01
    ws['C131'].value = dis_NTT5_I03
    ws['C132'].value = dis_NTT5_J01
    ws['C133'].value = dis_NTT5_J03
    ws['C134'].value = dis_NTT6_C
    ws['C135'].value = dis_OP12_A
    ws['C136'].value = dis_OP20_F1
    ws['C137'].value = dis_SG02_O
    ws['C138'].value = dis_SG02_P
    ws['C139'].value = dis_SG02_Q
    ws['C140'].value = dis_SG02_R
    ws['C141'].value = dis_SG02_T1
    ws['C142'].value = dis_SG02_U
    ws['C143'].value = dis_grand_total

    ##### net amount #####
    ws['D4'].value = net_BB14_B1
    ws['D5'].value = net_BB14_B10
    ws['D6'].value = net_BB14_B11
    ws['D7'].value = net_BB14_B2
    ws['D8'].value = net_BB14_B3
    ws['D9'].value = net_BB14_B4
    ws['D10'].value = net_BB14_B5
    ws['D11'].value = net_BB14_B6
    ws['D12'].value = net_BB14_B7
    ws['D13'].value = net_BB14_B8
    ws['D14'].value = net_BB14_C
    ws['D15'].value = net_BB14_E
    ws['D16'].value = net_CMG12_A
    ws['D17'].value = net_CMG12_C
    ws['D18'].value = net_CMG12_D
    ws['D19'].value = net_CMG2_B1
    ws['D20'].value = net_CMG2_B2
    ws['D21'].value = net_CMG2_B3
    ws['D22'].value = net_CMG2_C1
    ws['D23'].value = net_CMG2_C2
    ws['D24'].value = net_CMG2_C3
    ws['D25'].value = net_CMG2_C4
    ws['D26'].value = net_CMG2_C5_C
    ws['D27'].value = net_CMG2_C5_D
    ws['D28'].value = net_CMG2_D1
    ws['D29'].value = net_CMG2_D2
    ws['D30'].value = net_CMG2_D3
    ws['D31'].value = net_CMG2_E1
    ws['D32'].value = net_CMG2_E2
    ws['D33'].value = net_CMG2_E3
    ws['D34'].value = net_CMG2_F1
    ws['D35'].value = net_CMG2_F2
    ws['D36'].value = net_CMG2_F3
    ws['D37'].value = net_CMG3_A
    ws['D38'].value = net_CMG4_B
    ws['D39'].value = net_CMG4_B15
    ws['D40'].value = net_CMG4_B2
    ws['D41'].value = net_CMG4_E
    ws['D42'].value = net_CMG4_F
    ws['D43'].value = net_CMG4_G
    ws['D44'].value = net_CMG5_B
    ws['D45'].value = net_CMG5_E
    ws['D46'].value = net_CMG5_G
    ws['D47'].value = net_CMG6_B
    ws['D48'].value = net_CMG6_C
    ws['D49'].value = net_CMG6_E
    ws['D50'].value = net_CMG7_F
    ws['D51'].value = net_CMG8_B
    ws['D52'].value = net_CRA6_A
    ws['D53'].value = net_FIN23_C1
    ws['D54'].value = net_FIN23_C4
    ws['D55'].value = net_FIN23_C6
    ws['D56'].value = net_FIN9_C0201
    ws['D57'].value = net_FIN9_C0301
    ws['D58'].value = net_FIN9_C0305
    ws['D59'].value = net_GEG02_B
    ws['D60'].value = net_GEG02_C
    ws['D61'].value = net_GEG02_D
    ws['D62'].value = net_GEG02_E
    ws['D63'].value = net_GEG02_F
    ws['D64'].value = net_GEG02_G
    ws['D65'].value = net_GEG02_H
    ws['D66'].value = net_GEG02_I
    ws['D67'].value = net_GEG02_K
    ws['D68'].value = net_GEG02_L
    ws['D69'].value = net_GEG04_B3
    ws['D70'].value = net_GEG05_E
    ws['D71'].value = net_GEG09_D
    ws['D72'].value = net_GRTM_C503
    ws['D73'].value = net_ISG18_C2
    ws['D74'].value = net_NTT11_E
    ws['D75'].value = net_NTT2_B6
    ws['D76'].value = net_NTT2_C3
    ws['D77'].value = net_NTT2_C5
    ws['D78'].value = net_NTT2_C6
    ws['D79'].value = net_NTT2_D03_B
    ws['D80'].value = net_NTT2_D08_C
    ws['D81'].value = net_NTT2_D1
    ws['D82'].value = net_NTT2_D2_02
    ws['D83'].value = net_NTT2_D2_03
    ws['D84'].value = net_NTT2_D2_05
    ws['D85'].value = net_NTT2_D3_02
    ws['D86'].value = net_NTT2_D3_03
    ws['D87'].value = net_NTT2_D3_04
    ws['D88'].value = net_NTT2_D4_03
    ws['D89'].value = net_NTT2_D5_03
    ws['D90'].value = net_NTT2_D6
    ws['D91'].value = net_NTT2_D6_06
    ws['D92'].value = net_NTT2_D7_01
    ws['D93'].value = net_NTT2_D7_02
    ws['D94'].value = net_NTT2_D7_03
    ws['D95'].value = net_NTT2_E2
    ws['D96'].value = net_NTT2_E3
    ws['D97'].value = net_NTT2_F3A
    ws['D98'].value = net_NTT2_F3B
    ws['D99'].value = net_NTT2_F3D
    ws['D100'].value = net_NTT2_F3E
    ws['D101'].value = net_NTT2_G2
    ws['D102'].value = net_NTT2_G4
    ws['D103'].value = net_NTT3_B3
    ws['D104'].value = net_NTT3_B5
    ws['D105'].value = net_NTT3_C
    ws['D106'].value = net_NTT3_D2
    ws['D107'].value = net_NTT3_D3
    ws['D108'].value = net_NTT3_D4
    ws['D109'].value = net_NTT3_D5
    ws['D110'].value = net_NTT5_A01
    ws['D111'].value = net_NTT5_B02
    ws['D112'].value = net_NTT5_B06
    ws['D113'].value = net_NTT5_C01
    ws['D114'].value = net_NTT5_D01
    ws['D115'].value = net_NTT5_D02A
    ws['D116'].value = net_NTT5_D02B
    ws['D117'].value = net_NTT5_D2_03
    ws['D118'].value = net_NTT5_E01
    ws['D119'].value = net_NTT5_E03
    ws['D120'].value = net_NTT5_E04
    ws['D121'].value = net_NTT5_F01
    ws['D122'].value = net_NTT5_F02
    ws['D123'].value = net_NTT5_F03
    ws['D124'].value = net_NTT5_G
    ws['D125'].value = net_NTT5_G01
    ws['D126'].value = net_NTT5_G02
    ws['D127'].value = net_NTT5_G03
    ws['D128'].value = net_NTT5_H01
    ws['D129'].value = net_NTT5_H03
    ws['D130'].value = net_NTT5_I01
    ws['D131'].value = net_NTT5_I03
    ws['D132'].value = net_NTT5_J01
    ws['D133'].value = net_NTT5_J03
    ws['D134'].value = net_NTT6_C
    ws['D135'].value = net_OP12_A
    ws['D136'].value = net_OP20_F1
    ws['D137'].value = net_SG02_O
    ws['D138'].value = net_SG02_P
    ws['D139'].value = net_SG02_Q
    ws['D140'].value = net_SG02_R
    ws['D141'].value = net_SG02_T1
    ws['D142'].value = net_SG02_U
    ws['D143'].value = net_grand_total

    
    ws1 = wb.create_sheet("Raw Data")
    ws1.title = "Raw"
    columns = [
            'Record Number',
            'Statement No',
            'Account Number',
            'Statement Date',
            'Payment Due Date',
            'Period Covered',
            'Card Number',
            'Charging Department',
            'Embossed Name',
            'Plate Number',
            'Vehicle Description',
            'Invoice Date',
            'Station Name',
            'Station Address',
            'Invoice Number',
            'Product Name',
            'Product Quantity',
            'Product Amount',
            'Discount Per Litre',
            'Discount Amount',
            'Net Amount',
            'Odometer',
            'KM Driven',
            'Php/Km',
            'Km/Li',
            'Fuel Limit',
            'Fuel Limit Unit',

    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws1.cell(row=row_num, column=col_num)
        cell.value = column_title

    for report in report:
        row_num += 1
        row = [
            report.RecordNumber,
            report.StatementNo,
            report.Account_Number,
            report.StatementDate,
            report.PaymentDueDate,
            report.PeriodCovered,
            report.CardNumber,
            report.ChargingDepartment,
            report.EmbossedName,
            report.PlateNumber,
            report.VehicleDescription,
            report.InvoiceDate,
            report.StationName,
            report.StationAddress,
            report.InvoiceNumber,
            report.ProductName,
            report.ProductQuantity,
            report.ProductAmount,
            report.DiscountPerLitre,
            report.DiscountAmount,
            report.NetAmount,
            report.Odometer,
            report.KMDriven,
            report.Php_Km,
            report.Km_Li,
            report.FuelLimit,
            report.FuelLimitUnit,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = cell_value


    wb.save(output)

    return output
