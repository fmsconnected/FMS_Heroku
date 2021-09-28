from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from django.views import generic
import schedule
import time
from django.core.mail import get_connection,send_mail
from django.core import mail
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.template import Context
from django.shortcuts import render,HttpResponseRedirect, get_list_or_404,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
import datetime
from datetime import date, timedelta
from django.db.models import Q, Count
from .models import (
		CarRentalRequest
)
from masterlist.models import (
    EmployeeMasterlist,
    )
from vehicle_masterlist.models import VehicleMasterList
from leasingmasterlist.models import Leasing

from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from CustomerLog.models import (
    CS_log
)
from . forms import (
    carrequestform,

    )

from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )


                      #####################################  
                    #########################################
                   #####.       Car Rental Request       #####
                    #########################################
                     ######################################




def requestCreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    return render(request, 'car_rental/carrequest_new.html',{'Title':'Car - Car Request', 'emplist':emplist})

def requestsubmit(request):
    if request.method == 'POST':
        emp_id = request.POST.get('emp_id')
        date_received = request.POST.get('date_received')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        cnumber = request.POST.get('cnumber')
        company = request.POST.get('company')
        band = request.POST.get('band')
        dept = request.POST.get('dept')
        cost = request.POST.get('cost')
        div = request.POST.get('div')
        loc = request.POST.get('loc')
        section = request.POST.get('section')
        designation = request.POST.get('designation')
        atd = request.POST.get('atd')
        vname = request.POST.get('vname')
        date = request.POST.get('date')
        up_to = request.POST.get('up_to')
        time = request.POST.get('time')
        del_place = request.POST.get('del_place')
        type_rental = request.POST.get('type_rental')
        costcenter = request.POST.get('costcenter')
        rent_period = request.POST.get('rent_period')
        destination = request.POST.get('destination')
        del_date = request.POST.get('del_date')
        end_user = request.POST.get('end_user')
        vehicle_type = request.POST.get('vehicle_type')
        plate_no = request.POST.get('plate_no')
        supervisor = request.POST.get('supervisor')
        cr_sla = request.POST.get('cr_sla')
        

        saveto_req = CarRentalRequest(A_Employee = emp_id, Date_received = date_received,Assignee_Fname = fname,Assignee_Lname = lname,Assignee_No = cnumber,Assignee_Company = company,
                Assignee_band = band,Assignee_Dept = dept,Assignee_Cost = cost,Assignee_Div = div,Assignee_Loc = loc,Assignee_Section = section,
                Assignee_Designation = designation,Assignee_ATD = atd,Vendor_name = vname,Date = date,Up_to = up_to,Time = time,Place_of_del = del_place,
                type_rental=type_rental, Cost_center=costcenter, Rental_period=rent_period,Destination = destination,Delivery_date = del_date,End_user = end_user,Type_of_vehicle = vehicle_type,
                Plate_no = plate_no,Immediate_supervisor =supervisor ,CR_SLA = cr_sla)
        saveto_req.save()

        return HttpResponseRedirect('/CarRequest/Request/')

class requestListView(ListView):
	model = CarRentalRequest
	template_name = 'car_rental/carrequest_list.html'

class requestDetailView(DetailView):
	model = CarRentalRequest
	template_name = 'car_rental/carrequest_details.html'

class requestUpdateView(SuccessMessageMixin, UpdateView):
    model = CarRentalRequest
    form_class = carrequestform
    template_name = 'car_rental/carrequest_form.html'
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Car Rental Request Updated Successfully!"

class requestDeleteView(BSModalDeleteView):
    model = CarRentalRequest
    template_name = 'car_rental/car_delete.html'
    success_message = 'Success: Report was deleted.'
    success_url = reverse_lazy('carrequest_list')

def requestHistoryView(request):
    if request.method == "GET":
       obj = CarRentalRequest.history.all()

       return render(request, 'car_rental/carrequest_history.html', context={'object': obj})
def crr_deadline(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    # dl = datetime.datetime.today() - timedelta(days=3)
    dl = CarRentalRequest.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=1))
    dl2 = CarRentalRequest.objects.filter(Deadline__date = datetime.datetime.today() + timedelta(days=2))
    dl3 = CarRentalRequest.objects.filter(Deadline__date = datetime.datetime.today())
    return  render(request, 'car_rental/crrdeadline.html',{'title':'Car Request - Car Request Deadline', 'dl':dl, 'dl2':dl2, 'dl3':dl3})


def car_request_excel(request):
    rental_queryset = CarRentalRequest.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Car Rental Request.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Car Rental Request'

    columns = [
                'Assignee Employee Id' ,
                'Date Received' ,
                'Assignee First name' ,
                'Assignee Last name' ,
                'Assignee No' ,
                'Assignee Company' ,
                'Assignee Band' ,
                'Assignee Department' ,
                'Assignee Cost Center' ,
                'Assignee Division' ,
                'Assignee Loccation' ,
                'Assignee Section' ,
                'Assignee Designation' ,
                'Assignee ATD' ,
                'Vendor Name' ,
                'Date' ,
                'Up to' ,
                'Time' ,
                'Place of Delivery' ,
                'Type of Rental' ,
                'Cost Center' ,
                'Rental Reriod' ,
                'Destination' ,
                'Delivery Date' ,
                'End User' ,
                'Type of Vehicle' ,
                'Plate no' ,
                'Immediate Supervisor' ,
                'SLA' ,
                'Date Initiated' ,
                'Deadline',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for car in rental_queryset:
        row_num += 1
        row = [
                car.A_Employee ,
                car.Date_received ,
                car.Assignee_Fname ,
                car.Assignee_Lname ,
                car.Assignee_No ,
                car.Assignee_Company ,
                car.Assignee_band ,
                car.Assignee_Dept ,
                car.Assignee_Cost ,
                car.Assignee_Div ,
                car.Assignee_Loc ,
                car.Assignee_Section ,
                car.Assignee_Designation ,
                car.Assignee_ATD ,
                car.Vendor_name ,
                car.Date ,
                car.Up_to ,
                car.Time ,
                car.Place_of_del ,
                car.type_rental ,
                car.Cost_center ,
                car.Rental_period ,
                car.Destination ,
                car.Delivery_date ,
                car.End_user ,
                car.Type_of_vehicle ,
                car.Plate_no ,
                car.Immediate_supervisor ,
                car.CR_SLA ,
                car.Date_initiated ,
                car.Deadline,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
